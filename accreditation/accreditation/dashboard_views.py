"""
Dashboard views for PLP Accreditation System
Unified dashboard with role-based access control
Updated: 2025-11-03 - Fixed list/dict handling
"""

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from accreditation.decorators import login_required
from accreditation.firebase_utils import (
    get_all_documents, 
    get_document, 
    create_document, 
    update_document, 
    delete_document,
    query_documents
)
from accreditation.audit_utils import log_audit, get_client_ip
from accreditation.forms import UserManagementForm
from accreditation.firebase_auth import FirebaseUser
from accreditation.password_generator import generate_strong_password
from accreditation.cloudinary_utils import upload_image_to_cloudinary, delete_image_from_cloudinary
from accreditation.document_validator import validate_document_header
import json
import hashlib
import secrets
import traceback
from datetime import datetime


def get_session_user_dict(request):
    """
    Safely get user dict from session, ensuring it's always a dict.
    Prevents 'list' object has no attribute 'get' errors.
    """
    session_user = request.session.get('user', {})
    if not isinstance(session_user, dict):
        return {}
    return session_user


def safe_get_document(collection, doc_id, request=None):
    """
    Safely get a document from Firestore, ensuring it's always a dict.
    Handles cases where get_document returns a list instead of dict.
    
    Args:
        collection: Firestore collection name
        doc_id: Document ID
        request: Optional request object for caching
        
    Returns:
        dict: Document data or empty dict if not found
    """
    doc = get_document(collection, doc_id, request=request)
    
    # Handle list response
    if isinstance(doc, list):
        if len(doc) > 0:
            return doc[0] if isinstance(doc[0], dict) else {}
        return {}
    
    # Handle dict response
    if isinstance(doc, dict):
        return doc
    
    # Handle None or other types
    return {}


def hash_password(raw_password):
    """Hash a password using PBKDF2"""
    salt = secrets.token_hex(16)
    password_hash = hashlib.pbkdf2_hmac('sha256', 
                                      raw_password.encode('utf-8'),
                                      salt.encode('utf-8'),
                                      100000)
    return f"{salt}${password_hash.hex()}"


def get_user_from_session(request):
    """Build user object from session data"""
    # Get user data from session, ensuring it's a dict
    session_user = get_session_user_dict(request)
    
    return {
        'id': request.session.get('user_id'),
        'email': request.session.get('user_email'),
        'role': request.session.get('user_role'),
        'name': request.session.get('user_name'),
        'first_name': session_user.get('first_name', ''),
        'middle_name': session_user.get('middle_name', ''),
        'last_name': session_user.get('last_name', ''),
        'profile_image_url': session_user.get('profile_image_url', ''),
    }


@login_required
def dashboard_home(request):
    """
    Main dashboard home page with role-based content
    Shows stats cards, charts, and recent activity based on user role
    """
    user = get_user_from_session(request)
    user_role = user.get('role', '')
    
    context = {
        'active_page': 'dashboard',
        'user': user,
    }
    
    # Get role-specific dashboard data
    if user_role == 'qa_head' or user_role == 'qa_admin':
        # Both QA roles use the same dashboard data
        context.update(get_qa_admin_dashboard_data(user))
    elif user_role == 'department_user':
        context.update(get_department_dashboard_data(user))
    
    return render(request, 'dashboard/home.html', context)


@login_required
def calendar_view(request):
    """Calendar page"""
    context = {
        'active_page': 'calendar',
        'user': get_user_from_session(request),
    }
    return render(request, 'dashboard/calendar.html', context)


@login_required
def accreditation_view(request):
    """Accreditation page - displays all active items"""
    user = get_user_from_session(request)
    
    # Fetch all active departments
    try:
        departments = get_all_documents('departments')
        
        # Filter only active and non-archived departments
        departments = [
            dept for dept in departments 
            if dept.get('is_active', True) and not dept.get('is_archived', False)
        ]
        
        for dept in departments:
            # Set default values if not present
            if 'logo_url' not in dept:
                dept['logo_url'] = ''
        
        # Sort by name
        departments.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching departments: {str(e)}")
        departments = []
    
    # Fetch all active programs
    try:
        programs = get_all_documents('programs')
        
        # Filter only active and non-archived programs
        programs = [
            prog for prog in programs 
            if prog.get('is_active', True) and not prog.get('is_archived', False)
        ]
        
        # Sort by code
        programs.sort(key=lambda x: x.get('code', ''))
        
    except Exception as e:
        print(f"Error fetching programs: {str(e)}")
        programs = []
    
    # Fetch all active accreditation types
    try:
        types = get_all_documents('accreditation_types')
        
        # Filter only active and non-archived types
        types = [
            t for t in types 
            if t.get('is_active', True) and not t.get('is_archived', False)
        ]
        
        # Sort by type name
        types.sort(key=lambda x: x.get('type', ''))
        
    except Exception as e:
        print(f"Error fetching types: {str(e)}")
        types = []
    
    # Fetch all active areas
    try:
        areas = get_all_documents('areas')
        
        # Filter only active and non-archived areas
        areas = [
            mod for mod in areas 
            if mod.get('is_active', True) and not mod.get('is_archived', False)
        ]
        
        # Sort by name
        areas.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching areas: {str(e)}")
        areas = []
    
    # Calculate progress for each department based on its programs
    try:
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for dept in departments:
            dept_id = dept.get('id')
            # Get all programs for this department
            dept_programs = [p for p in programs if p.get('department_id') == dept_id]
            
            if not dept_programs:
                dept['progress'] = 0
                continue
            
            # Calculate progress for each program first
            program_progresses = []
            for prog in dept_programs:
                prog_id = prog.get('id')
                prog_types = [t for t in types if t.get('program_id') == prog_id]
                
                if not prog_types:
                    program_progresses.append(0)
                    continue
                
                # Calculate progress for each type
                type_progresses = []
                for prog_type in prog_types:
                    type_id = prog_type.get('id')
                    type_areas = [a for a in areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                    
                    if not type_areas:
                        type_progresses.append(0)
                        continue
                    
                    # Calculate progress for each area
                    area_progresses = []
                    for area in type_areas:
                        area_id = area.get('id')
                        area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
                        
                        if not area_checklists:
                            area_progresses.append(0)
                            continue
                        
                        total_checklists = len(area_checklists)
                        completed_checklists = 0
                        
                        for checklist in area_checklists:
                            checklist_id = checklist.get('id')
                            required_docs = [
                                doc for doc in all_documents 
                                if doc.get('checklist_id') == checklist_id 
                                and doc.get('is_required', False)
                                and not doc.get('is_archived', False)
                                and doc.get('status') == 'approved'
                            ]
                            if len(required_docs) > 0:
                                completed_checklists += 1
                        
                        area_progress = (completed_checklists / total_checklists) * 100 if total_checklists > 0 else 0
                        area_progresses.append(area_progress)
                    
                    # Type progress is average of its areas
                    type_progress = sum(area_progresses) / len(area_progresses) if area_progresses else 0
                    type_progresses.append(type_progress)
                
                # Program progress is average of its types
                program_progress = sum(type_progresses) / len(type_progresses) if type_progresses else 0
                program_progresses.append(program_progress)
            
            # Department progress is the average of its programs' progress
            dept['progress'] = round(sum(program_progresses) / len(program_progresses)) if program_progresses else 0
            
    except Exception as e:
        print(f"Error calculating department progress: {str(e)}")
        for dept in departments:
            dept['progress'] = 0
    
    context = {
        'active_page': 'accreditation',
        'user': user,
        'departments': departments,
        'programs': programs,
        'types': types,
        'areas': areas,
    }
    return render(request, 'dashboard/accreditation.html', context)


@login_required
def accreditation_department_programs_view(request, dept_id):
    """Accreditation Programs page - read-only view of programs under a department"""
    user = get_user_from_session(request)
    
    # Get department info
    try:
        department = get_document('departments', dept_id)
        if not department:
            messages.error(request, 'Department not found.')
            return redirect('dashboard:accreditation')
    except Exception as e:
        print(f"Error fetching department: {str(e)}")
        messages.error(request, 'Error fetching department.')
        return redirect('dashboard:accreditation')
    
    # Get all active programs for this department
    try:
        all_programs = get_all_documents('programs')
        programs = [
            prog for prog in all_programs 
            if prog.get('department_id') == dept_id 
            and prog.get('is_active', True) 
            and not prog.get('is_archived', False)
        ]
        programs.sort(key=lambda x: x.get('code', ''))
        
        # Calculate progress for each program based on its types
        all_types = get_all_documents('accreditation_types')
        all_areas = get_all_documents('areas')
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for prog in programs:
            prog_id = prog.get('id')
            # Get all types for this program
            prog_types = [t for t in all_types if t.get('program_id') == prog_id]
            
            if not prog_types:
                prog['progress'] = 0
                continue
            
            # Calculate progress for each type first
            type_progresses = []
            for prog_type in prog_types:
                type_id = prog_type.get('id')
                type_areas = [a for a in all_areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                
                if not type_areas:
                    type_progresses.append(0)
                    continue
                
                # Calculate progress for each area
                area_progresses = []
                for area in type_areas:
                    area_id = area.get('id')
                    area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
                    
                    if not area_checklists:
                        area_progresses.append(0)
                        continue
                    
                    total_checklists = len(area_checklists)
                    completed_checklists = 0
                    
                    for checklist in area_checklists:
                        checklist_id = checklist.get('id')
                        required_docs = [
                            doc for doc in all_documents 
                            if doc.get('checklist_id') == checklist_id 
                            and doc.get('is_required', False)
                            and not doc.get('is_archived', False)
                            and doc.get('status') == 'approved'
                        ]
                        if len(required_docs) > 0:
                            completed_checklists += 1
                    
                    area_progress = (completed_checklists / total_checklists) * 100 if total_checklists > 0 else 0
                    area_progresses.append(area_progress)
                
                # Type progress is average of its areas
                type_progress = sum(area_progresses) / len(area_progresses) if area_progresses else 0
                type_progresses.append(type_progress)
            
            # Program progress is the average of its types' progress
            prog['progress'] = round(sum(type_progresses) / len(type_progresses)) if type_progresses else 0
            
    except Exception as e:
        print(f"Error fetching programs: {str(e)}")
        programs = []
    
    context = {
        'active_page': 'accreditation',
        'user': user,
        'department': department,
        'programs': programs,
        'dept_id': dept_id,
    }
    return render(request, 'dashboard/accreditation_programs.html', context)


@login_required
def accreditation_program_types_view(request, dept_id, prog_id):
    """Accreditation Types page - read-only view of types under a program"""
    user = get_user_from_session(request)
    
    # Get department and program info
    try:
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        if not department or not program:
            messages.error(request, 'Department or Program not found.')
            return redirect('dashboard:accreditation')
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        messages.error(request, 'Error fetching data.')
        return redirect('dashboard:accreditation')
    
    # Get all active types for this program
    try:
        all_types = get_all_documents('accreditation_types')
        types = [
            t for t in all_types 
            if t.get('program_id') == prog_id 
            and t.get('is_active', True) 
            and not t.get('is_archived', False)
        ]
        types.sort(key=lambda x: x.get('type', ''))
        
        # Calculate progress for each type based on its areas
        all_areas = get_all_documents('areas')
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for accred_type in types:
            type_id = accred_type.get('id')
            # Get all areas for this type (check both type_id and accreditation_type_id)
            type_areas = [a for a in all_areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
            
            if not type_areas:
                accred_type['progress'] = 0
                continue
            
            # Calculate progress for each area first
            area_progresses = []
            for area in type_areas:
                area_id = area.get('id')
                area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
                
                if not area_checklists:
                    area_progresses.append(0)
                    continue
                
                total_checklists = len(area_checklists)
                completed_checklists = 0
                
                for checklist in area_checklists:
                    checklist_id = checklist.get('id')
                    required_docs = [
                        doc for doc in all_documents 
                        if doc.get('checklist_id') == checklist_id 
                        and doc.get('is_required', False)
                        and not doc.get('is_archived', False)
                        and doc.get('status') == 'approved'
                    ]
                    if len(required_docs) > 0:
                        completed_checklists += 1
                
                area_progress = (completed_checklists / total_checklists) * 100 if total_checklists > 0 else 0
                area_progresses.append(area_progress)
            
            # Type progress is the average of its areas' progress
            accred_type['progress'] = round(sum(area_progresses) / len(area_progresses)) if area_progresses else 0
            
    except Exception as e:
        print(f"Error fetching types: {str(e)}")
        types = []
    
    context = {
        'active_page': 'accreditation',
        'user': user,
        'department': department,
        'program': program,
        'types': types,
        'dept_id': dept_id,
        'prog_id': prog_id,
    }
    return render(request, 'dashboard/accreditation_types.html', context)


@login_required
def accreditation_type_areas_view(request, dept_id, prog_id, type_id):
    """Accreditation Areas page - read-only view of areas under a type"""
    user = get_user_from_session(request)
    
    # Get breadcrumb info
    try:
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        accreditation_type = get_document('accreditation_types', type_id)
        if not department or not program or not accreditation_type:
            messages.error(request, 'Data not found.')
            return redirect('dashboard:accreditation')
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        messages.error(request, 'Error fetching data.')
        return redirect('dashboard:accreditation')
    
    # Get all active areas for this type
    try:
        all_areas = get_all_documents('areas')
        areas = [
            mod for mod in all_areas 
            if (mod.get('type_id') == type_id or mod.get('accreditation_type_id') == type_id)
            and mod.get('is_active', True) 
            and not mod.get('is_archived', False)
        ]
        areas.sort(key=lambda x: x.get('name', ''))
        
        # Calculate progress for each area
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for area in areas:
            area_id = area.get('id')
            # Get all checklists for this area
            area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
            
            if not area_checklists:
                area['progress'] = 0
                continue
            
            # Calculate progress based on required documents
            total_checklists = len(area_checklists)
            completed_checklists = 0
            
            for checklist in area_checklists:
                checklist_id = checklist.get('id')
                required_docs = [
                    doc for doc in all_documents 
                    if doc.get('checklist_id') == checklist_id 
                    and doc.get('is_required', False)
                    and not doc.get('is_archived', False)
                    and doc.get('status') == 'approved'
                ]
                if len(required_docs) > 0:
                    completed_checklists += 1
            
            area['progress'] = round((completed_checklists / total_checklists) * 100) if total_checklists > 0 else 0
            
    except Exception as e:
        print(f"Error fetching areas: {str(e)}")
        areas = []
    
    context = {
        'active_page': 'accreditation',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'areas': areas,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'type_id': type_id,
    }
    return render(request, 'dashboard/accreditation_areas.html', context)


@login_required
def accreditation_area_checklists_view(request, dept_id, prog_id, type_id, area_id):
    """Accreditation Checklists page - read-only view of checklists under a area"""
    user = get_user_from_session(request)
    
    # Get breadcrumb info
    try:
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        accreditation_type = get_document('accreditation_types', type_id)
        area = get_document('areas', area_id)
        if not department or not program or not accreditation_type or not area:
            messages.error(request, 'Data not found.')
            return redirect('dashboard:accreditation')
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        messages.error(request, 'Error fetching data.')
        return redirect('dashboard:accreditation')
    
    # Get all active checklists for this area
    try:
        all_checklists = get_all_documents('checklists')
        checklists = [
            checklist for checklist in all_checklists 
            if checklist.get('area_id') == area_id 
            and checklist.get('is_active', True) 
            and not checklist.get('is_archived', False)
        ]
        checklists.sort(key=lambda x: x.get('name', ''))
        
        # Get all documents to calculate progress for each checklist
        all_documents = get_all_documents('documents')
        
        # Add progress percentage to each checklist
        for checklist in checklists:
            checklist_id = checklist.get('id')
            # Get required documents for this checklist
            required_docs = [
                doc for doc in all_documents 
                if doc.get('checklist_id') == checklist_id 
                and doc.get('is_required', False)
                and not doc.get('is_archived', False)
                and doc.get('status') == 'approved'
            ]
            # Progress is 100% if there's at least 1 required document, otherwise 0%
            checklist['progress'] = 100 if len(required_docs) > 0 else 0
            
    except Exception as e:
        print(f"Error fetching checklists: {str(e)}")
        checklists = []
    
    context = {
        'active_page': 'accreditation',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'area': area,
        'checklists': checklists,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'type_id': type_id,
        'area_id': area_id,
    }
    return render(request, 'dashboard/accreditation_checklists.html', context)


@login_required
def performance_view(request):
    """Performance Management page"""
    user = get_user_from_session(request)
    
    try:
        # Fetch all data
        departments = get_all_documents('departments')
        departments = [d for d in departments if d.get('is_active', True) and not d.get('is_archived', False)]
        
        programs = get_all_documents('programs')
        programs = [p for p in programs if p.get('is_active', True) and not p.get('is_archived', False)]
        
        types = get_all_documents('accreditation_types')
        types = [t for t in types if t.get('is_active', True) and not t.get('is_archived', False)]
        
        areas = get_all_documents('areas')
        areas = [a for a in areas if a.get('is_active', True) and not a.get('is_archived', False)]
        
        all_checklists = get_all_documents('checklists')
        # Filter to get only active checklists
        active_checklists = [c for c in all_checklists if c.get('is_active', False) and not c.get('is_archived', False)]
        
        all_documents = get_all_documents('documents')
        # Filter to get only active documents
        active_documents = [d for d in all_documents if d.get('is_active', False) and not d.get('is_archived', False)]
        
        # Calculate progress for each department
        department_stats = []
        total_required_docs = 0
        total_uploaded_docs = 0
        total_checklists_count = len(active_checklists)
        total_completed_checklists = 0
        
        for dept in departments:
            dept_id = dept.get('id')
            dept_programs = [p for p in programs if p.get('department_id') == dept_id]
            
            dept_required_docs = 0
            dept_uploaded_docs = 0
            dept_checklists = 0
            dept_completed_checklists = 0
            
            if not dept_programs:
                dept_progress = 0
            else:
                program_progresses = []
                for prog in dept_programs:
                    prog_id = prog.get('id')
                    prog_types = [t for t in types if t.get('program_id') == prog_id]
                    
                    if not prog_types:
                        program_progresses.append(0)
                        continue
                    
                    type_progresses = []
                    for prog_type in prog_types:
                        type_id = prog_type.get('id')
                        type_areas = [a for a in areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                        
                        if not type_areas:
                            type_progresses.append(0)
                            continue
                        
                        area_progresses = []
                        for area in type_areas:
                            area_id = area.get('id')
                            area_checklists = [c for c in active_checklists if c.get('area_id') == area_id]
                            
                            if not area_checklists:
                                area_progresses.append(0)
                                continue
                            
                            dept_checklists += len(area_checklists)
                            
                            area_completed = 0
                            for checklist in area_checklists:
                                checklist_id = checklist.get('id')
                                # Check if checklist has at least one approved document
                                checklist_docs = [
                                    doc for doc in active_documents 
                                    if doc.get('checklist_id') == checklist_id
                                ]
                                
                                has_approved = any(d.get('status') == 'approved' for d in checklist_docs)
                                
                                if has_approved:
                                    area_completed += 1
                                    dept_completed_checklists += 1
                                    total_completed_checklists += 1
                                
                                # Count approved documents for this checklist
                                approved_docs = [d for d in checklist_docs if d.get('status') == 'approved']
                                dept_uploaded_docs += len(approved_docs)
                                total_uploaded_docs += len(approved_docs)
                            
                            area_progress = (area_completed / len(area_checklists)) * 100 if area_checklists else 0
                            area_progresses.append(area_progress)
                        
                        type_progress = sum(area_progresses) / len(area_progresses) if area_progresses else 0
                        type_progresses.append(type_progress)
                    
                    prog_progress = sum(type_progresses) / len(type_progresses) if type_progresses else 0
                    program_progresses.append(prog_progress)
                
                dept_progress = sum(program_progresses) / len(program_progresses) if program_progresses else 0
            
            # Count types and areas for this department
            dept_types = []
            dept_areas = []
            for prog in dept_programs:
                prog_id = prog.get('id')
                prog_types = [t for t in types if t.get('program_id') == prog_id]
                dept_types.extend(prog_types)
                
                for prog_type in prog_types:
                    type_id = prog_type.get('id')
                    type_areas = [a for a in areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                    dept_areas.extend(type_areas)
            
            department_stats.append({
                'id': dept_id,
                'name': dept.get('name', 'Unknown'),
                'logo_url': dept.get('logo_url', ''),
                'progress': round(dept_progress),
                'programs_count': len(dept_programs),
                'types_count': len(dept_types),
                'areas_count': len(dept_areas),
                'checklists_count': dept_checklists,
                'completed_checklists': dept_completed_checklists,
                'required_docs': dept_required_docs,
                'uploaded_docs': dept_uploaded_docs,
                'approved_docs': dept_uploaded_docs,  # Since we're only counting approved docs
                'updated_at': dept.get('updated_at', '')
            })
        
        # Sort departments by progress (descending)
        department_stats.sort(key=lambda x: x['progress'], reverse=True)
        
        # Calculate overall completion rate (consistent with dashboard)
        # Completion rate = completed checklists / total checklists * 100
        avg_completion = (total_completed_checklists / total_checklists_count * 100) if total_checklists_count > 0 else 0
        
        # Status distribution
        excellent_count = len([d for d in department_stats if d['progress'] >= 80])
        good_count = len([d for d in department_stats if 60 <= d['progress'] < 80])
        needs_improvement_count = len([d for d in department_stats if 40 <= d['progress'] < 60])
        critical_count = len([d for d in department_stats if d['progress'] < 40])
        
        context = {
            'active_page': 'performance',
            'user': user,
            'departments': department_stats,
            'total_departments': len(department_stats),
            'avg_completion': round(avg_completion, 1),
            'total_programs': len(programs),
            'total_checklists': total_checklists_count,
            'total_completed_checklists': total_completed_checklists,
            'total_required_docs': total_required_docs,
            'total_uploaded_docs': total_uploaded_docs,
            'excellent_count': excellent_count,
            'good_count': good_count,
            'needs_improvement_count': needs_improvement_count,
            'critical_count': critical_count,
            'all_programs': programs,
            'all_types': types,
        }
        
    except Exception as e:
        print(f"Error in performance view: {str(e)}")
        context = {
            'active_page': 'performance',
            'user': user,
            'departments': [],
            'total_departments': 0,
            'avg_completion': 0,
            'total_programs': 0,
            'total_checklists': 0,
            'total_completed_checklists': 0,
            'total_required_docs': 0,
            'total_uploaded_docs': 0,
            'excellent_count': 0,
            'good_count': 0,
            'needs_improvement_count': 0,
            'critical_count': 0,
            'all_programs': [],
            'all_types': [],
        }
    
    return render(request, 'dashboard/performance.html', context)


@login_required
def reports_view(request):
    """Reports page with generation and history"""
    user = get_user_from_session(request)
    
    # Fetch filter options
    try:
        departments = get_all_documents('departments')
        departments = [d for d in departments if d.get('is_active', True) and not d.get('is_archived', False)]
        departments.sort(key=lambda x: x.get('name', ''))
        
        programs = get_all_documents('programs')
        programs = [p for p in programs if p.get('is_active', True) and not p.get('is_archived', False)]
        programs.sort(key=lambda x: x.get('code', ''))
        
        types = get_all_documents('accreditation_types')
        types = [t for t in types if t.get('is_active', True) and not t.get('is_archived', False)]
        types.sort(key=lambda x: x.get('name', ''))
        
        # Fetch reports history
        reports_history = get_all_documents('reports_history')
        
        # Enrich reports with user names from users collection
        users = get_all_documents('users')
        user_map = {u.get('email'): f"{u.get('first_name', '')} {u.get('last_name', '')}".strip() or u.get('email') for u in users}
        
        for report in reports_history:
            generated_by = report.get('generated_by', '')
            if generated_by in user_map and not report.get('generated_by_name'):
                report['generated_by_name'] = user_map[generated_by]
        
        # Convert Firebase timestamps to ISO format strings for sorting and display
        from datetime import datetime
        for report in reports_history:
            created_at = report.get('created_at')
            if created_at and not isinstance(created_at, str):
                # Convert Firebase DatetimeWithNanoseconds to ISO string
                try:
                    report['created_at'] = created_at.isoformat()
                except:
                    report['created_at'] = str(created_at)
        
        reports_history.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        # Calculate stats
        total_reports = len(reports_history)
        
        # Reports this month
        current_month = datetime.now().strftime('%Y-%m')
        reports_this_month = 0
        for r in reports_history:
            created_at = r.get('created_at', '')
            if isinstance(created_at, str) and created_at.startswith(current_month):
                reports_this_month += 1
        
        # Calculate total storage (in MB)
        total_storage = sum(r.get('file_size', 0) for r in reports_history) / (1024 * 1024)
        
        # Active users generating reports
        active_users = len(set(r.get('generated_by') for r in reports_history if r.get('generated_by')))
        
    except Exception as e:
        print(f"Error fetching reports data: {str(e)}")
        departments = []
        programs = []
        types = []
        reports_history = []
        total_reports = 0
        reports_this_month = 0
        total_storage = 0
        active_users = 0
    
    context = {
        'active_page': 'reports',
        'user': user,
        'departments': departments,
        'programs': programs,
        'types': types,
        'reports_history': reports_history,
        'total_reports': total_reports,
        'reports_this_month': reports_this_month,
        'total_storage': round(total_storage, 2),
        'active_users': active_users,
    }
    return render(request, 'dashboard/reports.html', context)


@login_required
def generate_report(request):
    """Generate report based on filters and type"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        import json
        from datetime import datetime
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        import io
        from .cloudinary_utils import upload_file_to_cloudinary
        
        user = get_user_from_session(request)
        data = json.loads(request.body)
        
        report_type = data.get('report_type')
        report_format = data.get('format', 'pdf')
        department_id = data.get('department_id', '')
        program_id = data.get('program_id', '')
        type_id = data.get('type_id', '')
        date_from = data.get('date_from', '')
        date_to = data.get('date_to', '')
        
        # Build scope description
        scope_parts = []
        if department_id:
            dept = get_document('departments', department_id)
            if dept:
                scope_parts.append(f"Department: {dept.get('name')}")
        if program_id:
            prog = get_document('programs', program_id)
            if prog:
                scope_parts.append(f"Program: {prog.get('name')}")
        if type_id:
            accred_type = get_document('accreditation_types', type_id)
            if accred_type:
                scope_parts.append(f"Type: {accred_type.get('name')}")
        if date_from and date_to:
            scope_parts.append(f"Period: {date_from} to {date_to}")
        
        scope = " | ".join(scope_parts) if scope_parts else "All Data"
        
        # Generate report based on format
        if report_format == 'pdf':
            file_data = generate_pdf_report(report_type, department_id, program_id, type_id, date_from, date_to, user)
            file_extension = 'pdf'
            content_type = 'application/pdf'
        elif report_format == 'excel':
            file_data = generate_excel_report(report_type, department_id, program_id, type_id, date_from, date_to, user)
            file_extension = 'xlsx'
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            return JsonResponse({'success': False, 'message': 'Invalid format'})
        
        # Upload to Cloudinary
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_type}_{timestamp}.{file_extension}"
        
        cloudinary_url = upload_file_to_cloudinary(file_data, filename, folder='reports')
        
        # Get user details from users collection
        user_email = user.get('email', 'Unknown')
        user_name = 'Unknown User'
        try:
            # Try to get user from users collection
            users = get_all_documents('users')
            user_doc = next((u for u in users if u.get('email') == user_email), None)
            if user_doc:
                first_name = user_doc.get('first_name', '')
                last_name = user_doc.get('last_name', '')
                user_name = f"{first_name} {last_name}".strip() or user_email
            else:
                # Fallback to user object if not found
                user_name = user.get('displayName', user.get('name', user_email))
        except Exception as e:
            print(f"Error fetching user details: {e}")
            user_name = user.get('displayName', user.get('name', user_email))
        
        # Save to reports history
        report_data = {
            'id': f"report_{timestamp}_{user.get('uid', 'unknown')}",
            'type': report_type,
            'generated_by': user_email,
            'generated_by_name': user_name,
            'scope': scope,
            'created_at': datetime.now().isoformat(),
            'format': file_extension.upper(),
            'status': 'completed',
            'file_size': len(file_data),
            'file_url': cloudinary_url,
            'file_path': cloudinary_url,
        }
        
        create_document('reports_history', report_data, report_data['id'])
        try:
            report_desc = f"{report_type.replace('_', ' ').title()} ({file_extension.upper()})"
            scope_desc = f" - {scope}" if scope != 'all' else ""
            log_audit(user, action_type='report_generation', resource_type='report', resource_id=report_data['id'], details=f"Generated report: {report_desc}{scope_desc}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Report generated successfully',
            'report_id': report_data['id'],
            'download_url': cloudinary_url
        })
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


def generate_pdf_report(report_type, department_id, program_id, type_id, date_from, date_to, user):
    """Generate PDF report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#4CAF50'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # Header
    story.append(Paragraph("PLP Accreditation System", title_style))
    report_title = {
        'complete_accreditation': 'Complete Accreditation Report',
        'results_incentives': 'Results and Incentives Report',
        'performance_analytics': 'Performance Analytics Report'
    }.get(report_type, 'System Report')
    
    story.append(Paragraph(report_title, heading_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Paragraph(f"Generated by: {user.get('displayName', 'Unknown User')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Fetch data based on report type
    if report_type == 'complete_accreditation':
        story.extend(generate_complete_accreditation_content(department_id, program_id, type_id, styles, heading_style))
    elif report_type == 'results_incentives':
        story.extend(generate_results_incentives_content(department_id, program_id, type_id, styles, heading_style))
    elif report_type == 'performance_analytics':
        story.extend(generate_performance_analytics_content(department_id, program_id, type_id, styles, heading_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_complete_accreditation_content(department_id, program_id, type_id, styles, heading_style):
    """Generate content for complete accreditation report"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    
    content = []
    content.append(Paragraph("Executive Summary", heading_style))
    
    # Fetch all data
    departments = get_all_documents('departments')
    programs = get_all_documents('programs')
    types = get_all_documents('accreditation_types')
    areas = get_all_documents('areas')
    checklists = get_all_documents('checklists')
    documents = get_all_documents('documents')
    
    # Filter by selections
    if department_id:
        departments = [d for d in departments if d.get('id') == department_id]
    if program_id:
        programs = [p for p in programs if p.get('id') == program_id]
    if type_id:
        types = [t for t in types if t.get('id') == type_id]
    
    departments = [d for d in departments if d.get('is_active', True) and not d.get('is_archived', False)]
    programs = [p for p in programs if p.get('is_active', True) and not p.get('is_archived', False)]
    types = [t for t in types if t.get('is_active', True) and not t.get('is_archived', False)]
    areas = [a for a in areas if a.get('is_active', True) and not a.get('is_archived', False)]
    
    # Summary statistics
    total_areas = len(areas)
    total_checklists = len(checklists)
    total_docs = len([d for d in documents if d.get('is_required', False)])
    approved_docs = len([d for d in documents if d.get('is_required', False) and d.get('status') == 'approved'])
    
    summary_data = [
        ['Metric', 'Count'],
        ['Departments', str(len(departments))],
        ['Programs', str(len(programs))],
        ['Accreditation Types', str(len(types))],
        ['Areas', str(total_areas)],
        ['Checklists', str(total_checklists)],
        ['Required Documents', str(total_docs)],
        ['Approved Documents', str(approved_docs)],
        ['Overall Progress', f"{round((approved_docs/total_docs*100) if total_docs > 0 else 0, 1)}%"],
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    
    content.append(summary_table)
    content.append(Spacer(1, 0.3*inch))
    
    # Department details
    content.append(Paragraph("Department Details", heading_style))
    
    for dept in departments:
        dept_programs = [p for p in programs if p.get('department_id') == dept.get('id')]
        
        dept_data = [[f"{dept.get('name')} ({dept.get('code')})"]]
        dept_table = Table(dept_data, colWidths=[6*inch])
        dept_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F5E9')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2E7D32')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        content.append(dept_table)
        content.append(Spacer(1, 0.1*inch))
        
        # Programs table
        if dept_programs:
            prog_data = [['Program Code', 'Program Name', 'Types', 'Areas', 'Progress']]
            
            for prog in dept_programs:
                prog_types = [t for t in types if t.get('program_id') == prog.get('id')]
                prog_areas = []
                for t in prog_types:
                    prog_areas.extend([a for a in areas if a.get('type_id') == t.get('id') or a.get('accreditation_type_id') == t.get('id')])
                
                # Calculate progress
                prog_checklists = []
                for area in prog_areas:
                    prog_checklists.extend([c for c in checklists if c.get('area_id') == area.get('id')])
                
                completed = 0
                for checklist in prog_checklists:
                    req_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id') and d.get('is_required', False) and d.get('status') == 'approved']
                    if len(req_docs) > 0:
                        completed += 1
                
                progress = round((completed / len(prog_checklists) * 100) if len(prog_checklists) > 0 else 0, 1)
                
                prog_data.append([
                    prog.get('code', ''),
                    prog.get('name', ''),
                    str(len(prog_types)),
                    str(len(prog_areas)),
                    f"{progress}%"
                ])
            
            prog_table = Table(prog_data, colWidths=[1*inch, 2.5*inch, 0.8*inch, 0.8*inch, 0.9*inch])
            prog_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
            ]))
            content.append(prog_table)
        
        content.append(Spacer(1, 0.2*inch))
    
    return content


def generate_results_incentives_content(department_id, program_id, type_id, styles, heading_style):
    """Generate content for results and incentives report"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    
    content = []
    content.append(Paragraph("Results and Incentives Overview", heading_style))
    
    # Fetch all data
    departments = get_all_documents('departments')
    programs = get_all_documents('programs')
    types = get_all_documents('accreditation_types')
    areas = get_all_documents('areas')
    checklists = get_all_documents('checklists')
    documents = get_all_documents('documents')
    
    # Filter by selections
    if department_id:
        departments = [d for d in departments if d.get('id') == department_id]
    if program_id:
        programs = [p for p in programs if p.get('id') == program_id]
    if type_id:
        types = [t for t in types if t.get('id') == type_id]
    
    departments = [d for d in departments if d.get('is_active', True) and not d.get('is_archived', False)]
    programs = [p for p in programs if p.get('is_active', True) and not p.get('is_archived', False)]
    types = [t for t in types if t.get('is_active', True) and not t.get('is_archived', False)]
    areas = [a for a in areas if a.get('is_active', True) and not a.get('is_archived', False)]
    
    # Build results data
    results_data = [['Department', 'Program', 'Type', 'Area', 'Progress', 'Certificate', 'Incentive']]
    
    for area in areas:
        type_id_val = area.get('type_id') or area.get('accreditation_type_id')
        accred_type = next((t for t in types if t.get('id') == type_id_val), None)
        if not accred_type:
            continue
        
        prog_id = accred_type.get('program_id')
        program = next((p for p in programs if p.get('id') == prog_id), None)
        if not program:
            continue
        
        dept_id = program.get('department_id')
        department = next((d for d in departments if d.get('id') == dept_id), None)
        if not department:
            continue
        
        # Calculate progress
        area_checklists = [c for c in checklists if c.get('area_id') == area.get('id')]
        if not area_checklists:
            progress = 0
        else:
            completed = 0
            for checklist in area_checklists:
                req_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id') and d.get('is_required', False) and d.get('status') == 'approved']
                if len(req_docs) > 0:
                    completed += 1
            progress = round((completed / len(area_checklists) * 100) if len(area_checklists) > 0 else 0, 1)
        
        certificate = "Issued" if area.get('certificate_issued', False) else "Pending"
        incentive = "Eligible" if progress >= 80 else "Not Eligible"
        
        results_data.append([
            department.get('code', ''),
            program.get('code', ''),
            accred_type.get('name', '')[:15],
            area.get('name', '')[:20],
            f"{progress}%",
            certificate,
            incentive
        ])
    
    if len(results_data) > 1:
        results_table = Table(results_data, colWidths=[0.8*inch, 0.8*inch, 1.2*inch, 1.5*inch, 0.7*inch, 0.8*inch, 0.9*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        content.append(results_table)
    else:
        content.append(Paragraph("No data available for selected filters.", styles['Normal']))
    
    content.append(Spacer(1, 0.3*inch))
    
    # Summary statistics
    total_areas = len(results_data) - 1
    certificates_issued = sum(1 for row in results_data[1:] if row[5] == "Issued")
    eligible_incentives = sum(1 for row in results_data[1:] if row[6] == "Eligible")
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Areas', str(total_areas)],
        ['Certificates Issued', str(certificates_issued)],
        ['Incentive Eligible', str(eligible_incentives)],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    content.append(summary_table)
    
    return content


def generate_performance_analytics_content(department_id, program_id, type_id, styles, heading_style):
    """Generate content for performance analytics report"""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    
    content = []
    content.append(Paragraph("Performance Analytics", heading_style))
    
    # Fetch all data
    departments = get_all_documents('departments')
    programs = get_all_documents('programs')
    types = get_all_documents('accreditation_types')
    areas = get_all_documents('areas')
    checklists = get_all_documents('checklists')
    documents = get_all_documents('documents')
    
    # Filter by selections
    if department_id:
        departments = [d for d in departments if d.get('id') == department_id]
    if program_id:
        programs = [p for p in programs if p.get('id') == program_id]
    if type_id:
        types = [t for t in types if t.get('id') == type_id]
    
    departments = [d for d in departments if d.get('is_active', True) and not d.get('is_archived', False)]
    programs = [p for p in programs if p.get('is_active', True) and not p.get('is_archived', False)]
    types = [t for t in types if t.get('is_active', True) and not t.get('is_archived', False)]
    
    # Calculate department performance
    dept_data = [['Rank', 'Department', 'Programs', 'Checklists', 'Documents', 'Progress', 'Status']]
    dept_performance = []
    
    for dept in departments:
        dept_programs = [p for p in programs if p.get('department_id') == dept.get('id')]
        dept_types = []
        dept_areas = []
        for prog in dept_programs:
            prog_types = [t for t in types if t.get('program_id') == prog.get('id')]
            dept_types.extend(prog_types)
            for t in prog_types:
                dept_areas.extend([a for a in areas if (a.get('type_id') == t.get('id') or a.get('accreditation_type_id') == t.get('id')) and a.get('is_active', True)])
        
        dept_checklists = []
        for area in dept_areas:
            dept_checklists.extend([c for c in checklists if c.get('area_id') == area.get('id')])
        
        completed_checklists = 0
        required_docs = 0
        approved_docs = 0
        
        for checklist in dept_checklists:
            checklist_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id') and d.get('is_required', False)]
            required_docs += len(checklist_docs)
            approved = [d for d in checklist_docs if d.get('status') == 'approved']
            approved_docs += len(approved)
            if len(approved) > 0:
                completed_checklists += 1
        
        progress = round((approved_docs / required_docs * 100) if required_docs > 0 else 0, 1)
        
        if progress >= 80:
            status = "Excellent"
        elif progress >= 60:
            status = "Good"
        elif progress >= 40:
            status = "Needs Work"
        else:
            status = "Critical"
        
        dept_performance.append({
            'name': dept.get('name', ''),
            'programs': len(dept_programs),
            'checklists': f"{completed_checklists}/{len(dept_checklists)}",
            'documents': f"{approved_docs}/{required_docs}",
            'progress': progress,
            'status': status
        })
    
    # Sort by progress
    dept_performance.sort(key=lambda x: x['progress'], reverse=True)
    
    for idx, dept in enumerate(dept_performance, 1):
        dept_data.append([
            str(idx),
            dept['name'],
            str(dept['programs']),
            dept['checklists'],
            dept['documents'],
            f"{dept['progress']}%",
            dept['status']
        ])
    
    if len(dept_data) > 1:
        perf_table = Table(dept_data, colWidths=[0.5*inch, 2*inch, 0.8*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
        perf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ]))
        content.append(perf_table)
    else:
        content.append(Paragraph("No data available for selected filters.", styles['Normal']))
    
    content.append(Spacer(1, 0.3*inch))
    
    # Performance summary
    if dept_performance:
        avg_progress = sum(d['progress'] for d in dept_performance) / len(dept_performance)
        excellent = sum(1 for d in dept_performance if d['status'] == 'Excellent')
        good = sum(1 for d in dept_performance if d['status'] == 'Good')
        needs_work = sum(1 for d in dept_performance if d['status'] == 'Needs Work')
        critical = sum(1 for d in dept_performance if d['status'] == 'Critical')
        
        summary_data = [
            ['Metric', 'Value'],
            ['Average Progress', f"{round(avg_progress, 1)}%"],
            ['Excellent Departments', str(excellent)],
            ['Good Departments', str(good)],
            ['Needs Improvement', str(needs_work)],
            ['Critical Departments', str(critical)],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        content.append(summary_table)
    
    return content


def generate_excel_report(report_type, department_id, program_id, type_id, date_from, date_to, user):
    """Generate Excel report"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Header styling
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    table_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    table_header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    
    cell_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'PLP Accreditation System'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.row_dimensions[1].height = 25
    
    # Report type
    ws.merge_cells('A2:G2')
    report_title = {
        'complete_accreditation': 'Complete Accreditation Report',
        'results_incentives': 'Results and Incentives Report',
        'performance_analytics': 'Performance Analytics Report'
    }.get(report_type, 'System Report')
    ws['A2'] = report_title
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Metadata
    ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A4'] = f"Generated by: {user.get('displayName', 'Unknown User')}"
    
    current_row = 6
    
    # Fetch data based on report type
    if report_type == 'complete_accreditation':
        current_row = generate_excel_complete_accreditation(ws, department_id, program_id, type_id, current_row, table_header_font, table_header_fill, cell_alignment, border)
    elif report_type == 'results_incentives':
        current_row = generate_excel_results_incentives(ws, department_id, program_id, type_id, current_row, table_header_font, table_header_fill, cell_alignment, border)
    elif report_type == 'performance_analytics':
        current_row = generate_excel_performance_analytics(ws, department_id, program_id, type_id, current_row, table_header_font, table_header_fill, cell_alignment, border)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_complete_accreditation(ws, department_id, program_id, type_id, start_row, header_font, header_fill, cell_alignment, border):
    """Generate Excel content for complete accreditation report"""
    from openpyxl.styles import Alignment
    
    # Fetch data
    departments = get_all_documents('departments')
    programs = get_all_documents('programs')
    types = get_all_documents('accreditation_types')
    areas = get_all_documents('areas')
    checklists = get_all_documents('checklists')
    documents = get_all_documents('documents')
    
    # Filter
    if department_id:
        departments = [d for d in departments if d.get('id') == department_id]
    if program_id:
        programs = [p for p in programs if p.get('id') == program_id]
    if type_id:
        types = [t for t in types if t.get('id') == type_id]
    
    departments = [d for d in departments if d.get('is_active', True)]
    programs = [p for p in programs if p.get('is_active', True)]
    types = [t for t in types if t.get('is_active', True)]
    areas = [a for a in areas if a.get('is_active', True)]
    
    # Headers
    headers = ['Department', 'Program Code', 'Program Name', 'Type', 'Area', 'Checklists', 'Progress']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    current_row = start_row + 1
    
    # Data rows
    for dept in departments:
        dept_programs = [p for p in programs if p.get('department_id') == dept.get('id')]
        for prog in dept_programs:
            prog_types = [t for t in types if t.get('program_id') == prog.get('id')]
            for ptype in prog_types:
                type_areas = [a for a in areas if a.get('type_id') == ptype.get('id') or a.get('accreditation_type_id') == ptype.get('id')]
                for area in type_areas:
                    area_checklists = [c for c in checklists if c.get('area_id') == area.get('id')]
                    
                    completed = 0
                    for checklist in area_checklists:
                        req_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id') and d.get('is_required', False) and d.get('status') == 'approved']
                        if len(req_docs) > 0:
                            completed += 1
                    
                    progress = round((completed / len(area_checklists) * 100) if len(area_checklists) > 0 else 0, 1)
                    
                    row_data = [
                        dept.get('name', ''),
                        prog.get('code', ''),
                        prog.get('name', ''),
                        ptype.get('name', ''),
                        area.get('name', ''),
                        f"{completed}/{len(area_checklists)}",
                        f"{progress}%"
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.alignment = cell_alignment
                        cell.border = border
                    
                    current_row += 1
    
    return current_row + 2


def generate_excel_results_incentives(ws, department_id, program_id, type_id, start_row, header_font, header_fill, cell_alignment, border):
    """Generate Excel content for results and incentives report"""
    from openpyxl.styles import Alignment
    
    # Fetch data
    departments = get_all_documents('departments')
    programs = get_all_documents('programs')
    types = get_all_documents('accreditation_types')
    areas = get_all_documents('areas')
    checklists = get_all_documents('checklists')
    documents = get_all_documents('documents')
    
    # Filter
    if department_id:
        departments = [d for d in departments if d.get('id') == department_id]
    if program_id:
        programs = [p for p in programs if p.get('id') == program_id]
    if type_id:
        types = [t for t in types if t.get('id') == type_id]
    
    departments = [d for d in departments if d.get('is_active', True)]
    programs = [p for p in programs if p.get('is_active', True)]
    types = [t for t in types if t.get('is_active', True)]
    areas = [a for a in areas if a.get('is_active', True)]
    
    # Headers
    headers = ['Department', 'Program', 'Type', 'Area', 'Progress', 'Certificate', 'Incentive Eligible']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    current_row = start_row + 1
    
    # Data rows
    for area in areas:
        type_id_val = area.get('type_id') or area.get('accreditation_type_id')
        accred_type = next((t for t in types if t.get('id') == type_id_val), None)
        if not accred_type:
            continue
        
        prog_id = accred_type.get('program_id')
        program = next((p for p in programs if p.get('id') == prog_id), None)
        if not program:
            continue
        
        dept_id = program.get('department_id')
        department = next((d for d in departments if d.get('id') == dept_id), None)
        if not department:
            continue
        
        # Calculate progress
        area_checklists = [c for c in checklists if c.get('area_id') == area.get('id')]
        if not area_checklists:
            progress = 0
        else:
            completed = 0
            for checklist in area_checklists:
                req_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id') and d.get('is_required', False) and d.get('status') == 'approved']
                if len(req_docs) > 0:
                    completed += 1
            progress = round((completed / len(area_checklists) * 100) if len(area_checklists) > 0 else 0, 1)
        
        certificate = "Issued" if area.get('certificate_issued', False) else "Pending"
        incentive = "Yes" if progress >= 80 else "No"
        
        row_data = [
            department.get('code', ''),
            program.get('code', ''),
            accred_type.get('name', ''),
            area.get('name', ''),
            f"{progress}%",
            certificate,
            incentive
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
        
        current_row += 1
    
    return current_row + 2


def generate_excel_performance_analytics(ws, department_id, program_id, type_id, start_row, header_font, header_fill, cell_alignment, border):
    """Generate Excel content for performance analytics report"""
    from openpyxl.styles import Alignment
    
    # Fetch data
    departments = get_all_documents('departments')
    programs = get_all_documents('programs')
    types = get_all_documents('accreditation_types')
    areas = get_all_documents('areas')
    checklists = get_all_documents('checklists')
    documents = get_all_documents('documents')
    
    # Filter
    if department_id:
        departments = [d for d in departments if d.get('id') == department_id]
    if program_id:
        programs = [p for p in programs if p.get('id') == program_id]
    if type_id:
        types = [t for t in types if t.get('id') == type_id]
    
    departments = [d for d in departments if d.get('is_active', True)]
    programs = [p for p in programs if p.get('is_active', True)]
    types = [t for t in types if t.get('is_active', True)]
    
    # Headers
    headers = ['Rank', 'Department', 'Programs', 'Checklists', 'Documents', 'Progress', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Calculate performance
    dept_performance = []
    for dept in departments:
        dept_programs = [p for p in programs if p.get('department_id') == dept.get('id')]
        dept_types = []
        dept_areas = []
        for prog in dept_programs:
            prog_types = [t for t in types if t.get('program_id') == prog.get('id')]
            dept_types.extend(prog_types)
            for t in prog_types:
                dept_areas.extend([a for a in areas if (a.get('type_id') == t.get('id') or a.get('accreditation_type_id') == t.get('id')) and a.get('is_active', True)])
        
        dept_checklists = []
        for area in dept_areas:
            dept_checklists.extend([c for c in checklists if c.get('area_id') == area.get('id')])
        
        completed_checklists = 0
        required_docs = 0
        approved_docs = 0
        
        for checklist in dept_checklists:
            checklist_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id') and d.get('is_required', False)]
            required_docs += len(checklist_docs)
            approved = [d for d in checklist_docs if d.get('status') == 'approved']
            approved_docs += len(approved)
            if len(approved) > 0:
                completed_checklists += 1
        
        progress = round((approved_docs / required_docs * 100) if required_docs > 0 else 0, 1)
        
        if progress >= 80:
            status = "Excellent"
        elif progress >= 60:
            status = "Good"
        elif progress >= 40:
            status = "Needs Work"
        else:
            status = "Critical"
        
        dept_performance.append({
            'name': dept.get('name', ''),
            'programs': len(dept_programs),
            'checklists': f"{completed_checklists}/{len(dept_checklists)}",
            'documents': f"{approved_docs}/{required_docs}",
            'progress': progress,
            'status': status
        })
    
    # Sort by progress
    dept_performance.sort(key=lambda x: x['progress'], reverse=True)
    
    current_row = start_row + 1
    for idx, dept in enumerate(dept_performance, 1):
        row_data = [
            idx,
            dept['name'],
            dept['programs'],
            dept['checklists'],
            dept['documents'],
            f"{dept['progress']}%",
            dept['status']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
        
        current_row += 1
    
    return current_row + 2


@login_required
def delete_report(request, report_id):
    """Delete a report from history"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        user = get_user_from_session(request)
        # Get report
        report = get_document('reports_history', report_id)
        if not report:
            return JsonResponse({'success': False, 'message': 'Report not found'})
        
        # Delete from Cloudinary (optional - you might want to keep files)
        # Note: Cloudinary deletion would require additional setup
        
        # Delete from database
        delete_document('reports_history', report_id)
        try:
            report_type = report.get('type', 'Unknown')
            report_format = report.get('format', 'Unknown')
            log_audit(user, action_type='delete', resource_type='report', resource_id=report_id, details=f"Deleted report: {report_type} ({report_format})", status='success')
        except Exception:
            pass
        
        return JsonResponse({'success': True, 'message': 'Report deleted successfully'})
        
    except Exception as e:
        print(f"Error deleting report: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def results_view(request):
    """Results and Incentives page"""
    user = get_user_from_session(request)
    
    # Fetch all departments, programs, and types for filters
    try:
        departments = get_all_documents('departments')
        departments = [d for d in departments if d.get('is_active', True) and not d.get('is_archived', False)]
        departments.sort(key=lambda x: x.get('name', ''))
        
        programs = get_all_documents('programs')
        programs = [p for p in programs if p.get('is_active', True) and not p.get('is_archived', False)]
        programs.sort(key=lambda x: x.get('code', ''))
        
        types = get_all_documents('accreditation_types')
        types = [t for t in types if t.get('is_active', True) and not t.get('is_archived', False)]
        types.sort(key=lambda x: x.get('name', ''))
        
        # Fetch all areas with their complete hierarchy
        all_areas = get_all_documents('areas')
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        # Build area data with progress and hierarchy info
        areas_data = []
        for area in all_areas:
            if not area.get('is_active', True) or area.get('is_archived', False):
                continue
            
            area_id = area.get('id')
            type_id = area.get('type_id') or area.get('accreditation_type_id')
            
            # Get type info
            accred_type = next((t for t in types if t.get('id') == type_id), None)
            if not accred_type:
                continue
            
            # Get program info
            prog_id = accred_type.get('program_id')
            program = next((p for p in programs if p.get('id') == prog_id), None)
            if not program:
                continue
            
            # Get department info
            dept_id = program.get('department_id')
            department = next((d for d in departments if d.get('id') == dept_id), None)
            if not department:
                continue
            
            # Calculate area progress
            area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
            if not area_checklists:
                progress = 0
            else:
                total_checklists = len(area_checklists)
                completed_checklists = 0
                
                for checklist in area_checklists:
                    checklist_id = checklist.get('id')
                    required_docs = [
                        doc for doc in all_documents 
                        if doc.get('checklist_id') == checklist_id 
                        and doc.get('is_required', False)
                        and not doc.get('is_archived', False)
                        and doc.get('status') == 'approved'
                    ]
                    if len(required_docs) > 0:
                        completed_checklists += 1
                
                progress = round((completed_checklists / total_checklists) * 100) if total_checklists > 0 else 0
            
            areas_data.append({
                'area_id': area_id,
                'area_name': area.get('name', ''),
                'dept_code': department.get('code', ''),
                'dept_name': department.get('name', ''),
                'prog_code': program.get('code', ''),
                'prog_name': program.get('name', ''),
                'type_id': type_id,
                'type_name': accred_type.get('name', ''),
                'progress': progress,
                'certificate_issued': area.get('certificate_issued', False)
            })
        
        # Sort by department, program, type, area
        areas_data.sort(key=lambda x: (x['dept_code'], x['prog_code'], x['type_name'], x['area_name']))
        
    except Exception as e:
        print(f"Error fetching results data: {str(e)}")
        departments = []
        programs = []
        types = []
        areas_data = []
    
    context = {
        'active_page': 'results',
        'user': user,
        'departments': departments,
        'programs': programs,
        'types': types,
        'areas': areas_data,
    }
    return render(request, 'dashboard/results.html', context)


@login_required
@require_http_methods(["POST"])
def toggle_certificate_view(request, area_id):
    """Toggle certificate issuance status for an area"""
    try:
        import json
        
        # Parse request body
        data = json.loads(request.body)
        action = data.get('action')  # 'issue' or 'revoke'
        
        # Get the area document
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        # Calculate area progress to validate 100% requirement
        checklists = [c for c in get_all_documents('checklists') if c.get('area_id') == area_id]
        documents = get_all_documents('documents')
        
        total_checklists = len(checklists)
        completed_checklists = 0
        
        for checklist in checklists:
            checklist_id = checklist.get('id')
            # Check if checklist has at least one required document
            has_required_doc = any(
                doc.get('checklist_id') == checklist_id and doc.get('is_required', False)
                for doc in documents
            )
            if has_required_doc:
                completed_checklists += 1
        
        area_progress = round((completed_checklists / total_checklists * 100)) if total_checklists > 0 else 0
        
        # Validate that area is 100% complete before issuing certificate
        if action == 'issue' and area_progress < 100:
            return JsonResponse({
                'success': False,
                'message': f'Cannot issue certificate. Area is only {area_progress}% complete.'
            }, status=400)
        
        # Update certificate status
        new_status = (action == 'issue')
        update_document('areas', area_id, {'certificate_issued': new_status})
        
        # Return success response
        return JsonResponse({
            'success': True,
            'message': f'Certificate {"issued" if new_status else "revoked"} successfully',
            'certificate_issued': new_status
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
def audit_view(request):
    """Audit Trail page"""
    context = {
        'active_page': 'audit',
        'user': get_user_from_session(request),
    }
    return render(request, 'dashboard/audit.html', context)


@login_required
def archive_view(request):
    """Archive page"""
    context = {
        'active_page': 'archive',
        'user': get_user_from_session(request),
    }
    return render(request, 'dashboard/archive.html', context)


@login_required
def settings_view(request):
    """
    Profile Settings page - Shows available settings based on user role
    """
    user = get_user_from_session(request)
    
    # Fetch complete user data from Firestore
    from accreditation.firebase_utils import get_document
    from accreditation.firebase_auth import UserRole
    from datetime import datetime
    user_doc = safe_get_document('users', user['id'])
    
    if user_doc:
        # Get department information if user has a department
        department_name = 'N/A'
        if user_doc.get('department_id'):
            dept_doc = get_document('departments', user_doc['department_id'])
            if dept_doc:
                department_name = dept_doc.get('name', 'N/A')
        
        # Get role display name
        role_display = UserRole.get_role_display(user.get('role', ''))
        
        # Convert Firestore timestamps to datetime objects
        created_at = user_doc.get('created_at')
        print(f"DEBUG - created_at raw: {created_at}, type: {type(created_at)}")
        
        if created_at:
            if hasattr(created_at, 'timestamp'):
                # Firestore timestamp object
                created_at = datetime.fromtimestamp(created_at.timestamp())
                print(f"DEBUG - created_at converted (timestamp): {created_at}")
            elif isinstance(created_at, str):
                # ISO format string - parse it
                from datetime import datetime as dt
                try:
                    # Try parsing ISO format: 2025-10-25T03:03:01.081234
                    created_at = dt.fromisoformat(created_at.replace('Z', '+00:00'))
                    print(f"DEBUG - created_at converted (string): {created_at}")
                except Exception as e:
                    print(f"DEBUG - Error parsing created_at: {e}")
                    created_at = None
        else:
            print("DEBUG - created_at is None or empty")
        
        last_login = user_doc.get('last_login')
        if last_login:
            if hasattr(last_login, 'timestamp'):
                last_login = datetime.fromtimestamp(last_login.timestamp())
            elif isinstance(last_login, str):
                from datetime import datetime as dt
                try:
                    last_login = dt.fromisoformat(last_login.replace('Z', '+00:00'))
                except:
                    last_login = None
        
        last_password_change = user_doc.get('last_password_change')
        if last_password_change:
            if hasattr(last_password_change, 'timestamp'):
                last_password_change = datetime.fromtimestamp(last_password_change.timestamp())
            elif isinstance(last_password_change, str):
                from datetime import datetime as dt
                try:
                    last_password_change = dt.fromisoformat(last_password_change.replace('Z', '+00:00'))
                except:
                    last_password_change = None
        
        print(f"DEBUG - Final created_at being sent to template: {created_at}")
        
        # Merge session user with complete user data
        user.update({
            'first_name': user_doc.get('first_name', ''),
            'middle_name': user_doc.get('middle_name', ''),
            'last_name': user_doc.get('last_name', ''),
            'profile_image_url': user_doc.get('profile_image_url', ''),
            'department_name': department_name,
            'role_display': role_display,
            'is_active': user_doc.get('is_active', True),
            'created_at': created_at,
            'last_login': last_login,
            'last_password_change': last_password_change,
        })
    
    context = {
        'active_page': 'settings',
        'user': user,
    }
    return render(request, 'dashboard/settings.html', context)


@login_required
def upload_profile_image_view(request):
    """Upload user profile image"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    user = get_user_from_session(request)
    
    if 'profile_image' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No image provided'})
    
    image_file = request.FILES['profile_image']
    
    # Validate file size (2MB)
    if image_file.size > 2 * 1024 * 1024:
        return JsonResponse({'success': False, 'message': 'File size must be less than 2MB'})
    
    # Validate file type
    allowed_types = ['image/jpeg', 'image/jpg', 'image/png']
    if image_file.content_type not in allowed_types:
        return JsonResponse({'success': False, 'message': 'Only JPG and PNG files are allowed'})
    
    try:
        from accreditation.firebase_utils import update_document
        from accreditation.cloudinary_utils import upload_image_to_cloudinary, delete_image_from_cloudinary
        from datetime import datetime
        
        # Get old profile image URL if exists
        old_image_url = user.get('profile_image_url', '')
        
        # Delete old image from Cloudinary if it exists
        if old_image_url and 'cloudinary.com' in old_image_url:
            delete_image_from_cloudinary(old_image_url)
        
        # Upload new image to Cloudinary in 'profile' folder
        image_url = upload_image_to_cloudinary(image_file, folder='profile')
        
        if not image_url:
            return JsonResponse({'success': False, 'message': 'Failed to upload image to Cloudinary'})
        
        # Update user document
        update_data = {
            'profile_image_url': image_url,
            'updated_at': datetime.now()
        }
        
        success = update_document('users', user['id'], update_data)
        
        if success:
            try:
                log_audit(user, action_type='update', resource_type='user', resource_id=user['id'], details=f"Updated profile image for user: {user.get('name', user.get('email'))}", status='success')
            except Exception:
                pass
            # Update session
            request.session['user']['profile_image_url'] = image_url
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': 'Profile picture updated successfully',
                'image_url': image_url
            })
        else:
            return JsonResponse({'success': False, 'message': 'Failed to update profile picture'})
            
    except Exception as e:
        print(f"Error uploading profile image: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred while uploading'})


@login_required
def remove_profile_image_view(request):
    """Remove user profile image"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    user = get_user_from_session(request)
    
    try:
        from accreditation.firebase_utils import update_document
        from accreditation.cloudinary_utils import delete_image_from_cloudinary
        from datetime import datetime
        
        # Get old profile image URL if exists
        old_image_url = user.get('profile_image_url', '')
        
        # Delete old image from Cloudinary if it exists
        if old_image_url and 'cloudinary.com' in old_image_url:
            delete_image_from_cloudinary(old_image_url)
        
        # Remove image URL from user document
        update_data = {
            'profile_image_url': None,
            'updated_at': datetime.now()
        }
        
        success = update_document('users', user['id'], update_data)
        
        if success:
            try:
                log_audit(user, action_type='update', resource_type='user', resource_id=user['id'], details=f"Removed profile image for user: {user.get('name', user.get('email'))}", status='success')
            except Exception:
                pass
            # Update session
            if 'profile_image_url' in request.session['user']:
                del request.session['user']['profile_image_url']
            request.session.modified = True
            
            # Generate default avatar URL
            default_image = f"https://ui-avatars.com/api/?name={user['first_name']}+{user['last_name']}&background=4CAF50&color=fff&size=150"
            
            return JsonResponse({
                'success': True,
                'message': 'Profile picture removed successfully',
                'default_image': default_image
            })
        else:
            return JsonResponse({'success': False, 'message': 'Failed to remove profile picture'})
            
    except Exception as e:
        print(f"Error removing profile image: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred'})


@login_required
def update_personal_info_view(request):
    """Update user personal information"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    user = get_user_from_session(request)
    
    try:
        import json
        from datetime import datetime
        data = json.loads(request.body)
        
        # Check if user role is department_user or qa_admin - they cannot edit name fields
        user_role = user.get('role', '')
        if user_role in ['department_user', 'qa_admin']:
            return JsonResponse({
                'success': False, 
                'message': 'Department users and QA admins cannot modify their name. Please contact the administrator.'
            })
        
        first_name = data.get('first_name', '').strip()
        middle_name = data.get('middle_name', '').strip()
        last_name = data.get('last_name', '').strip()
        
        # Validate required fields
        if not first_name or not last_name:
            return JsonResponse({'success': False, 'message': 'First name and last name are required'})
        
        from accreditation.firebase_utils import update_document
        
        update_data = {
            'first_name': first_name,
            'middle_name': middle_name,
            'last_name': last_name,
            'updated_at': datetime.now()
        }
        
        success = update_document('users', user['id'], update_data)
        
        if success:
            # Update session
            request.session['user']['first_name'] = first_name
            request.session['user']['middle_name'] = middle_name
            request.session['user']['last_name'] = last_name
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': 'Personal information updated successfully'
            })
        else:
            return JsonResponse({'success': False, 'message': 'Failed to update information'})
            
    except Exception as e:
        print(f"Error updating personal info: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred'})


@login_required
def change_password_view(request):
    """Change user password"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    user = get_user_from_session(request)
    
    try:
        import json
        from datetime import datetime
        from accreditation.firebase_utils import get_document, update_document
        from werkzeug.security import check_password_hash, generate_password_hash
        
        data = json.loads(request.body)
        
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')
        confirm_password = data.get('confirm_password', '')
        
        # Validate required fields
        if not current_password or not new_password or not confirm_password:
            return JsonResponse({'success': False, 'message': 'All fields are required'})
        
        # Validate new passwords match
        if new_password != confirm_password:
            return JsonResponse({'success': False, 'message': 'New passwords do not match'})
        
        # Validate password length
        if len(new_password) < 8:
            return JsonResponse({'success': False, 'message': 'Password must be at least 8 characters long'})
        
        # Get user from database
        user_doc = safe_get_document('users', user['id'])
        
        if not user_doc:
            return JsonResponse({'success': False, 'message': 'User not found'})
        
        # Verify current password
        if not check_password_hash(user_doc.get('password', ''), current_password):
            return JsonResponse({'success': False, 'message': 'Current password is incorrect'})
        
        # Hash new password
        hashed_password = generate_password_hash(new_password)
        
        # Update password and last password change timestamp
        update_data = {
            'password': hashed_password,
            'last_password_change': datetime.now(),
            'updated_at': datetime.now()
        }
        
        success = update_document('users', user['id'], update_data)
        
        if success:
            return JsonResponse({
                'success': True,
                'message': 'Password changed successfully'
            })
        else:
            return JsonResponse({'success': False, 'message': 'Failed to change password'})
            
    except Exception as e:
        print(f"Error changing password: {str(e)}")
        return JsonResponse({'success': False, 'message': 'An error occurred'})


@login_required
def user_management_view(request):
    """User Management page (QA Head only) - List all users"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head
    if user.get('role') != 'qa_head':
        messages.error(request, 'Access denied. Only QA Head can access User Management.')
        return redirect('dashboard:home')
    
    # Import UserRole for role display
    from accreditation.firebase_auth import UserRole
    
    # Get filter parameters
    department_filter = request.GET.get('department', '')
    role_filter = request.GET.get('role', '')
    search_query = request.GET.get('search', '')
    
    # Get all users
    try:
        all_users = get_all_documents('users')
        
        # Apply filters
        filtered_users = all_users
        
        if department_filter:
            filtered_users = [u for u in filtered_users if u.get('department') == department_filter]
        
        if role_filter:
            filtered_users = [u for u in filtered_users if u.get('role') == role_filter]
        
        if search_query:
            search_query = search_query.lower()
            filtered_users = [
                u for u in filtered_users 
                if search_query in u.get('first_name', '').lower()
                or search_query in u.get('middle_name', '').lower()
                or search_query in u.get('last_name', '').lower()
                or search_query in u.get('email', '').lower()
            ]
        
        # Get departments and roles for filters
        departments = get_all_documents('departments')
        roles = get_all_documents('roles')
        
        # Create department mapping (code -> name)
        dept_mapping = {dept.get('code'): dept.get('name') for dept in departments if dept.get('code')}
        
        # Add department names and construct full names for users
        for user_item in filtered_users:
            # Construct full name from first, middle, and last name
            first_name = user_item.get('first_name', '')
            middle_name = user_item.get('middle_name', '')
            last_name = user_item.get('last_name', '')
            
            # Build full name
            if middle_name:
                user_item['name'] = f"{first_name} {middle_name} {last_name}"
            else:
                user_item['name'] = f"{first_name} {last_name}"
            
            # Add department name
            dept_code = user_item.get('department', '')
            user_item['department_name'] = dept_mapping.get(dept_code, dept_code if dept_code else '—')
            
            # Add role display name
            user_role = user_item.get('role', '')
            user_item['role_display'] = UserRole.get_role_display(user_role)
        
    except Exception as e:
        messages.error(request, f'Error loading users: {str(e)}')
        filtered_users = []
        departments = []
        roles = []
    
    context = {
        'active_page': 'user_management',
        'user': user,
        'users': filtered_users,
        'departments': departments,
        'roles': roles,
        'department_filter': department_filter,
        'role_filter': role_filter,
        'search_query': search_query,
    }
    return render(request, 'dashboard/user_management.html', context)


@login_required
@require_http_methods(["POST"])
def user_add_view(request):
    """Add a new user (QA Head only)"""
    user = get_user_from_session(request)
    
    if user.get('role') != 'qa_head':
        return JsonResponse({'success': False, 'message': 'Access denied.'}, status=403)
    
    try:
        form = UserManagementForm(request.POST)
        
        if form.is_valid():
            # Get form data
            first_name = form.cleaned_data['first_name']
            middle_name = form.cleaned_data.get('middle_name', '')
            last_name = form.cleaned_data['last_name']
            email_prefix = form.cleaned_data['email_prefix']
            department = form.cleaned_data['department']
            role = form.cleaned_data['role']
            status = form.cleaned_data['status']
            
            # Build full name
            if middle_name:
                full_name = f"{first_name} {middle_name} {last_name}"
            else:
                full_name = f"{first_name} {last_name}"
            
            # Build full email
            email = f"{email_prefix}@plpasig.edu.ph"
            
            # Auto-generate strong password
            generated_password = generate_strong_password()
            
            # Check if user already exists
            existing_users = query_documents('users', 'email', '==', email)
            if existing_users:
                return JsonResponse({
                    'success': False, 
                    'message': 'A user with this email already exists.'
                }, status=400)
            
            # Create user document in Firestore
            user_data = {
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'name': full_name,
                'email': email,
                'email_prefix': email_prefix,
                'department': department,
                'role': role,
                'is_active': (status == 'active'),
                'password_hash': hash_password(generated_password),  # Hash the password
                'is_password_changed': False,  # User must change password on first login
                'profile_picture': 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284240/default-profile-account-unknown-icon-black-silhouette-free-vector_jdlpve.jpg',  # Default profile picture
            }
            
            user_id = create_document('users', user_data)
            try:
                log_audit(user, action_type='create', resource_type='user', resource_id=user_id, details=f"Created new user: {full_name} ({email})", status='success')
            except Exception:
                pass
            
            return JsonResponse({
                'success': True, 
                'message': 'User created successfully!',
                'password': generated_password  # Return password for admin to give to new user
            })
        else:
            errors = dict(form.errors.items())
            return JsonResponse({
                'success': False, 
                'message': 'Please correct the errors below.',
                'errors': errors
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error creating user: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def user_edit_view(request, user_id):
    """Edit an existing user (QA Head only)"""
    user = get_user_from_session(request)
    
    if user.get('role') != 'qa_head':
        return JsonResponse({'success': False, 'message': 'Access denied.'}, status=403)
    
    try:
        # Get the user to edit
        user_to_edit = safe_get_document('users', user_id)
        if not user_to_edit:
            return JsonResponse({
                'success': False, 
                'message': 'User not found.'
            }, status=404)
        
        form = UserManagementForm(request.POST, is_edit=True)
        
        if form.is_valid():
            # Get form data
            first_name = form.cleaned_data['first_name']
            middle_name = form.cleaned_data.get('middle_name', '')
            last_name = form.cleaned_data['last_name']
            email_prefix = form.cleaned_data['email_prefix']
            department = form.cleaned_data['department']
            role = form.cleaned_data['role']
            status = form.cleaned_data['status']
            
            # Build full name
            if middle_name:
                full_name = f"{first_name} {middle_name} {last_name}"
            else:
                full_name = f"{first_name} {last_name}"
            
            # Build full email
            email = f"{email_prefix}@plpasig.edu.ph"
            
            # Update user data
            update_data = {
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'name': full_name,
                'email': email,
                'email_prefix': email_prefix,
                'department': department,
                'role': role,
                'is_active': (status == 'active'),
            }
            
            update_document('users', user_id, update_data)
            
            return JsonResponse({
                'success': True, 
                'message': 'User updated successfully!'
            })
        else:
            errors = dict(form.errors.items())
            return JsonResponse({
                'success': False, 
                'message': 'Please correct the errors below.',
                'errors': errors
            }, status=400)
            
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error updating user: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def user_delete_view(request, user_id):
    """Delete a user (QA Head only)"""
    user = get_user_from_session(request)
    
    if user.get('role') != 'qa_head':
        return JsonResponse({'success': False, 'message': 'Access denied.'}, status=403)
    
    try:
        # Get the user to delete
        user_to_delete = safe_get_document('users', user_id)
        if not user_to_delete:
            return JsonResponse({
                'success': False, 
                'message': 'User not found.'
            }, status=404)
        
        # Prevent deleting self
        if user_to_delete.get('email') == user.get('email'):
            return JsonResponse({
                'success': False, 
                'message': 'You cannot delete your own account.'
            }, status=400)
        
        # Delete the user
        delete_document('users', user_id)
        try:
            user_name = user_to_delete.get('name', user_to_delete.get('email', 'Unknown'))
            log_audit(user, action_type='delete', resource_type='user', resource_id=user_id, details=f"Deleted user: {user_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True, 
            'message': 'User deleted successfully!'
        })
        
    except Exception as e:
        # Print full traceback for debugging
        print("="*80)
        print("ERROR in user_delete_view:")
        print(traceback.format_exc())
        print("="*80)
        return JsonResponse({
            'success': False, 
            'message': f'Error deleting user: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def user_toggle_status_view(request, user_id):
    """Toggle user active status (QA Head only)"""
    user = get_user_from_session(request)
    
    if user.get('role') != 'qa_head':
        return JsonResponse({'success': False, 'message': 'Access denied.'}, status=403)
    
    try:
        data = json.loads(request.body)
        is_active = data.get('is_active', True)
        
        # Get the user to update
        user_to_update = safe_get_document('users', user_id)
        if not user_to_update:
            return JsonResponse({
                'success': False, 
                'message': 'User not found.'
            }, status=404)
        
        # Prevent deactivating self
        if user_to_update.get('email') == user.get('email') and not is_active:
            return JsonResponse({
                'success': False, 
                'message': 'You cannot deactivate your own account.'
            }, status=400)
        
        # Update the user status
        update_document('users', user_id, {'is_active': is_active})
        
        action = 'activated' if is_active else 'deactivated'
        return JsonResponse({
            'success': True, 
            'message': f'User {action} successfully!'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request data.'
        }, status=400)
    except Exception as e:
        # Print full traceback for debugging
        print("="*80)
        print("ERROR in user_toggle_status_view:")
        print(traceback.format_exc())
        print("="*80)
        return JsonResponse({
            'success': False, 
            'message': f'Error updating user status: {str(e)}'
        }, status=500)


@login_required
def user_get_view(request, user_id):
    """Get user details (QA Head only)"""
    user = get_user_from_session(request)
    
    if user.get('role') != 'qa_head':
        return JsonResponse({'success': False, 'message': 'Access denied.'}, status=403)
    
    try:
        user_data = safe_get_document('users', user_id)
        if not user_data:
            return JsonResponse({
                'success': False, 
                'message': 'User not found.'
            }, status=404)
        
        # Remove password from response
        user_data.pop('password', None)
        
        return JsonResponse({
            'success': True,
            'user': user_data
        })
        
    except Exception as e:
        # Print full traceback for debugging
        print("="*80)
        print("ERROR in user_get_view:")
        print(traceback.format_exc())
        print("="*80)
        return JsonResponse({
            'success': False, 
            'message': f'Error retrieving user: {str(e)}'
        }, status=500)


@login_required
def accreditation_settings_view(request):
    """Accreditation Settings page (QA Head and QA Admin)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied. Only QA Head and QA Admin can access Accreditation Settings.')
        return render(request, 'dashboard_base.html', {
            'active_page': 'dashboard',
            'user': user,
        })
    
    # Fetch all departments
    try:
        departments = get_all_documents('departments')
        
        for dept in departments:
            # Set default values if not present
            if 'is_archived' not in dept:
                dept['is_archived'] = False
            if 'logo_url' not in dept:
                dept['logo_url'] = ''  # Use blank as default
        
        # Sort by name
        departments.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching departments: {str(e)}")
        departments = []
    
    context = {
        'active_page': 'accreditation_settings',
        'user': user,
        'departments': departments,
    }
    return render(request, 'dashboard/accreditation_settings.html', context)


@login_required
def system_appearance_view(request):
    """System Appearance page (QA Head and QA Admin)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied. Only QA Head and QA Admin can access System Appearance.')
        return render(request, 'dashboard_base.html', {
            'active_page': 'dashboard',
            'user': user,
        })
    
    context = {
        'active_page': 'system_appearance',
        'user': user,
    }
    return render(request, 'dashboard/system_appearance.html', context)


@require_http_methods(["GET"])
def get_appearance_settings(request):
    """Get current appearance settings - PUBLIC endpoint for login page"""
    try:
        # Get settings from Firestore
        settings_docs = get_all_documents('system_settings')
        
        # Find appearance settings document
        appearance_settings = None
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                appearance_settings = doc
                break
        
        # Default settings if not found
        if not appearance_settings:
            appearance_settings = {
                'theme_color': '#4a9d4f',
                'system_title': 'PLP Accreditation System',
                'logo_url': '',
                'login_bg_url': 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284342/bg_qhybsq.jpg'
            }
        
        return JsonResponse({
            'success': True,
            'settings': appearance_settings
        })
    except Exception as e:
        print(f"Error loading appearance settings: {e}")
        # Return default settings on error
        return JsonResponse({
            'success': True,
            'settings': {
                'theme_color': '#4a9d4f',
                'system_title': 'PLP Accreditation System',
                'logo_url': '',
                'login_bg_url': 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284342/bg_qhybsq.jpg'
            }
        })


@login_required
@require_http_methods(["POST"])
def save_theme_color(request):
    """Save theme color setting"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied. Only QA Head and QA Admin can modify appearance settings.'
        }, status=403)
    
    try:
        data = json.loads(request.body)
        color = data.get('color', '#4a9d4f')
        
        # Validate color format (hex color)
        if not color.startswith('#') or len(color) != 7:
            return JsonResponse({
                'success': False,
                'message': 'Invalid color format. Please use hex color format (e.g., #4a9d4f).'
            }, status=400)
        
        # Get existing settings
        settings_docs = get_all_documents('system_settings')
        appearance_settings = None
        settings_id = None
        
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                appearance_settings = doc
                settings_id = doc.get('id')
                break
        
        # Update or create settings
        if settings_id:
            update_document('system_settings', settings_id, {
                'theme_color': color,
                'updated_at': datetime.now()
            })
        else:
            create_document('system_settings', {
                'setting_type': 'appearance',
                'theme_color': color,
                'system_title': 'PLP Accreditation System',
                'logo_url': '',
                'login_bg_url': 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284342/bg_qhybsq.jpg'
            })
        
        # Log audit trail
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='theme_color',
            details=f"Changed theme color to {color}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Theme color saved successfully.'
        })
    except Exception as e:
        # Log failed audit
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='theme_color',
            details=f"Failed to change theme color: {str(e)}",
            status='failed',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def upload_logo(request):
    """Upload system logo"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied. Only QA Head and QA Admin can modify appearance settings.'
        }, status=403)
    
    try:
        if 'logo' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No logo file provided.'
            }, status=400)
        
        logo_file = request.FILES['logo']
        
        # Validate file size (5MB max)
        if logo_file.size > 5 * 1024 * 1024:
            return JsonResponse({
                'success': False,
                'message': 'File size must be less than 5MB.'
            }, status=400)
        
        # Get existing settings to check for old logo
        settings_docs = get_all_documents('system_settings')
        appearance_settings = None
        settings_id = None
        old_logo_url = None
        
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                appearance_settings = doc
                settings_id = doc.get('id')
                old_logo_url = doc.get('logo_url')
                break
        
        # Delete old logo from Cloudinary if exists
        if old_logo_url and 'cloudinary' in old_logo_url:
            try:
                delete_image_from_cloudinary(old_logo_url)
            except:
                pass  # Continue even if deletion fails
        
        # Upload new logo to Cloudinary
        logo_url = upload_image_to_cloudinary(logo_file, folder='system/logos')
        
        # Update or create settings
        if settings_id:
            update_document('system_settings', settings_id, {
                'logo_url': logo_url,
                'updated_at': datetime.now()
            })
        else:
            create_document('system_settings', {
                'setting_type': 'appearance',
                'theme_color': '#4a9d4f',
                'system_title': 'PLP Accreditation System',
                'logo_url': logo_url,
                'login_bg_url': 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284342/bg_qhybsq.jpg'
            })
        
        # Log audit trail
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='logo',
            details=f"Uploaded new system logo: {logo_url}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Logo uploaded successfully.',
            'logo_url': logo_url
        })
    except Exception as e:
        # Log failed audit
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='logo',
            details=f"Failed to upload logo: {str(e)}",
            status='failed',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def remove_logo(request):
    """Remove system logo"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied. Only QA Head and QA Admin can modify appearance settings.'
        }, status=403)
    
    try:
        # Get existing settings
        settings_docs = get_all_documents('system_settings')
        settings_id = None
        old_logo_url = None
        
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                settings_id = doc.get('id')
                old_logo_url = doc.get('logo_url')
                break
        
        if not settings_id:
            return JsonResponse({
                'success': False,
                'message': 'No settings found.'
            }, status=404)
        
        # Delete logo from Cloudinary if exists
        if old_logo_url and 'cloudinary' in old_logo_url:
            try:
                delete_image_from_cloudinary(old_logo_url)
            except:
                pass  # Continue even if deletion fails
        
        # Update settings
        update_document('system_settings', settings_id, {
            'logo_url': '',
            'updated_at': datetime.now()
        })
        
        # Log audit trail
        log_audit(
            user=user,
            action_type='delete',
            resource_type='system_appearance',
            resource_id='logo',
            details=f"Removed system logo",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Logo removed successfully.'
        })
    except Exception as e:
        # Log failed audit
        log_audit(
            user=user,
            action_type='delete',
            resource_type='system_appearance',
            resource_id='logo',
            details=f"Failed to remove logo: {str(e)}",
            status='failed',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def save_system_title(request):
    """Save system title setting"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied. Only QA Head and QA Admin can modify appearance settings.'
        }, status=403)
    
    try:
        data = json.loads(request.body)
        title = data.get('title', 'PLP Accreditation System').strip()
        
        # Validate title
        if not title:
            return JsonResponse({
                'success': False,
                'message': 'System title cannot be empty.'
            }, status=400)
        
        if len(title) > 50:
            return JsonResponse({
                'success': False,
                'message': 'System title must be 50 characters or less.'
            }, status=400)
        
        # Get existing settings
        settings_docs = get_all_documents('system_settings')
        appearance_settings = None
        settings_id = None
        
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                appearance_settings = doc
                settings_id = doc.get('id')
                break
        
        # Update or create settings
        if settings_id:
            update_document('system_settings', settings_id, {
                'system_title': title,
                'updated_at': datetime.now()
            })
        else:
            create_document('system_settings', {
                'setting_type': 'appearance',
                'theme_color': '#4a9d4f',
                'system_title': title,
                'logo_url': '',
                'login_bg_url': 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284342/bg_qhybsq.jpg'
            })
        
        # Log audit trail
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='system_title',
            details=f"Changed system title to: {title}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'System title saved successfully.'
        })
    except Exception as e:
        # Log failed audit
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='system_title',
            details=f"Failed to change system title: {str(e)}",
            status='failed',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def upload_background(request):
    """Upload login background image"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied. Only QA Head and QA Admin can modify appearance settings.'
        }, status=403)
    
    try:
        if 'background' not in request.FILES:
            return JsonResponse({
                'success': False,
                'message': 'No background file provided.'
            }, status=400)
        
        bg_file = request.FILES['background']
        
        # Validate file size (5MB max)
        if bg_file.size > 5 * 1024 * 1024:
            return JsonResponse({
                'success': False,
                'message': 'File size must be less than 5MB.'
            }, status=400)
        
        # Get existing settings to check for old background
        settings_docs = get_all_documents('system_settings')
        appearance_settings = None
        settings_id = None
        old_bg_url = None
        
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                appearance_settings = doc
                settings_id = doc.get('id')
                old_bg_url = doc.get('login_bg_url')
                break
        
        # Delete old background from Cloudinary if exists and not default
        if old_bg_url and 'cloudinary' in old_bg_url and 'bg_qhybsq' not in old_bg_url:
            try:
                delete_image_from_cloudinary(old_bg_url)
            except:
                pass  # Continue even if deletion fails
        
        # Upload new background to Cloudinary
        bg_url = upload_image_to_cloudinary(bg_file, folder='system/backgrounds')
        
        # Update or create settings
        if settings_id:
            update_document('system_settings', settings_id, {
                'login_bg_url': bg_url,
                'updated_at': datetime.now()
            })
        else:
            create_document('system_settings', {
                'setting_type': 'appearance',
                'theme_color': '#4a9d4f',
                'system_title': 'PLP Accreditation System',
                'logo_url': '',
                'login_bg_url': bg_url
            })
        
        # Log audit trail
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='login_background',
            details=f"Uploaded new login background: {bg_url}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Background uploaded successfully.',
            'bg_url': bg_url
        })
    except Exception as e:
        # Log failed audit
        log_audit(
            user=user,
            action_type='update',
            resource_type='system_appearance',
            resource_id='login_background',
            details=f"Failed to upload background: {str(e)}",
            status='failed',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def remove_background(request):
    """Remove login background image"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied. Only QA Head and QA Admin can modify appearance settings.'
        }, status=403)
    
    try:
        # Get existing settings
        settings_docs = get_all_documents('system_settings')
        settings_id = None
        old_bg_url = None
        
        for doc in settings_docs:
            if doc.get('setting_type') == 'appearance':
                settings_id = doc.get('id')
                old_bg_url = doc.get('login_bg_url')
                break
        
        if not settings_id:
            return JsonResponse({
                'success': False,
                'message': 'No settings found.'
            }, status=404)
        
        # Delete background from Cloudinary if exists and not default
        if old_bg_url and 'cloudinary' in old_bg_url and 'bg_qhybsq' not in old_bg_url:
            try:
                delete_image_from_cloudinary(old_bg_url)
            except:
                pass  # Continue even if deletion fails
        
        # Reset to default background
        default_bg = 'https://res.cloudinary.com/dygrh6ztt/image/upload/v1761284342/bg_qhybsq.jpg'
        update_document('system_settings', settings_id, {
            'login_bg_url': default_bg,
            'updated_at': datetime.now()
        })
        
        # Log audit trail
        log_audit(
            user=user,
            action_type='delete',
            resource_type='system_appearance',
            resource_id='login_background',
            details=f"Removed login background, reset to default",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Background removed successfully.'
        })
    except Exception as e:
        # Log failed audit
        log_audit(
            user=user,
            action_type='delete',
            resource_type='system_appearance',
            resource_id='login_background',
            details=f"Failed to remove background: {str(e)}",
            status='failed',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def change_password_view(request):
    """Handle password change for first-time login"""
    user = get_user_from_session(request)
    
    try:
        data = json.loads(request.body)
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        
        # Validate passwords
        if not new_password or not confirm_password:
            return JsonResponse({
                'success': False,
                'message': 'Both password fields are required.'
            }, status=400)
        
        if new_password != confirm_password:
            return JsonResponse({
                'success': False,
                'message': 'Passwords do not match.'
            }, status=400)
        
        if len(new_password) < 6:
            return JsonResponse({
                'success': False,
                'message': 'Password must be at least 6 characters long.'
            }, status=400)
        
        # Get user from database
        user_id = user.get('id')
        users = query_documents('users', 'email', '==', user.get('email'))
        
        if not users:
            return JsonResponse({
                'success': False,
                'message': 'User not found.'
            }, status=404)
        
        user_doc = users[0]
        user_doc_id = user_doc.get('id')
        
        # Update password and mark as changed
        update_data = {
            'password_hash': hash_password(new_password),  # Hash the new password
            'is_password_changed': True
        }
        
        update_document('users', user_doc_id, update_data)
        
        # Update session
        request.session['is_password_changed'] = True
        
        return JsonResponse({
            'success': True,
            'message': 'Password changed successfully!'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request data.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error changing password: {str(e)}'
        }, status=500)


@login_required
def toast_demo_view(request):
    """Toast notification demo page"""
    context = {
        'active_page': 'toast_demo',
        'user': get_user_from_session(request),
    }
    return render(request, 'dashboard/toast_demo.html', context)


# Legacy views for backward compatibility - redirect to unified dashboard
@login_required
def qa_head_dashboard(request):
    """Redirect to unified dashboard"""
    return dashboard_home(request)


@login_required
def qa_admin_dashboard(request):
    """Redirect to unified dashboard"""
    return dashboard_home(request)


@login_required
def department_dashboard(request):
    """Redirect to unified dashboard"""
    return dashboard_home(request)


@login_required
def dashboard_view(request):
    """Redirect to unified dashboard"""
    return dashboard_home(request)


# Department CRUD Views
@login_required
@require_http_methods(["POST"])
def department_add_view(request):
    """Add a new department"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        name = request.POST.get('name', '').strip()
        logo_file = request.FILES.get('logo')
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Department name is required']
        if not logo_file:
            errors['logo'] = ['Department logo is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Upload logo to Cloudinary
        logo_url = ''
        if logo_file:
            try:
                logo_url = upload_image_to_cloudinary(logo_file, folder='departments')
                if not logo_url:
                    return JsonResponse({
                        'success': False,
                        'message': 'Error uploading logo. Please check Cloudinary configuration.',
                        'errors': {'logo': ['Upload failed. Check console for details.']}
                    })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error uploading logo: {str(e)}',
                    'errors': {'logo': [f'Upload failed: {str(e)}']}
                })
        
        # Generate department code (e.g., "CCS", "CBA")
        code = ''.join([word[0].upper() for word in name.split()[:3]])
        
        # Check if code already exists
        existing_dept = get_document('departments', code)
        if existing_dept:
            # Try adding a number suffix
            counter = 1
            while get_document('departments', f"{code}{counter}"):
                counter += 1
            code = f"{code}{counter}"
        
        # Create department
        dept_data = {
            'name': name,
            'code': code,
            'logo_url': logo_url,
            'is_archived': False,
            'is_active': True
        }
        
        create_document('departments', dept_data, code)
        try:
            log_audit(user, action_type='create', resource_type='department', resource_id=code, details=f"Created department: {name} ({code})", status='success')
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Department "{name}" created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating department: {str(e)}'
        }, status=500)


@login_required
def department_get_view(request, dept_id):
    """Get department details"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        dept = get_document('departments', dept_id)
        
        if not dept:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'department': dept
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving department: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def department_edit_view(request, dept_id):
    """Edit department"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if department exists
        dept = get_document('departments', dept_id)
        if not dept:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        logo_file = request.FILES.get('logo')
        current_logo_url = request.POST.get('current_logo_url', '').strip()
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Department name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Upload new logo if provided, otherwise keep existing
        logo_url = current_logo_url
        if logo_file:
            try:
                # Delete old logo from Cloudinary if it exists and is from our account
                if current_logo_url and 'dygrh6ztt' in current_logo_url:
                    from accreditation.cloudinary_utils import delete_image_from_cloudinary
                    delete_image_from_cloudinary(current_logo_url)
                
                # Upload new logo
                logo_url = upload_image_to_cloudinary(logo_file, folder='departments')
                if not logo_url:
                    return JsonResponse({
                        'success': False,
                        'message': 'Error uploading logo. Please check Cloudinary configuration.',
                        'errors': {'logo': ['Upload failed. Check console for details.']}
                    })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error uploading logo: {str(e)}',
                    'errors': {'logo': [f'Upload failed: {str(e)}']}
                })
        
        # Update department
        update_data = {
            'name': name,
            'logo_url': logo_url
        }
        
        update_document('departments', dept_id, update_data)
        
        return JsonResponse({
            'success': True,
            'message': f'Department "{name}" updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating department: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def department_archive_view(request, dept_id):
    """Archive/Unarchive department"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if department exists
        dept = get_document('departments', dept_id)
        if not dept:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        # Get is_archived from request body
        data = json.loads(request.body)
        is_archived = data.get('is_archived', False)
        
        # Update department
        update_document('departments', dept_id, {'is_archived': is_archived})
        
        action = 'archived' if is_archived else 'unarchived'
        return JsonResponse({
            'success': True,
            'message': f'Department "{dept.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error archiving department: {str(e)}'
        }, status=500)


def department_delete_view(request, dept_id):
    """Delete department"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if department exists
        dept = get_document('departments', dept_id)
        if not dept:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        dept_name = dept.get('name')
        logo_url = dept.get('logo_url', '')
        
        # Delete logo from Cloudinary if it exists and is from our account
        if logo_url and 'dygrh6ztt' in logo_url:
            from accreditation.cloudinary_utils import delete_image_from_cloudinary
            try:
                delete_image_from_cloudinary(logo_url)
            except Exception as e:
                print(f"Warning: Could not delete logo from Cloudinary: {str(e)}")
                # Continue with deletion even if logo deletion fails
        
        # Delete department
        delete_document('departments', dept_id)
        try:
            log_audit(user, action_type='delete', resource_type='department', resource_id=dept_id, details=f"Deleted department: {dept_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Department "{dept_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting department: {str(e)}'
        }, status=500)


def department_toggle_active_view(request, dept_id):
    """Toggle department active status"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if department exists
        dept = get_document('departments', dept_id)
        if not dept:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        # Get is_active from request body
        data = json.loads(request.body)
        is_active = data.get('is_active', False)
        
        # Update department
        update_document('departments', dept_id, {'is_active': is_active})
        
        action = 'activated' if is_active else 'deactivated'
        return JsonResponse({
            'success': True,
            'message': f'Department "{dept.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating department status: {str(e)}'
        }, status=500)


# ============================================
# PROGRAMS MANAGEMENT VIEWS
# ============================================

@login_required
def department_programs_view(request, dept_id):
    """View programs for a specific department"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied. Only QA Head and QA Admin can access Programs.')
        return render(request, 'dashboard_base.html', {
            'active_page': 'dashboard',
            'user': user,
        })
    
    # Get department details
    try:
        department = get_document('departments', dept_id)
        if not department:
            messages.error(request, 'Department not found.')
            return redirect('dashboard:accreditation_settings')
        
        # Fetch programs for this department
        all_programs = get_all_documents('programs')
        programs = [p for p in all_programs if p.get('department_id') == dept_id]
        
        # Set default values if not present
        for prog in programs:
            if 'is_archived' not in prog:
                prog['is_archived'] = False
            if 'is_active' not in prog:
                prog['is_active'] = True
        
        # Sort by code
        programs.sort(key=lambda x: x.get('code', ''))
        
    except Exception as e:
        print(f"Error fetching programs: {str(e)}")
        programs = []
        department = {'name': 'Unknown Department', 'code': dept_id}
    
    context = {
        'active_page': 'accreditation_settings',
        'user': user,
        'department': department,
        'programs': programs,
    }
    return render(request, 'dashboard/department_programs.html', context)


@login_required
@require_http_methods(["POST"])
def program_add_view(request, dept_id):
    """Add a new program to a department"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Verify department exists
        department = get_document('departments', dept_id)
        if not department:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        code = request.POST.get('code', '').strip().upper()
        name = request.POST.get('name', '').strip()
        
        # Validation
        errors = {}
        if not code:
            errors['code'] = ['Program code is required']
        if not name:
            errors['name'] = ['Program name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Check if code already exists
        existing_prog = get_document('programs', code)
        if existing_prog:
            return JsonResponse({
                'success': False,
                'message': 'Program code already exists',
                'errors': {'code': ['This program code is already in use']}
            })
        
        # Create program
        prog_data = {
            'code': code,
            'name': name,
            'department_id': dept_id,
            'is_archived': False,
            'is_active': True
        }
        
        create_document('programs', prog_data, code)
        try:
            log_audit(user, action_type='create', resource_type='program', resource_id=code, details=f"Created program: {name} ({code})", status='success')
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Program "{name}" created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating program: {str(e)}'
        }, status=500)


@login_required
def program_get_view(request, dept_id, prog_id):
    """Get program details"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        prog = get_document('programs', prog_id)
        
        if not prog:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'program': prog
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving program: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def program_edit_view(request, dept_id, prog_id):
    """Edit program"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if program exists
        prog = get_document('programs', prog_id)
        if not prog:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Program name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Update program (code cannot be changed)
        update_data = {
            'name': name
        }
        
        update_document('programs', prog_id, update_data)
        
        return JsonResponse({
            'success': True,
            'message': f'Program "{name}" updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating program: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def program_archive_view(request, dept_id, prog_id):
    """Archive/Unarchive program"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if program exists
        prog = get_document('programs', prog_id)
        if not prog:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        # Get is_archived from request body
        data = json.loads(request.body)
        is_archived = data.get('is_archived', False)
        
        # Update program
        update_document('programs', prog_id, {'is_archived': is_archived})
        
        action = 'archived' if is_archived else 'unarchived'
        return JsonResponse({
            'success': True,
            'message': f'Program "{prog.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error archiving program: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def program_delete_view(request, dept_id, prog_id):
    """Delete program"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if program exists
        prog = get_document('programs', prog_id)
        if not prog:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        prog_name = prog.get('name')
        
        # Delete program
        delete_document('programs', prog_id)
        try:
            log_audit(user, action_type='delete', resource_type='program', resource_id=prog_id, details=f"Deleted program: {prog_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Program "{prog_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting program: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def program_toggle_active_view(request, dept_id, prog_id):
    """Toggle program active status"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if program exists
        prog = get_document('programs', prog_id)
        if not prog:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        # Get is_active from request body
        data = json.loads(request.body)
        is_active = data.get('is_active', False)
        
        # Update program
        update_document('programs', prog_id, {'is_active': is_active})
        
        action = 'activated' if is_active else 'deactivated'
        return JsonResponse({
            'success': True,
            'message': f'Program "{prog.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating program status: {str(e)}'
        }, status=500)


# =============================================
# ACCREDITATION TYPES MANAGEMENT VIEWS
# =============================================

@login_required
def program_types_view(request, dept_id, prog_id):
    """View accreditation types for a specific program"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied. Only QA Head and QA Admin can access Accreditation Types.')
        return render(request, 'dashboard_base.html', {
            'active_page': 'dashboard',
            'user': user,
        })
    
    # Get department and program details
    try:
        department = get_document('departments', dept_id)
        if not department:
            messages.error(request, 'Department not found.')
            return redirect('dashboard:accreditation_settings')
        
        program = get_document('programs', prog_id)
        if not program:
            messages.error(request, 'Program not found.')
            return redirect('dashboard:department_programs', dept_id=dept_id)
        
        # Fetch accreditation types for this program
        all_types = get_all_documents('accreditation_types')
        types = [t for t in all_types if t.get('program_id') == prog_id]
        
        # Set default values if not present
        for type_item in types:
            if 'is_archived' not in type_item:
                type_item['is_archived'] = False
            if 'is_active' not in type_item:
                type_item['is_active'] = True
            if 'logo_url' not in type_item:
                type_item['logo_url'] = ''
        
        # Sort by name
        types.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching accreditation types: {str(e)}")
        types = []
        program = {'name': 'Unknown Program', 'code': prog_id}
        department = {'name': 'Unknown Department', 'code': dept_id}
    
    context = {
        'active_page': 'accreditation_settings',
        'user': user,
        'department': department,
        'program': program,
        'types': types,
    }
    return render(request, 'dashboard/program_types.html', context)


@login_required
@require_http_methods(["POST"])
def accreditation_type_add_view(request, dept_id, prog_id):
    """Add a new accreditation type to a program"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Verify program exists
        program = get_document('programs', prog_id)
        if not program:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        logo_file = request.FILES.get('logo')
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Type name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Generate unique ID for the type
        import uuid
        type_id = str(uuid.uuid4())
        
        # Upload logo if provided
        logo_url = ''
        if logo_file:
            try:
                logo_url = upload_image_to_cloudinary(logo_file, 'departments')
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error uploading logo: {str(e)}'
                }, status=500)
        
        # Create accreditation type
        type_data = {
            'id': type_id,
            'name': name,
            'program_id': prog_id,
            'logo_url': logo_url,
            'is_archived': False,
            'is_active': True
        }
        
        create_document('accreditation_types', type_data, type_id)
        try:
            log_audit(user, action_type='create', resource_type='accreditation_type', resource_id=type_id, details=f"Created accreditation type: {name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Accreditation type "{name}" created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating accreditation type: {str(e)}'
        }, status=500)


@login_required
def accreditation_type_get_view(request, dept_id, prog_id, type_id):
    """Get accreditation type details"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        type_item = get_document('accreditation_types', type_id)
        if not type_item:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'type': type_item
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving accreditation type: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def accreditation_type_edit_view(request, dept_id, prog_id, type_id):
    """Edit an existing accreditation type"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if type exists
        type_item = get_document('accreditation_types', type_id)
        if not type_item:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        logo_file = request.FILES.get('logo')
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Type name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Prepare update data
        update_data = {
            'name': name
        }
        
        # Upload new logo if provided and delete old logo
        if logo_file:
            try:
                # Delete old logo if exists
                old_logo_url = type_item.get('logo_url')
                if old_logo_url:
                    delete_image_from_cloudinary(old_logo_url)
                
                logo_url = upload_image_to_cloudinary(logo_file, 'departments')
                update_data['logo_url'] = logo_url
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Error uploading logo: {str(e)}'
                }, status=500)
        
        # Update type
        update_document('accreditation_types', type_id, update_data)
        
        return JsonResponse({
            'success': True,
            'message': f'Accreditation type "{name}" updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating accreditation type: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def accreditation_type_archive_view(request, dept_id, prog_id, type_id):
    """Archive accreditation type"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if type exists
        type_item = get_document('accreditation_types', type_id)
        if not type_item:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        # Get is_archived from request body
        data = json.loads(request.body)
        is_archived = data.get('is_archived', False)
        
        # Update type
        update_document('accreditation_types', type_id, {'is_archived': is_archived})
        
        action = 'archived' if is_archived else 'unarchived'
        return JsonResponse({
            'success': True,
            'message': f'Accreditation type "{type_item.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error archiving accreditation type: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def accreditation_type_delete_view(request, dept_id, prog_id, type_id):
    """Delete accreditation type"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if type exists
        type_item = get_document('accreditation_types', type_id)
        if not type_item:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        type_name = type_item.get('name')
        
        # Delete logo from Cloudinary if exists
        logo_url = type_item.get('logo_url')
        if logo_url:
            try:
                delete_image_from_cloudinary(logo_url)
            except Exception as e:
                print(f"Error deleting logo from Cloudinary: {str(e)}")
        
        # Delete type
        delete_document('accreditation_types', type_id)
        try:
            log_audit(user, action_type='delete', resource_type='accreditation_type', resource_id=type_id, details=f"Deleted accreditation type: {type_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Accreditation type "{type_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting accreditation type: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def accreditation_type_toggle_active_view(request, dept_id, prog_id, type_id):
    """Toggle accreditation type active status"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if type exists
        type_item = get_document('accreditation_types', type_id)
        if not type_item:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        # Get is_active from request body
        data = json.loads(request.body)
        is_active = data.get('is_active', False)
        
        # Update type
        update_document('accreditation_types', type_id, {'is_active': is_active})
        
        action = 'activated' if is_active else 'deactivated'
        return JsonResponse({
            'success': True,
            'message': f'Accreditation type "{type_item.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating accreditation type status: {str(e)}'
        }, status=500)


# =============================================
# AREAS MANAGEMENT VIEWS
# =============================================

@login_required
def type_areas_view(request, dept_id, prog_id, type_id):
    """View areas for a specific accreditation type"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied. Only QA Head and QA Admin can access Areas.')
        return render(request, 'dashboard_base.html', {
            'active_page': 'dashboard',
            'user': user,
        })
    
    # Get department, program, and type details
    try:
        department = get_document('departments', dept_id)
        if not department:
            messages.error(request, 'Department not found.')
            return redirect('dashboard:accreditation_settings')
        
        program = get_document('programs', prog_id)
        if not program:
            messages.error(request, 'Program not found.')
            return redirect('dashboard:department_programs', dept_id=dept_id)
        
        accreditation_type = get_document('accreditation_types', type_id)
        if not accreditation_type:
            messages.error(request, 'Accreditation type not found.')
            return redirect('dashboard:program_types', dept_id=dept_id, prog_id=prog_id)
        
        # Fetch areas for this type
        all_areas = get_all_documents('areas')
        areas = [m for m in all_areas if m.get('accreditation_type_id') == type_id]
        
        # Set default values if not present
        for area in areas:
            if 'is_archived' not in area:
                area['is_archived'] = False
            if 'is_active' not in area:
                area['is_active'] = True
        
        # Sort by name
        areas.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching areas: {str(e)}")
        areas = []
        accreditation_type = {'name': 'Unknown Type', 'id': type_id}
        program = {'name': 'Unknown Program', 'code': prog_id}
        department = {'name': 'Unknown Department', 'code': dept_id}
    
    context = {
        'active_page': 'accreditation_settings',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'areas': areas,
    }
    return render(request, 'dashboard/type_areas.html', context)


@login_required
@require_http_methods(["POST"])
def area_add_view(request, dept_id, prog_id, type_id):
    """Add a new area to an accreditation type"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Verify type exists
        accreditation_type = get_document('accreditation_types', type_id)
        if not accreditation_type:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Area name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Generate unique ID for the area
        import uuid
        area_id = str(uuid.uuid4())
        
        # Create area
        module_data = {
            'id': area_id,
            'name': name,
            'accreditation_type_id': type_id,
            'is_archived': False,
            'is_active': True
        }
        
        create_document('areas', module_data, area_id)
        try:
            log_audit(user, action_type='create', resource_type='area', resource_id=area_id, details=f"Created area: {name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Area "{name}" created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating area: {str(e)}'
        }, status=500)


@login_required
def area_get_view(request, dept_id, prog_id, type_id, area_id):
    """Get area details"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'area': area
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving area: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def area_edit_view(request, dept_id, prog_id, type_id, area_id):
    """Edit an existing area"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if area exists
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Area name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Update area
        update_document('areas', area_id, {'name': name})
        
        return JsonResponse({
            'success': True,
            'message': f'Area "{name}" updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating area: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def area_archive_view(request, dept_id, prog_id, type_id, area_id):
    """Archive area"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if area exists
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        # Get is_archived from request body
        data = json.loads(request.body)
        is_archived = data.get('is_archived', False)
        
        # Update area
        update_document('areas', area_id, {'is_archived': is_archived})
        
        action = 'archived' if is_archived else 'unarchived'
        return JsonResponse({
            'success': True,
            'message': f'Area "{area.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error archiving area: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def area_delete_view(request, dept_id, prog_id, type_id, area_id):
    """Delete area"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if area exists
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        module_name = area.get('name')
        
        # Delete area
        delete_document('areas', area_id)
        try:
            log_audit(user, action_type='delete', resource_type='area', resource_id=area_id, details=f"Deleted area: {module_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Area "{module_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting area: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def area_toggle_active_view(request, dept_id, prog_id, type_id, area_id):
    """Toggle area active status"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if area exists
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        # Get is_active from request body
        data = json.loads(request.body)
        is_active = data.get('is_active', False)
        
        # Update area
        update_document('areas', area_id, {'is_active': is_active})
        
        action = 'activated' if is_active else 'deactivated'
        return JsonResponse({
            'success': True,
            'message': f'Area "{area.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating area status: {str(e)}'
        }, status=500)


# =============================================
# CHECKLISTS MANAGEMENT VIEWS
# =============================================

@login_required
def area_checklists_view(request, dept_id, prog_id, type_id, area_id):
    """View checklists for a specific area"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied. Only QA Head and QA Admin can access Checklists.')
        return render(request, 'dashboard_base.html', {
            'active_page': 'dashboard',
            'user': user,
        })
    
    # Get department, program, type, and area details
    try:
        department = get_document('departments', dept_id)
        if not department:
            messages.error(request, 'Department not found.')
            return redirect('dashboard:accreditation_settings')
        
        program = get_document('programs', prog_id)
        if not program:
            messages.error(request, 'Program not found.')
            return redirect('dashboard:department_programs', dept_id=dept_id)
        
        accreditation_type = get_document('accreditation_types', type_id)
        if not accreditation_type:
            messages.error(request, 'Accreditation type not found.')
            return redirect('dashboard:program_types', dept_id=dept_id, prog_id=prog_id)
        
        area = get_document('areas', area_id)
        if not area:
            messages.error(request, 'Area not found.')
            return redirect('dashboard:type_areas', dept_id=dept_id, prog_id=prog_id, type_id=type_id)
        
        # Fetch checklists for this area
        all_checklists = get_all_documents('checklists')
        checklists = [c for c in all_checklists if c.get('area_id') == area_id]
        
        # Get all documents to calculate progress for each checklist
        all_documents = get_all_documents('documents')
        
        # Set default values if not present and add progress
        for checklist in checklists:
            if 'is_archived' not in checklist:
                checklist['is_archived'] = False
            if 'is_active' not in checklist:
                checklist['is_active'] = True
            
            # Calculate progress based on required documents
            checklist_id = checklist.get('id')
            required_docs = [
                doc for doc in all_documents 
                if doc.get('checklist_id') == checklist_id 
                and doc.get('is_required', False)
                and not doc.get('is_archived', False)
                and doc.get('status') == 'approved'
            ]
            # Progress is 100% if there's at least 1 required document, otherwise 0%
            checklist['progress'] = 100 if len(required_docs) > 0 else 0
        
        # Sort by checklist number (extract number from "Checklist X")
        def get_checklist_number(checklist):
            name = checklist.get('name', '')
            try:
                # Extract number from "Checklist X" format
                if 'Checklist' in name:
                    return int(name.replace('Checklist', '').strip())
                return 0
            except:
                return 0
        
        checklists.sort(key=get_checklist_number)
        
    except Exception as e:
        print(f"Error fetching checklists: {str(e)}")
        checklists = []
        area = {'name': 'Unknown Area', 'id': area_id}
        accreditation_type = {'name': 'Unknown Type', 'id': type_id}
        program = {'name': 'Unknown Program', 'code': prog_id}
        department = {'name': 'Unknown Department', 'code': dept_id}
    
    context = {
        'active_page': 'accreditation_settings',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'area': area,
        'checklists': checklists,
    }
    return render(request, 'dashboard/area_checklists.html', context)


@login_required
@require_http_methods(["POST"])
def checklist_add_view(request, dept_id, prog_id, type_id, area_id):
    """Add a new checklist to a area"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Verify area exists
        area = get_document('areas', area_id)
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Checklist name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Generate unique ID for the checklist
        import uuid
        checklist_id = str(uuid.uuid4())
        
        # Create checklist
        checklist_data = {
            'id': checklist_id,
            'name': name,
            'area_id': area_id,
            'is_archived': False,
            'is_active': True
        }
        
        create_document('checklists', checklist_data, checklist_id)
        try:
            log_audit(user, action_type='create', resource_type='checklist', resource_id=checklist_id, details=f"Created checklist: {name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Checklist "{name}" created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating checklist: {str(e)}'
        }, status=500)


@login_required
def checklist_get_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """Get checklist details"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        checklist = get_document('checklists', checklist_id)
        if not checklist:
            return JsonResponse({
                'success': False,
                'message': 'Checklist not found'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'checklist': checklist
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving checklist: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def checklist_edit_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """Edit an existing checklist"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if checklist exists
        checklist = get_document('checklists', checklist_id)
        if not checklist:
            return JsonResponse({
                'success': False,
                'message': 'Checklist not found'
            }, status=404)
        
        name = request.POST.get('name', '').strip()
        
        # Validation
        errors = {}
        if not name:
            errors['name'] = ['Checklist name is required']
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': 'Validation failed',
                'errors': errors
            })
        
        # Update checklist
        update_document('checklists', checklist_id, {'name': name})
        
        return JsonResponse({
            'success': True,
            'message': f'Checklist "{name}" updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating checklist: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def checklist_archive_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """Archive checklist"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if checklist exists
        checklist = get_document('checklists', checklist_id)
        if not checklist:
            return JsonResponse({
                'success': False,
                'message': 'Checklist not found'
            }, status=404)
        
        # Get is_archived from request body
        data = json.loads(request.body)
        is_archived = data.get('is_archived', False)
        
        # Update checklist
        update_document('checklists', checklist_id, {'is_archived': is_archived})
        
        action = 'archived' if is_archived else 'unarchived'
        return JsonResponse({
            'success': True,
            'message': f'Checklist "{checklist.get("name")}" {action} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error archiving checklist: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def checklist_delete_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """Delete checklist"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if checklist exists
        checklist = get_document('checklists', checklist_id)
        if not checklist:
            return JsonResponse({
                'success': False,
                'message': 'Checklist not found'
            }, status=404)
        
        checklist_name = checklist.get('name')
        
        # Delete checklist
        delete_document('checklists', checklist_id)
        try:
            log_audit(user, action_type='delete', resource_type='checklist', resource_id=checklist_id, details=f"Deleted checklist: {checklist_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Checklist "{checklist_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting checklist: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def checklist_toggle_active_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """Toggle checklist active status"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Check if checklist exists
        checklist = get_document('checklists', checklist_id)
        if not checklist:
            return JsonResponse({
                'success': False,
                'message': 'Checklist not found'
            }, status=404)
        
        # Get is_active from request body
        data = json.loads(request.body)
        is_active = data.get('is_active', False)
        
        # Update checklist
        update_document('checklists', checklist_id, {'is_active': is_active})
        
        action = 'activated' if is_active else 'deactivated'
        return JsonResponse({
            'success': True,
            'message': f'Checklist "{checklist.get("name")}" {action} successfully'
        })
        
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating checklist status: {str(e)}'
        }, status=500)


# ============================================
# ARCHIVE API VIEWS
# ============================================

@login_required
def archive_api_departments(request):
    """Get archived departments"""
    user = get_user_from_session(request)
    
    try:
        # Get all archived departments
        archived_depts = query_documents('departments', 'is_archived', '==', True)
        
        return JsonResponse({
            'success': True,
            'items': archived_depts
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving archived departments: {str(e)}'
        }, status=500)


@login_required
def archive_api_programs(request):
    """Get archived programs"""
    user = get_user_from_session(request)
    
    try:
        # Get all archived programs
        archived_programs = query_documents('programs', 'is_archived', '==', True)
        
        # Fetch department info for each program
        for program in archived_programs:
            if program.get('department_id'):
                dept = get_document('departments', program['department_id'])
                if dept:
                    program['department_code'] = dept.get('code', 'N/A')
                    program['department_name'] = dept.get('name', 'Unknown Department')
                    program['department_logo'] = dept.get('logo_url', '')
        
        return JsonResponse({
            'success': True,
            'items': archived_programs
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving archived programs: {str(e)}'
        }, status=500)


@login_required
def archive_api_types(request):
    """Get archived accreditation types"""
    user = get_user_from_session(request)
    
    try:
        # Get all archived types
        archived_types = query_documents('accreditation_types', 'is_archived', '==', True)
        
        # Fetch department and program info for each type
        for type_item in archived_types:
            if type_item.get('program_id'):
                program = get_document('programs', type_item['program_id'])
                if program:
                    type_item['program_code'] = program.get('code', 'N/A')
                    type_item['program_name'] = program.get('name', 'Unknown Program')
                    
                    # Get department info
                    if program.get('department_id'):
                        dept = get_document('departments', program['department_id'])
                        if dept:
                            type_item['department_code'] = dept.get('code', 'N/A')
                            type_item['department_name'] = dept.get('name', 'Unknown Department')
        
        return JsonResponse({
            'success': True,
            'items': archived_types
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving archived types: {str(e)}'
        }, status=500)


@login_required
def archive_api_areas(request):
    """Get archived areas"""
    user = get_user_from_session(request)
    
    try:
        # Get all archived areas
        archived_areas = query_documents('areas', 'is_archived', '==', True)
        
        # Fetch type, program, and department info for each area
        for area in archived_areas:
            # Initialize with defaults
            area['department_code'] = 'N/A'
            area['program_code'] = 'N/A'
            area['type_name'] = 'Unknown Type'
            
            type_id = area.get('type_id') or area.get('accreditation_type_id')
            if type_id:
                type_item = get_document('accreditation_types', type_id)
                if type_item:
                    area['type_name'] = type_item.get('name', 'Unknown Type')
                    
                    # Get program info
                    program_id = type_item.get('program_id')
                    if program_id:
                        program = get_document('programs', program_id)
                        if program:
                            area['program_code'] = program.get('code', 'N/A')
                            area['program_name'] = program.get('name', 'Unknown Program')
                            
                            # Get department info
                            dept_id = program.get('department_id')
                            if dept_id:
                                dept = get_document('departments', dept_id)
                                if dept:
                                    area['department_code'] = dept.get('code', 'N/A')
                                    area['department_name'] = dept.get('name', 'Unknown Department')
        
        return JsonResponse({
            'success': True,
            'items': archived_areas
        })
        
    except Exception as e:
        import traceback
        print(f"Error in archive_api_areas: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving archived areas: {str(e)}'
        }, status=500)


@login_required
def archive_api_checklists(request):
    """Get archived checklists"""
    user = get_user_from_session(request)
    
    try:
        # Get all archived checklists
        archived_checklists = query_documents('checklists', 'is_archived', '==', True)
        
        # Fetch area, type, program, and department info for each checklist
        for checklist in archived_checklists:
            # Initialize with defaults
            checklist['department_code'] = 'N/A'
            checklist['program_code'] = 'N/A'
            checklist['type_name'] = 'Unknown Type'
            checklist['module_name'] = 'Unknown Area'
            
            area_id = checklist.get('area_id')
            if area_id:
                area = get_document('areas', area_id)
                if area:
                    checklist['module_name'] = area.get('name', 'Unknown Area')
                    
                    # Get type info
                    type_id = area.get('type_id') or area.get('accreditation_type_id')
                    if type_id:
                        type_item = get_document('accreditation_types', type_id)
                        if type_item:
                            checklist['type_name'] = type_item.get('name', 'Unknown Type')
                            
                            # Get program info
                            program_id = type_item.get('program_id')
                            if program_id:
                                program = get_document('programs', program_id)
                                if program:
                                    checklist['program_code'] = program.get('code', 'N/A')
                                    checklist['program_name'] = program.get('name', 'Unknown Program')
                                    
                                    # Get department info
                                    dept_id = program.get('department_id')
                                    if dept_id:
                                        dept = get_document('departments', dept_id)
                                        if dept:
                                            checklist['department_code'] = dept.get('code', 'N/A')
                                            checklist['department_name'] = dept.get('name', 'Unknown Department')
        
        return JsonResponse({
            'success': True,
            'items': archived_checklists
        })
        
    except Exception as e:
        import traceback
        print(f"Error in archive_api_checklists: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'Error retrieving archived checklists: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def archive_api_unarchive(request, item_type, item_id):
    """Unarchive any item"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Map item type to collection name
        collection_map = {
            'departments': 'departments',
            'programs': 'programs',
            'types': 'accreditation_types',
            'areas': 'areas',
            'checklists': 'checklists'
        }
        
        collection = collection_map.get(item_type)
        if not collection:
            return JsonResponse({
                'success': False,
                'message': 'Invalid item type'
            }, status=400)
        
        # Get the item
        item = get_document(collection, item_id)
        if not item:
            return JsonResponse({
                'success': False,
                'message': 'Item not found'
            }, status=404)
        
        # Unarchive the item
        update_document(collection, item_id, {'is_archived': False})
        
        item_name = item.get('name') or item.get('code', 'Item')
        
        return JsonResponse({
            'success': True,
            'message': f'"{item_name}" unarchived successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error unarchiving item: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def archive_api_delete(request, item_type, item_id):
    """Permanently delete archived item"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Map item type to collection name
        collection_map = {
            'departments': 'departments',
            'programs': 'programs',
            'types': 'accreditation_types',
            'areas': 'areas',
            'checklists': 'checklists'
        }
        
        collection = collection_map.get(item_type)
        if not collection:
            return JsonResponse({
                'success': False,
                'message': 'Invalid item type'
            }, status=400)
        
        # Get the item
        item = get_document(collection, item_id)
        if not item:
            return JsonResponse({
                'success': False,
                'message': 'Item not found'
            }, status=404)
        
        item_name = item.get('name') or item.get('code', 'Item')
        
        # Delete image if it's a department or type
        if item_type in ['departments', 'types']:
            logo_url = item.get('logo_url')
            if logo_url:
                try:
                    delete_image_from_cloudinary(logo_url)
                except Exception as e:
                    print(f"Error deleting image from Cloudinary: {str(e)}")
        
        # Delete the item
        delete_document(collection, item_id)
        
        
        return JsonResponse({
            'success': True,
            'message': f'"{item_name}" permanently deleted'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting item: {str(e)}'
        }, status=500)


# ============================================
# DOCUMENT MANAGEMENT VIEWS
# ============================================

@login_required
def checklist_documents_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """View documents for a specific checklist"""
    user = get_user_from_session(request)
    
    try:
        # Get breadcrumb data
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        accreditation_type = get_document('accreditation_types', type_id)
        area = get_document('areas', area_id)
        checklist = get_document('checklists', checklist_id)
        
        if not all([department, program, accreditation_type, area, checklist]):
            messages.error(request, 'Data not found.')
            return redirect('dashboard:accreditation')
        
        # Get documents for this checklist
        all_documents = get_all_documents('documents')
        documents = [
            doc for doc in all_documents 
            if doc.get('checklist_id') == checklist_id 
            and not doc.get('is_archived', False)
        ]
        
        # Separate required and additional documents
        required_documents = [doc for doc in documents if doc.get('is_required', False)]
        additional_documents = [doc for doc in documents if not doc.get('is_required', False)]
        
        # Sort required documents by creation date (most recent first)
        required_documents.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        # Sort additional documents by creation date
        additional_documents.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        # Get the most recent required document for backward compatibility
        required_document = required_documents[0] if required_documents else None
        
    except Exception as e:
        print(f"Error fetching documents: {str(e)}")
        messages.error(request, 'Error fetching documents.')
        return redirect('dashboard:accreditation')
    
    context = {
        'active_page': 'accreditation',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'area': area,
        'checklist': checklist,
        'required_document': required_document,
        'required_documents': required_documents,
        'additional_documents': additional_documents,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'type_id': type_id,
        'area_id': area_id,
        'checklist_id': checklist_id,
    }
    return render(request, 'dashboard/checklist_documents.html', context)


@login_required
@require_http_methods(["POST"])
def document_add_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """Add documents to a checklist"""
    from accreditation.cloudinary_utils import upload_document_to_cloudinary
    import uuid
    
    user = get_user_from_session(request)
    
    try:
        # Get IDs from dropdown selections (they override URL parameters)
        selected_dept_id = request.POST.get('department_id', dept_id)
        selected_prog_id = request.POST.get('program_id', prog_id)
        selected_type_id = request.POST.get('accreditation_type_id', type_id)
        selected_area_id = request.POST.get('area_id', area_id)
        selected_checklist_id = request.POST.get('checklist_id', checklist_id)
        
        # Handle multiple required documents
        required_files = request.FILES.getlist('required_documents[]')
        required_names = request.POST.getlist('required_document_names[]')
        
        if not required_files or len(required_files) == 0:
            return JsonResponse({
                'success': False,
                'message': 'At least one required document is required'
            }, status=400)
        
        if len(required_names) != len(required_files):
            return JsonResponse({
                'success': False,
                'message': 'Document names and files count mismatch'
            }, status=400)
        
        # Validate and upload required documents
        allowed_formats = ['doc', 'docx']
        uploaded_count = 0
        
        for idx, required_file in enumerate(required_files):
            required_doc_name = required_names[idx].strip() if idx < len(required_names) else ''
            
            if not required_doc_name:
                return JsonResponse({
                    'success': False,
                    'message': f'Document name is required for file: {required_file.name}'
                }, status=400)
            
            if not required_file.name:
                continue
            
            # Validate required document format
            file_ext = required_file.name.split('.')[-1].lower()
            
            if file_ext not in allowed_formats:
                return JsonResponse({
                    'success': False,
                    'message': f'Invalid file format for {required_file.name}. Allowed formats: {", ".join(allowed_formats)}'
                }, status=400)
            
            # ============================================
            # VALIDATE DOCUMENT HEADER AGAINST TEMPLATE
            # ============================================
            if file_ext in ['doc', 'docx']:
                is_valid, error_message = validate_document_header(required_file)
                if not is_valid:
                    return JsonResponse({
                        'success': False,
                        'message': f'Header validation failed for {required_file.name}: {error_message}'
                    }, status=400)
                
                # Reset file pointer after validation
                required_file.seek(0)
            
            # ============================================
            # ADD FOOTER WATERMARK TO WORD DOCUMENTS
            # ============================================
            file_to_upload = required_file
            upload_filename = required_file.name
            
            if file_ext in ['doc', 'docx']:
                try:
                    from accreditation.document_footer_utils import (
                        add_footer_code_to_document, 
                        count_existing_required_documents
                    )
                    
                    # Get metadata for footer code
                    # Fetch department, program, type, area, and checklist info
                    dept = get_document('departments', selected_dept_id)
                    program = get_document('programs', selected_prog_id)
                    accred_type = get_document('accreditation_types', selected_type_id)
                    area = get_document('areas', selected_area_id)
                    checklist = get_document('checklists', selected_checklist_id)
                    
                    # Count existing required documents to get next number
                    doc_number = count_existing_required_documents(
                        selected_dept_id, 
                        selected_prog_id, 
                        selected_type_id, 
                        selected_area_id, 
                        selected_checklist_id
                    )
                    
                    # Prepare metadata
                    metadata = {
                        'accreditation_type_name': accred_type.get('name', '') if accred_type else '',
                        'dept_code': dept.get('code', 'DEPT') if dept else 'DEPT',
                        'program_code': program.get('code', 'PROG') if program else 'PROG',
                        'area_name': area.get('name', 'Area 1') if area else 'Area 1',
                        'checklist_name': checklist.get('name', 'Checklist 1') if checklist else 'Checklist 1',
                        'document_number': doc_number,
                        'checklist_id': selected_checklist_id,  # Add checklist_id for versioning
                        'document_name': required_doc_name  # Add document name for versioning
                    }
                    
                    # Add footer to document and get modified file with new filename
                    file_to_upload, upload_filename = add_footer_code_to_document(
                        required_file, 
                        metadata, 
                        required_file.name
                    )
                    
                    # Update file extension to 'docx' since we saved it as .docx
                    file_ext = 'docx'
                    
                except Exception as footer_error:
                    print(f"Warning: Failed to add footer to document: {str(footer_error)}")
                    # Continue with original file if footer addition fails
                    required_file.seek(0)
                    file_to_upload = required_file
                    upload_filename = required_file.name
            
            # Upload required document (with footer if Word document)
            file_url = upload_document_to_cloudinary(
                file_to_upload, 
                folder=f'documents/{selected_dept_id}/{selected_prog_id}/{selected_type_id}/{selected_area_id}/{selected_checklist_id}',
                filename=upload_filename
            )
            
            if not file_url:
                return JsonResponse({
                    'success': False,
                    'message': f'Failed to upload document: {required_file.name}'
                }, status=500)
            
            # Create required document record
            from datetime import datetime
            required_doc_data = {
                'department_id': selected_dept_id,
                'program_id': selected_prog_id,
                'accreditation_type_id': selected_type_id,
                'area_id': selected_area_id,
                'checklist_id': selected_checklist_id,
                'name': required_doc_name,
                'file_url': file_url,
                'format': file_ext,
                'uploaded_by': user.get('email', 'Unknown'),
                'uploaded_at': datetime.now().isoformat(),
                'is_required': True,
                'status': 'submitted',
                'comment': '',  # Comment will be added by disapprover later
                'is_active': True,
                'is_archived': False,
                'user_id': user.get('id'),  # Add user_id for notifications
            }
            
            doc_id = str(uuid.uuid4())
            create_document('documents', required_doc_data, doc_id)
            try:
                log_audit(user, action_type='document_upload', resource_type='document', resource_id=doc_id, details=f"Uploaded required document: {required_doc_name}", status='success', ip=get_client_ip(request))
            except Exception:
                pass
            
            # Create notification for document upload
            try:
                from accreditation.notification_utils import notify_document_upload
                
                print(f"[NOTIFICATION DEBUG] Creating notification for document: {doc_id}")
                
                # Get names for notification
                department = get_document('departments', selected_dept_id)
                program = get_document('programs', selected_prog_id)
                acc_type = get_document('accreditation_types', selected_type_id)
                area = get_document('areas', selected_area_id)
                checklist = get_document('checklists', selected_checklist_id)
                
                print(f"[NOTIFICATION DEBUG] Department: {department.get('name') if department else 'None'}")
                print(f"[NOTIFICATION DEBUG] User: {user.get('email')}")
                
                notify_document_upload(
                    document_id=doc_id,
                    document_name=required_doc_name,
                    department_name=department.get('name', 'Unknown') if department else 'Unknown',
                    program_name=program.get('name', 'Unknown') if program else 'Unknown',
                    type_name=acc_type.get('name', 'Unknown') if acc_type else 'Unknown',
                    area_name=area.get('name', 'Unknown') if area else 'Unknown',
                    checklist_name=checklist.get('name', 'Unknown') if checklist else 'Unknown',
                    uploader_email=user.get('email', 'Unknown'),
                    uploader_name=f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or user.get('email', 'Unknown')
                )
                
                print(f"[NOTIFICATION DEBUG] Notification created successfully!")
            except Exception as notif_error:
                import traceback
                print(f"[NOTIFICATION ERROR] Failed to create upload notification: {str(notif_error)}")
                print(f"[NOTIFICATION ERROR] Traceback: {traceback.format_exc()}")
            
            uploaded_count += 1
        
        # Handle additional documents with individual names
        additional_files = request.FILES.getlist('additional_documents[]')
        additional_names = request.POST.getlist('additional_document_names[]')
        additional_allowed_formats = ['doc', 'docx', 'pdf', 'ppt', 'pptx', 'xls', 'xlsx', 
                                     'jpg', 'jpeg', 'png', 'gif']
        
        for idx, additional_file in enumerate(additional_files):
            if not additional_file.name:
                continue
            
            add_file_ext = additional_file.name.split('.')[-1].lower()
            
            if add_file_ext not in additional_allowed_formats:
                continue
            
            # Get the name for this additional document
            add_doc_name = additional_names[idx].strip() if idx < len(additional_names) and additional_names[idx].strip() else additional_file.name
            
            # Upload additional document
            add_file_url = upload_document_to_cloudinary(
                additional_file,
                folder=f'documents/{selected_dept_id}/{selected_prog_id}/{selected_type_id}/{selected_area_id}/{selected_checklist_id}/additional'
            )
            
            if add_file_url:
                # Create additional document record
                additional_doc_data = {
                    'department_id': selected_dept_id,
                    'program_id': selected_prog_id,
                    'accreditation_type_id': selected_type_id,
                    'area_id': selected_area_id,
                    'checklist_id': selected_checklist_id,
                    'name': add_doc_name,
                    'file_url': add_file_url,
                    'format': add_file_ext,
                    'uploaded_by': user.get('email', 'Unknown'),
                    'uploaded_at': datetime.now().isoformat(),
                    'is_required': False,
                    'status': 'submitted',
                    'comment': '',  # Comment will be added by disapprover later
                    'is_active': True,
                    'is_archived': False,
                    'user_id': user.get('id'),  # Add user_id for notifications
                }
                
                add_doc_id = str(uuid.uuid4())
                create_document('documents', additional_doc_data, add_doc_id)
                try:
                    log_audit(user, action_type='document_upload', resource_type='document', resource_id=add_doc_id, details=f"Uploaded additional document: {add_doc_name}", status='success', ip=get_client_ip(request))
                except Exception:
                    pass
                
                # Create notification for additional document upload  
                try:
                    from accreditation.notification_utils import notify_document_upload
                    
                    # Get names for notification (reuse from above or fetch)
                    department = get_document('departments', selected_dept_id)
                    program = get_document('programs', selected_prog_id)
                    acc_type = get_document('accreditation_types', selected_type_id)
                    area = get_document('areas', selected_area_id)
                    checklist = get_document('checklists', selected_checklist_id)
                    
                    notify_document_upload(
                        document_id=add_doc_id,
                        document_name=add_doc_name,
                        department_name=department.get('name', 'Unknown') if department else 'Unknown',
                        program_name=program.get('name', 'Unknown') if program else 'Unknown',
                        type_name=acc_type.get('name', 'Unknown') if acc_type else 'Unknown',
                        area_name=area.get('name', 'Unknown') if area else 'Unknown',
                        checklist_name=checklist.get('name', 'Unknown') if checklist else 'Unknown',
                        uploader_email=user.get('email', 'Unknown'),
                        uploader_name=f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or user.get('email', 'Unknown')
                    )
                except Exception as notif_error:
                    print(f"Error creating upload notification: {str(notif_error)}")
        
        total_docs = uploaded_count + len([f for f in additional_files if f.name])
        return JsonResponse({
            'success': True,
            'message': f'{total_docs} document(s) uploaded successfully'
        })
        
    except Exception as e:
        print(f"Error adding documents: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error uploading documents: {str(e)}'
        }, status=500)


@login_required
def document_view(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):
    """Get document details for viewing"""
    try:
        document = get_document('documents', document_id)
        
        if not document:
            return JsonResponse({
                'success': False,
                'message': 'Document not found'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'document': document
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error fetching document: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def document_update_status_view(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):
    """Update document status (approve/disapprove)"""
    user = get_user_from_session(request)
    
    # Only QA Head and QA Admin can update status
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Unauthorized'
        }, status=403)
    
    try:
        import json as json_module
        data = json_module.loads(request.body)
        status = data.get('status')
        comment = data.get('comment', '').strip()
        
        if status not in ['approved', 'disapproved']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status'
            }, status=400)
        
        # Validate comment for disapproval
        if status == 'disapproved' and not comment:
            return JsonResponse({
                'success': False,
                'message': 'Comment is required for disapproval'
            }, status=400)
        
        document = get_document('documents', document_id)
        
        if not document:
            return JsonResponse({
                'success': False,
                'message': 'Document not found'
            }, status=404)
        
        # Update document status
        update_data = {
            'status': status,
            'approved_by': user.get('email', 'Unknown')
        }
        
        # Add comment if provided
        if comment:
            update_data['comment'] = comment
        
        update_document('documents', document_id, update_data)
        try:
            doc_name = document.get('name', 'Unknown document')
            action_text = "approved" if status == "approved" else "disapproved"
            log_audit(user, action_type='update', resource_type='document', resource_id=document_id, details=f"{action_text.capitalize()} document: {doc_name}", status='success')
        except Exception:
            pass
        
        # Create notification for document status change
        try:
            from accreditation.notification_utils import notify_document_status_change
            
            print(f"[STATUS NOTIFICATION] Processing status change for document: {document_id}")
            
            # Get document details for notification
            uploader_id = document.get('user_id')
            
            print(f"[STATUS NOTIFICATION] Uploader ID: {uploader_id}")
            
            if uploader_id:  # Only notify if we have uploader info
                department = get_document('departments', document.get('department_id'))
                program = get_document('programs', document.get('program_id'))
                acc_type = get_document('accreditation_types', document.get('accreditation_type_id'))
                area = get_document('areas', document.get('area_id'))
                checklist = get_document('checklists', document.get('checklist_id'))
                
                notify_document_status_change(
                    document_id=document_id,
                    document_name=document.get('name', 'Unknown'),
                    status=status,
                    uploader_id=uploader_id,
                    department_name=department.get('name', 'Unknown') if department else 'Unknown',
                    program_name=program.get('name', 'Unknown') if program else 'Unknown',
                    type_name=acc_type.get('name', 'Unknown') if acc_type else 'Unknown',
                    area_name=area.get('name', 'Unknown') if area else 'Unknown',
                    checklist_name=checklist.get('name', 'Unknown') if checklist else 'Unknown',
                    reviewer_name=f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or user.get('email', 'Reviewer'),
                    comment=comment
                )
                print(f"[STATUS NOTIFICATION] Notification sent successfully!")
            else:
                print(f"[STATUS NOTIFICATION] No uploader_id found, skipping notification")
        except Exception as notif_error:
            import traceback
            print(f"[STATUS NOTIFICATION ERROR] Failed: {str(notif_error)}")
            print(f"[STATUS NOTIFICATION ERROR] Traceback: {traceback.format_exc()}")
        
        return JsonResponse({
            'success': True,
            'message': f'Document {status} successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating status: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def document_delete_view(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):
    """Delete a document"""
    from accreditation.cloudinary_utils import delete_document_from_cloudinary
    user = get_user_from_session(request)
    
    try:
        document = get_document('documents', document_id)
        
        if not document:
            return JsonResponse({
                'success': False,
                'message': 'Document not found'
            }, status=404)
        
        # Delete file from Cloudinary
        file_url = document.get('file_url', '')
        if file_url:
            delete_document_from_cloudinary(file_url)
        
        # Delete document record
        delete_document('documents', document_id)
        try:
            doc_name = document.get('name', 'Unknown document')
            log_audit(user, action_type='delete', resource_type='document', resource_id=document_id, details=f"Deleted document: {doc_name}", status='success', ip=get_client_ip(request))
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Document deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting document: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def document_proxy_view(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):
    """Proxy view to serve documents from Cloudinary"""
    import requests
    from django.http import HttpResponse, StreamingHttpResponse
    import cloudinary
    import cloudinary.api
    import cloudinary.uploader
    import os
    
    try:
        # Get document details
        document = get_document('documents', document_id)
        
        if not document:
            return HttpResponse('Document not found', status=404)
        
        file_url = document.get('file_url')
        if not file_url:
            return HttpResponse('Document file URL not found', status=404)
        
        # Configure Cloudinary
        api_key = os.environ.get('CLOUDINARY_API_KEY', '')
        api_secret = os.environ.get('CLOUDINARY_API_SECRET', '')
        
        cloudinary.config(
            cloud_name='dygrh6ztt',
            api_key=api_key,
            api_secret=api_secret,
            secure=True
        )
        
        from accreditation.cloudinary_utils import extract_public_id_from_url
        public_id = extract_public_id_from_url(file_url)
        
        if not public_id:
            return HttpResponse('Invalid Cloudinary URL', status=400)
        
        # Use Cloudinary Admin API to get the file directly
        try:
            # Use Cloudinary's API to generate a temporary download URL
            import hashlib
            import time
            
            timestamp = int(time.time())
            
            # Create signature for authenticated access
            to_sign = f"public_id={public_id}&timestamp={timestamp}{api_secret}"
            signature = hashlib.sha1(to_sign.encode('utf-8')).hexdigest()
            
            # Build authenticated download URL
            auth_url = f"https://api.cloudinary.com/v1_1/dygrh6ztt/raw/download?public_id={public_id}&timestamp={timestamp}&signature={signature}&api_key={api_key}"
            
            print(f"Trying authenticated download URL...")
            response = requests.get(auth_url, stream=True, timeout=30)
            print(f"Response status: {response.status_code}")
            
            if response.status_code == 200:
                # Success! Stream the file
                format_lower = (document.get('format') or '').lower()
                
                content_type_map = {
                    'pdf': 'application/pdf',
                    'doc': 'application/msword',
                    'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    'xls': 'application/vnd.ms-excel',
                    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'ppt': 'application/vnd.ms-powerpoint',
                    'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
                }
                
                content_type = content_type_map.get(format_lower, 'application/pdf')
                
                django_response = StreamingHttpResponse(
                    response.iter_content(chunk_size=8192),
                    content_type=content_type
                )
                django_response['Content-Disposition'] = f'inline; filename="{document.get("name", "document.pdf")}"'
                django_response['Cache-Control'] = 'no-cache'
                
                return django_response
            
        except Exception as e:
            print(f"Authenticated download failed: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # If everything fails, return error message
        return HttpResponse(
            '<html><body style="font-family: Arial; padding: 40px; text-align: center;">'
            '<h2>Unable to load PDF</h2>'
            '<p>Please check your Cloudinary settings or contact support.</p>'
            '</body></html>',
            content_type='text/html'
        )
        
    except Exception as e:
        print(f"Error in proxy view: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Error: {str(e)}', status=500)
        return HttpResponse(f'Error loading document: {str(e)}', status=500)


# API endpoints for document modal dropdowns
@login_required
@require_http_methods(["GET"])
def api_get_departments(request):
    """Get all departments for modal dropdown"""
    try:
        departments = get_all_documents('departments')
        active_departments = [
            {'id': dept['id'], 'name': dept.get('name', 'Unknown')}
            for dept in departments 
            if dept.get('is_active', True) and not dept.get('is_archived', False)
        ]
        return JsonResponse({'departments': active_departments})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_department_programs(request, dept_id):
    """Get programs for a specific department"""
    try:
        # Query by department_id
        programs = query_documents('programs', 'department_id', '==', dept_id)
        
        # Filter for active and non-archived
        filtered_programs = [
            prog for prog in programs
            if prog.get('is_active', True) and not prog.get('is_archived', False)
        ]
        
        program_list = [
            {
                'id': prog['id'], 
                'name': prog.get('name', 'Unknown'),
                'code': prog.get('code', '')
            }
            for prog in filtered_programs
        ]
        return JsonResponse({'programs': program_list})
    except Exception as e:
        print(f"Error in api_get_department_programs: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e), 'programs': []}, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_program_types(request, prog_id):
    """Get accreditation types for a specific program"""
    try:
        # Query by program_id
        types = query_documents('accreditation_types', 'program_id', '==', prog_id)
        
        # Filter for active and non-archived
        filtered_types = [
            t for t in types
            if t.get('is_active', True) and not t.get('is_archived', False)
        ]
        
        type_list = [
            {'id': t['id'], 'name': t.get('name', 'Unknown')}
            for t in filtered_types
        ]
        return JsonResponse({'types': type_list})
    except Exception as e:
        print(f"Error in api_get_program_types: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e), 'types': []}, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_type_areas(request, type_id):
    """Get areas for a specific accreditation type"""
    try:
        # Query by accreditation_type_id
        areas = query_documents('areas', 'accreditation_type_id', '==', type_id)
        
        # Filter for active and non-archived
        filtered_areas = [
            area for area in areas
            if area.get('is_active', True) and not area.get('is_archived', False)
        ]
        
        area_list = [
            {'id': area['id'], 'name': area.get('name', 'Unknown')}
            for area in filtered_areas
        ]
        return JsonResponse({'areas': area_list})
    except Exception as e:
        print(f"Error in api_get_type_areas: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e), 'areas': []}, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_area_checklists(request, area_id):
    """Get checklists for a specific area"""
    try:
        # Query by area_id
        checklists = query_documents('checklists', 'area_id', '==', area_id)
        
        # Filter for active and non-archived
        filtered_checklists = [
            cl for cl in checklists
            if cl.get('is_active', True) and not cl.get('is_archived', False)
        ]
        
        checklist_list = [
            {'id': cl['id'], 'name': cl.get('name', 'Unknown')}
            for cl in filtered_checklists
        ]
        return JsonResponse({'checklists': checklist_list})
    except Exception as e:
        print(f"Error in api_get_area_checklists: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e), 'checklists': []}, status=500)


@login_required
@require_http_methods(["GET"])
def download_template_view(request):
    """Download the official document template"""
    import os
    from django.http import FileResponse, Http404
    from django.conf import settings
    
    try:
        # Path to the template file
        template_path = os.path.join(settings.BASE_DIR, 'Template.docx')
        
        if not os.path.exists(template_path):
            raise Http404("Template file not found")
        
        # Open and return the file
        response = FileResponse(
            open(template_path, 'rb'),
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        response['Content-Disposition'] = 'attachment; filename="Template.docx"'
        
        return response
        
    except Exception as e:
        print(f"Error downloading template: {str(e)}")
        raise Http404("Error downloading template file")


# =====================================================
# CALENDAR EVENT VIEWS
# =====================================================

@login_required
def calendar_view(request):
    """Calendar page view"""
    user = get_user_from_session(request)
    return render(request, 'dashboard/calendar.html', {
        'user': user,
        'active_page': 'calendar'
    })


@login_required
def get_calendar_events(request):
    """Get all calendar events"""
    try:
        # Get all events (returns a list of dicts with 'id' field already included)
        events_list = get_all_documents('calendar_events')
        
        # Sort by date
        events_list.sort(key=lambda x: x.get('date', ''))
        
        return JsonResponse({
            'success': True,
            'events': events_list
        })
    except Exception as e:
        print(f"Error getting calendar events: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def create_calendar_event(request):
    """Create a new calendar event"""
    try:
        user = get_user_from_session(request)
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['event_type', 'title', 'date', 'description']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False,
                    'message': f'{field} is required'
                }, status=400)
        
        # Validate event type
        if data['event_type'] not in ['schedules', 'announcements', 'updates']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid event type'
            }, status=400)
        
        # Create event data
        from datetime import datetime
        import uuid
        
        event_data = {
            'event_type': data['event_type'],
            'title': data['title'],
            'date': data['date'],
            'description': data['description'],
            'status': data.get('status', 'active'),
            'is_archived': False,
            'created_by': user.get('email', 'Unknown'),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        # Generate event ID
        event_id = str(uuid.uuid4())
        
        # Save to database
        create_document('calendar_events', event_data, event_id)
        try:
            event_title = data['title']
            event_type = data['event_type']
            event_date = data['date']
            log_audit(user, action_type='create', resource_type='calendar_event', resource_id=event_id, details=f"Created calendar event: {event_title} ({event_type}) on {event_date}", status='success')
        except Exception:
            pass
        
        # Create notification for new event
        try:
            from accreditation.notification_utils import notify_event_created
            
            notify_event_created(
                event_id=event_id,
                event_title=data['title'],
                event_date=data['date'],
                event_description=data['description'],
                created_by=f"{user.get('first_name', '')} {user.get('last_name', '')}".strip() or user.get('email', 'Admin')
            )
        except Exception as notif_error:
            print(f"Error creating event notification: {str(notif_error)}")
        
        return JsonResponse({
            'success': True,
            'message': 'Event created successfully',
            'event_id': event_id
        })
        
    except Exception as e:
        print(f"Error creating calendar event: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def update_calendar_event(request, event_id):
    """Update a calendar event"""
    try:
        user = get_user_from_session(request)
        data = json.loads(request.body)
        
        # Get existing event
        event = get_document('calendar_events', event_id)
        if not event:
            return JsonResponse({
                'success': False,
                'message': 'Event not found'
            }, status=404)
        
        # Update fields
        from datetime import datetime
        
        update_data = {
            'event_type': data.get('event_type', event.get('event_type')),
            'title': data.get('title', event.get('title')),
            'date': data.get('date', event.get('date')),
            'description': data.get('description', event.get('description')),
            'status': data.get('status', event.get('status')),
            'updated_by': user.get('email', 'Unknown'),
            'updated_at': datetime.now().isoformat()
        }
        
        # Update in database
        update_document('calendar_events', event_id, update_data)
        
        return JsonResponse({
            'success': True,
            'message': 'Event updated successfully'
        })
        
    except Exception as e:
        print(f"Error updating calendar event: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def delete_calendar_event(request, event_id):
    """Delete a calendar event"""
    user = get_user_from_session(request)
    
    try:
        # Check if event exists
        event = get_document('calendar_events', event_id)
        if not event:
            return JsonResponse({
                'success': False,
                'message': 'Event not found'
            }, status=404)
        
        # Delete from database
        delete_document('calendar_events', event_id)
        try:
            event_title = event.get('title', 'Unknown Event')
            event_type = event.get('event_type', 'Unknown Type')
            log_audit(user, action_type='delete', resource_type='calendar_event', resource_id=event_id, details=f"Deleted calendar event: {event_title} ({event_type})", status='success')
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Event deleted successfully'
        })
        
    except Exception as e:
        print(f"Error deleting calendar event: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def archive_calendar_event(request, event_id):
    """Archive a calendar event"""
    try:
        user = get_user_from_session(request)
        
        # Check if event exists
        event = get_document('calendar_events', event_id)
        if not event:
            return JsonResponse({
                'success': False,
                'message': 'Event not found'
            }, status=404)
        
        # Update archive status
        from datetime import datetime
        
        update_data = {
            'is_archived': True,
            'archived_by': user.get('email', 'Unknown'),
            'archived_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        update_document('calendar_events', event_id, update_data)
        
        return JsonResponse({
            'success': True,
            'message': 'Event archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving calendar event: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def audit_trail_view(request):
    """Audit Trail page - displays all system audit logs"""
    from datetime import datetime
    
    user = get_user_from_session(request)
    
    # Only QA Head and QA Admin can view audit trail
    if user.get('role') not in ['qa_head', 'qa_admin']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    try:
        # Fetch all audit logs
        audit_logs = get_all_documents('audit_trail')
        
        print(f"DEBUG: Fetched {len(audit_logs)} audit logs from database")
        
        # Sort by timestamp descending (newest first)
        audit_logs.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
        
        # Map old action types to new categorized action types
        action_category_map = {
            'login': 'login_history',
            'logout': 'login_history',
            'document_upload': 'document_upload',
            'report_generation': 'report_generation',
            'report_download': 'report_generation',
            'create': 'record_modification',
            'update': 'record_modification',
            'delete': 'record_modification',
        }
        
        # Readable labels for display
        action_type_labels = {
            'login_history': 'Login History',
            'document_upload': 'Document Upload',
            'report_generation': 'Report Generation',
            'record_modification': 'Record Modification',
        }
        
        # Calculate stats
        today = datetime.now().date().isoformat()
        document_upload_count = 0
        record_modification_count = 0
        today_count = 0
        
        for log in audit_logs:
            # Categorize action type
            original_action = log.get('action_type', '')
            categorized_action = action_category_map.get(original_action, 'record_modification')
            log['action_category'] = categorized_action
            log['action_type_label'] = action_type_labels.get(categorized_action, 'Unknown')
            
            # Count for stats
            if categorized_action == 'document_upload':
                document_upload_count += 1
            elif categorized_action == 'record_modification':
                record_modification_count += 1
            
            # Parse timestamp for better display
            timestamp_str = log.get('timestamp', '')
            if timestamp_str:
                try:
                    dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                    log['timestamp_display'] = dt.strftime('%B %d, %Y at %I:%M %p')
                    # Check if today
                    if dt.date().isoformat() == today:
                        today_count += 1
                except Exception:
                    log['timestamp_display'] = timestamp_str
            else:
                log['timestamp_display'] = 'Unknown'
            
            # Get status (success/failed)
            log['status_value'] = log.get('status', 'success')
            
            # Details is already a string from audit_utils
            if not log.get('details'):
                log['details'] = 'No details provided'
        
    except Exception as e:
        print(f"Error fetching audit logs: {str(e)}")
        import traceback
        traceback.print_exc()
        audit_logs = []
        document_upload_count = 0
        record_modification_count = 0
        today_count = 0
    
    context = {
        'active_page': 'audit',
        'user': user,
        'logs': audit_logs,  # Changed from 'audit_logs' to 'logs' to match template
        'document_upload_count': document_upload_count,
        'record_modification_count': record_modification_count,
        'today_count': today_count,
    }
    
    print(f"DEBUG: Context has {len(audit_logs)} logs")
    print(f"DEBUG: Stats - Today: {today_count}, Docs: {document_upload_count}, Records: {record_modification_count}")
    
    return render(request, 'dashboard/audit.html', context)


# ================================================================================
# DASHBOARD DATA FUNCTIONS - Role-Based Dashboard Content
# ================================================================================

def get_qa_head_dashboard_data(user):
    """Get comprehensive system statistics for QA Head dashboard"""
    from datetime import datetime, timedelta
    import json
    
    try:
        # Get all users
        users = get_all_documents('users')
        total_users = len([u for u in users if not u.get('archived', False)])
        
        # Count users added this month
        now = datetime.now()
        month_start = datetime(now.year, now.month, 1)
        users_this_month = len([
            u for u in users 
            if u.get('created_at') and isinstance(u.get('created_at'), datetime) and u.get('created_at') >= month_start
        ])
        
        # Get all documents
        documents = get_all_documents('documents')
        total_documents = len(documents)
        
        # Count documents this week
        week_ago = now - timedelta(days=7)
        documents_this_week = len([
            d for d in documents
            if d.get('uploaded_at') and isinstance(d.get('uploaded_at'), datetime) and d.get('uploaded_at') >= week_ago
        ])
        
        # Get programs and departments
        programs = get_all_documents('programs')
        total_programs = len([p for p in programs if p.get('status') == 'active'])
        
        departments = get_all_documents('departments')
        active_departments = len([d for d in departments if d.get('active', True) and not d.get('archived', False)])
        
        # Count pending approvals (documents with pending status)
        pending_approvals = len([d for d in documents if d.get('status') == 'pending'])
        
        # Calculate completion rate (documents with status 'approved' vs total required)
        approved_docs = len([d for d in documents if d.get('status') == 'approved'])
        completion_rate = round((approved_docs / total_documents * 100) if total_documents > 0 else 0, 1)
        
        # Get recent activities from audit trail
        audit_logs = get_all_documents('audit_logs')
        audit_logs.sort(key=lambda x: x.get('timestamp', datetime.min), reverse=True)
        recent_activities = audit_logs[:10]
        
        # Format activities
        for activity in recent_activities:
            activity['user_name'] = activity.get('user_name', 'Unknown User')
            if isinstance(activity.get('timestamp'), datetime):
                activity['timestamp'] = activity['timestamp']
        
        # Prepare department chart data
        dept_doc_counts = {}
        for doc in documents:
            dept_name = doc.get('department_name', 'Unknown')
            dept_doc_counts[dept_name] = dept_doc_counts.get(dept_name, 0) + 1
        
        department_labels = json.dumps(list(dept_doc_counts.keys()))
        department_data = json.dumps(list(dept_doc_counts.values()))
        
        # Prepare timeline data (last 30 days)
        timeline_data = []
        timeline_labels = []
        for i in range(29, -1, -1):
            day = now - timedelta(days=i)
            day_str = day.strftime('%m/%d')
            timeline_labels.append(day_str)
            
            count = len([
                d for d in documents
                if d.get('uploaded_at') and isinstance(d.get('uploaded_at'), datetime) and d.get('uploaded_at').date() == day.date()
            ])
            timeline_data.append(count)
        
        # Department progress
        department_progress = []
        for dept in departments[:10]:  # Top 10 departments
            if dept.get('active', True) and not dept.get('archived', False):
                dept_docs = [d for d in documents if d.get('department_id') == dept.get('id')]
                approved = len([d for d in dept_docs if d.get('status') == 'approved'])
                total = len(dept_docs)
                progress = round((approved / total * 100) if total > 0 else 0, 1)
                
                department_progress.append({
                    'name': dept.get('name', 'Unknown'),
                    'progress': progress
                })
        
        return {
            'stats': {
                'total_users': total_users,
                'users_this_month': users_this_month,
                'total_documents': total_documents,
                'documents_this_week': documents_this_week,
                'total_programs': total_programs,
                'pending_approvals': pending_approvals,
                'active_departments': active_departments,
                'completion_rate': completion_rate,
            },
            'recent_activities': recent_activities,
            'department_labels': department_labels,
            'department_data': department_data,
            'timeline_labels': json.dumps(timeline_labels),
            'timeline_data': json.dumps(timeline_data),
            'department_progress': department_progress,
        }
    except Exception as e:
        print(f"Error in get_qa_head_dashboard_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'stats': {},
            'recent_activities': [],
            'department_labels': '[]',
            'department_data': '[]',
            'timeline_labels': '[]',
            'timeline_data': '[]',
            'department_progress': [],
        }


def get_qa_admin_dashboard_data(user):
    """Get accreditation-focused statistics for QA Admin dashboard"""
    from datetime import datetime, timedelta
    import json
    
    def safe_get_datetime(doc, field_name):
        """Safely convert timestamp to datetime object"""
        value = doc.get(field_name)
        if isinstance(value, datetime):
            return value
        elif isinstance(value, str):
            try:
                # Try ISO format first
                return datetime.fromisoformat(value.replace('Z', '+00:00'))
            except:
                try:
                    # Try strptime with common formats
                    return datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f')
                except:
                    return datetime.min
        return datetime.min
    
    try:
        # Get all documents
        documents = get_all_documents('documents')
        total_documents = len(documents)
        
        # Count documents by status - 'submitted' is the pending status
        pending_reviews = len([d for d in documents if d.get('status') == 'submitted'])
        approved_documents = len([d for d in documents if d.get('status') == 'approved'])
        rejected_documents = len([d for d in documents if d.get('status') in ['rejected', 'disapproved']])
        in_review = len([d for d in documents if d.get('status') == 'in_review'])
        
        # Count documents uploaded today
        now = datetime.now()
        today = now.date()
        documents_today = len([
            d for d in documents
            if safe_get_datetime(d, 'uploaded_at').date() == today
        ])
        
        # Get checklists - only active and not archived
        all_checklists = get_all_documents('checklists')
        active_checklists = [c for c in all_checklists if c.get('is_active', False) and not c.get('is_archived', False)]
        total_checklists = len(active_checklists)
        
        # Calculate completion rate based on completed checklists
        # A checklist is complete if it has at least one approved document
        completed_checklists = 0
        for checklist in active_checklists:
            checklist_docs = [d for d in documents if d.get('checklist_id') == checklist.get('id')]
            if checklist_docs:
                # Check if there's at least one approved document for this checklist
                has_approved = any(d.get('status') == 'approved' for d in checklist_docs)
                if has_approved:
                    completed_checklists += 1
        
        completion_rate = round((completed_checklists / total_checklists * 100) if total_checklists > 0 else 0, 1)
        
        # Get departments - use is_active and is_archived fields
        departments = get_all_documents('departments')
        active_departments = len([d for d in departments if d.get('is_active', False) and not d.get('is_archived', False)])
        
        # Get programs - use is_active and is_archived fields
        programs = get_all_documents('programs')
        active_programs = len([p for p in programs if p.get('is_active', False) and not p.get('is_archived', False)])
        
        # Get recent document uploads - only active and not archived
        active_documents = [d for d in documents if d.get('is_active', False) and not d.get('is_archived', False)]
        active_documents.sort(key=lambda x: safe_get_datetime(x, 'uploaded_at'), reverse=True)
        
        # Get all documents for the modal (all active documents)
        all_documents_list = active_documents.copy()
        
        # Get recent documents (limit to 12 for dashboard display)
        recent_documents = active_documents[:12]
        
        # Format all documents - convert timestamps and get related info
        from accreditation.firebase_utils import get_document
        
        # Get lookup dictionaries for efficiency
        users_cache = {}
        departments_cache = {}
        programs_cache = {}
        types_cache = {}
        areas_cache = {}
        checklists_cache = {}
        
        def get_user_name(user_identifier):
            """Get user's full name from email or user ID"""
            if not user_identifier:
                return 'Unknown User'
            if user_identifier not in users_cache:
                try:
                    # First try to get by ID
                    user_doc = safe_get_document('users', user_identifier)
                    
                    # If not found by ID, search by email
                    if not user_doc:
                        all_users = get_all_documents('users')
                        user_doc = next((u for u in all_users if u.get('email') == user_identifier), None)
                    
                    if user_doc:
                        name = f"{user_doc.get('first_name', '')} {user_doc.get('last_name', '')}".strip()
                        users_cache[user_identifier] = name if name else user_doc.get('email', 'Unknown User')
                    else:
                        users_cache[user_identifier] = 'Unknown User'
                except:
                    users_cache[user_identifier] = 'Unknown User'
            return users_cache[user_identifier]
        
        # Format all documents with full details
        for doc in all_documents_list:
            doc['uploaded_at'] = safe_get_datetime(doc, 'uploaded_at')
            doc['uploader_name'] = get_user_name(doc.get('uploaded_by'))
            
            # Get related information for the modal table
            dept_id = doc.get('department_id')
            prog_id = doc.get('program_id')
            type_id = doc.get('accreditation_type_id')
            area_id = doc.get('area_id')
            checklist_id = doc.get('checklist_id')
            
            # Get department name
            if dept_id and dept_id not in departments_cache:
                try:
                    dept = get_document('departments', dept_id)
                    departments_cache[dept_id] = dept.get('name', 'Unknown') if dept else 'Unknown'
                except:
                    departments_cache[dept_id] = 'Unknown'
            doc['department_name'] = departments_cache.get(dept_id, 'Unknown')
            
            # Get program name
            if prog_id and prog_id not in programs_cache:
                try:
                    prog = get_document('programs', prog_id)
                    programs_cache[prog_id] = prog.get('name', 'Unknown') if prog else 'Unknown'
                except:
                    programs_cache[prog_id] = 'Unknown'
            doc['program_name'] = programs_cache.get(prog_id, 'Unknown')
            
            # Get accreditation type name
            if type_id and type_id not in types_cache:
                try:
                    atype = get_document('accreditation_types', type_id)
                    types_cache[type_id] = atype.get('name', 'Unknown') if atype else 'Unknown'
                except:
                    types_cache[type_id] = 'Unknown'
            doc['type_name'] = types_cache.get(type_id, 'Unknown')
            
            # Get area name
            if area_id and area_id not in areas_cache:
                try:
                    area = get_document('areas', area_id)
                    areas_cache[area_id] = area.get('name', 'Unknown') if area else 'Unknown'
                except:
                    areas_cache[area_id] = 'Unknown'
            doc['area_name'] = areas_cache.get(area_id, 'Unknown')
            
            # Get checklist name
            if checklist_id and checklist_id not in checklists_cache:
                try:
                    checklist = get_document('checklists', checklist_id)
                    checklists_cache[checklist_id] = checklist.get('name', 'Unknown') if checklist else 'Unknown'
                except:
                    checklists_cache[checklist_id] = 'Unknown'
            doc['checklist_name'] = checklists_cache.get(checklist_id, 'Unknown')
            
            # Build the URL to the checklist documents page
            if dept_id and prog_id and type_id and area_id and checklist_id:
                doc['checklist_url'] = f"/dashboard/accreditation/department/{dept_id}/programs/{prog_id}/types/{type_id}/areas/{area_id}/checklists/{checklist_id}/documents/"
            else:
                doc['checklist_url'] = "#"
        
        # Prepare department uploads data - count documents by department
        dept_uploads = {}
        for doc in active_documents:
            dept_name = doc.get('department_name', 'Unknown')
            dept_uploads[dept_name] = dept_uploads.get(dept_name, 0) + 1
        
        department_labels = json.dumps(list(dept_uploads.keys()))
        department_uploads = json.dumps(list(dept_uploads.values()))
        
        # Weekly trends (last 7 days)
        weekly_data = []
        weekly_labels = []
        for i in range(6, -1, -1):
            day = now - timedelta(days=i)
            day_str = day.strftime('%a')
            weekly_labels.append(day_str)
            
            count = len([
                d for d in active_documents
                if safe_get_datetime(d, 'uploaded_at').date() == day.date()
            ])
            weekly_data.append(count)
        
        # Get program progress (instead of area progress)
        all_types = get_all_documents('accreditation_types')
        all_areas = get_all_documents('areas')
        all_checklists = get_all_documents('checklists')
        
        program_progress = []
        for prog in programs:
            if not prog.get('is_active', False) or prog.get('is_archived', False):
                continue
                
            prog_id = prog.get('id')
            prog_name = prog.get('name', 'Unknown Program')
            
            # Get all types for this program
            prog_types = [t for t in all_types if t.get('program_id') == prog_id]
            
            if not prog_types:
                program_progress.append({
                    'id': prog_id,
                    'name': prog_name,
                    'progress': 0
                })
                continue
            
            # Calculate progress for each type
            type_progresses = []
            for prog_type in prog_types:
                type_id = prog_type.get('id')
                type_areas = [a for a in all_areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                
                if not type_areas:
                    type_progresses.append(0)
                    continue
                
                # Calculate progress for each area
                area_progresses = []
                for area in type_areas:
                    area_id = area.get('id')
                    area_checklists = [c for c in active_checklists if c.get('area_id') == area_id]
                    
                    if not area_checklists:
                        area_progresses.append(0)
                        continue
                    
                    area_total_checklists = len(area_checklists)
                    completed_checklists = 0
                    
                    for checklist in area_checklists:
                        checklist_id = checklist.get('id')
                        # Check if checklist has at least one approved document
                        checklist_docs = [
                            d for d in active_documents 
                            if d.get('checklist_id') == checklist_id
                        ]
                        has_approved = any(d.get('status') == 'approved' for d in checklist_docs)
                        if has_approved:
                            completed_checklists += 1
                    
                    area_progress = (completed_checklists / area_total_checklists) * 100 if area_total_checklists > 0 else 0
                    area_progresses.append(area_progress)
                
                # Type progress is average of its areas
                type_progress = sum(area_progresses) / len(area_progresses) if area_progresses else 0
                type_progresses.append(type_progress)
            
            # Program progress is the average of its types' progress
            prog_progress = round(sum(type_progresses) / len(type_progresses), 1) if type_progresses else 0
            
            program_progress.append({
                'id': prog_id,
                'name': prog_name,
                'progress': prog_progress
            })
        
        # Get the most recent submitted document's checklist URL for "Review Documents" quick action
        recent_submitted_url = "#"
        submitted_docs = [d for d in active_documents if d.get('status') == 'submitted']
        if submitted_docs:
            # Get the most recent submitted document
            most_recent = submitted_docs[0]  # Already sorted by uploaded_at descending
            dept_id = most_recent.get('department_id')
            prog_id = most_recent.get('program_id')
            type_id = most_recent.get('accreditation_type_id')
            area_id = most_recent.get('area_id')
            checklist_id = most_recent.get('checklist_id')
            
            if dept_id and prog_id and type_id and area_id and checklist_id:
                recent_submitted_url = f"/dashboard/accreditation/department/{dept_id}/programs/{prog_id}/types/{type_id}/areas/{area_id}/checklists/{checklist_id}/documents/"
        
        return {
            'stats': {
                'total_documents': total_documents,
                'documents_today': documents_today,
                'total_checklists': total_checklists,
                'completion_rate': completion_rate,
                'active_departments': active_departments,
                'active_programs': active_programs,
                'pending_reviews': pending_reviews,
                'approved_documents': approved_documents,
                'rejected_documents': rejected_documents,
                'in_review': in_review,
            },
            'recent_documents': recent_documents,
            'all_documents_list': all_documents_list,
            'department_labels': department_labels,
            'department_uploads': department_uploads,
            'weekly_labels': json.dumps(weekly_labels),
            'weekly_data': json.dumps(weekly_data),
            'program_progress': program_progress,
            'recent_submitted_url': recent_submitted_url,
        }
    except Exception as e:
        print(f"Error in get_qa_admin_dashboard_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'stats': {
                'total_documents': 0,
                'documents_today': 0,
                'total_checklists': 0,
                'completion_rate': 0,
                'active_departments': 0,
                'active_programs': 0,
                'pending_reviews': 0,
                'approved_documents': 0,
                'rejected_documents': 0,
                'in_review': 0,
            },
            'recent_documents': [],
            'department_labels': '[]',
            'department_uploads': '[]',
            'weekly_labels': '[]',
            'weekly_data': '[]',
            'area_progress': [],
        }


def get_department_dashboard_data(user):
    """Get department-specific statistics for Department dashboard"""
    from datetime import datetime, timedelta
    import json
    
    def safe_get_datetime(doc, field_name):
        """Safely convert timestamp to datetime object"""
        value = doc.get(field_name)
        if isinstance(value, datetime):
            return value
        elif isinstance(value, str):
            try:
                # Try ISO format first
                return datetime.fromisoformat(value.replace('Z', '+00:00'))
            except:
                try:
                    # Try strptime with common formats
                    return datetime.strptime(value, '%Y-%m-%d %H:%M:%S.%f')
                except:
                    return datetime.min
        return datetime.min
    
    try:
        user_dept_id = user.get('department_id')
        if not user_dept_id:
            return {'stats': {}, 'area_progress': [], 'recent_activities': []}
        
        # Get department's documents
        all_documents = get_all_documents('documents')
        my_documents = [d for d in all_documents if d.get('department_id') == user_dept_id]
        
        # Count by status
        pending_uploads = len([d for d in my_documents if d.get('status') == 'pending'])
        approved = len([d for d in my_documents if d.get('status') == 'approved'])
        needs_revision = len([d for d in my_documents if d.get('status') in ['rejected', 'needs_revision']])
        
        # Calculate completion
        total_docs = len(my_documents)
        completion_percentage = round((approved / total_docs * 100) if total_docs > 0 else 0, 1)
        
        # Get next deadline (placeholder - would need actual deadline system)
        days_to_deadline = 15  # Placeholder
        
        # Get area progress for this department
        areas = get_all_documents('areas')
        area_progress = []
        
        for area in areas:
            # Get checklists for this area
            area_checklists = [
                c for c in get_all_documents('checklists')
                if c.get('area_id') == area.get('id')
            ]
            
            if area_checklists:
                checklist_ids = [c.get('id') for c in area_checklists]
                area_docs = [d for d in my_documents if d.get('checklist_id') in checklist_ids]
                
                required = len(area_checklists)
                uploaded = len(area_docs)
                percentage = round((uploaded / required * 100) if required > 0 else 0, 1)
                
                area_progress.append({
                    'number': area.get('area_number', ''),
                    'name': area.get('name', 'Unknown'),
                    'uploaded': uploaded,
                    'required': required,
                    'percentage': percentage
                })
        
        # Get recent activities - sort safely
        my_documents.sort(key=lambda x: safe_get_datetime(x, 'uploaded_at'), reverse=True)
        recent_activities = []
        
        for doc in my_documents[:10]:
            activity_type = 'upload'
            if doc.get('status') == 'approved':
                activity_type = 'approved'
            elif doc.get('status') in ['rejected', 'needs_revision']:
                activity_type = 'rejected'
            
            recent_activities.append({
                'type': activity_type,
                'title': doc.get('title', 'Untitled Document'),
                'description': f"Area: {doc.get('area_name', 'N/A')}",
                'timestamp': safe_get_datetime(doc, 'uploaded_at')
            })
        
        # Prepare upload progress chart data (last 30 days)
        now = datetime.now()
        upload_data = []
        upload_labels = []
        
        for i in range(29, -1, -1):
            day = now - timedelta(days=i)
            day_str = day.strftime('%m/%d')
            upload_labels.append(day_str)
            
            count = len([
                d for d in my_documents
                if safe_get_datetime(d, 'uploaded_at').date() == day.date()
            ])
            upload_data.append(count)
        
        # Upcoming deadlines (placeholder)
        upcoming_deadlines = []
        
        return {
            'stats': {
                'my_documents': total_docs,
                'pending_uploads': pending_uploads,
                'approved': approved,
                'needs_revision': needs_revision,
                'completion_percentage': completion_percentage,
                'days_to_deadline': days_to_deadline,
            },
            'area_progress': area_progress,
            'recent_activities': recent_activities,
            'upload_labels': json.dumps(upload_labels),
            'upload_data': json.dumps(upload_data),
            'upcoming_deadlines': upcoming_deadlines,
        }
    except Exception as e:
        print(f"Error in get_department_dashboard_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'stats': {},
            'area_progress': [],
            'recent_activities': [],
            'upload_labels': '[]',
            'upload_data': '[]',
            'upcoming_deadlines': [],
        }


# ============================================================================
# MY ACCREDITATION VIEWS - Department User's Personal Accreditation View
# ============================================================================

@login_required
def my_accreditation_view(request):
    """My Accreditation page - displays all active items with user's assigned department"""
    user = get_user_from_session(request)
    
    # Get user's assigned department from session
    user_department_id = request.session.get('user', {}).get('department_id', '')
    
    # Fetch all active departments
    try:
        departments = get_all_documents('departments')
        
        # Filter only active and non-archived departments
        departments = [
            dept for dept in departments 
            if dept.get('is_active', True) and not dept.get('is_archived', False)
        ]
        
        for dept in departments:
            # Set default values if not present
            if 'logo_url' not in dept:
                dept['logo_url'] = ''
        
        # Sort by name
        departments.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching departments: {str(e)}")
        departments = []
    
    # Fetch all active programs
    try:
        programs = get_all_documents('programs')
        
        # Filter only active and non-archived programs
        programs = [
            prog for prog in programs 
            if prog.get('is_active', True) and not prog.get('is_archived', False)
        ]
        
        # Sort by code
        programs.sort(key=lambda x: x.get('code', ''))
        
    except Exception as e:
        print(f"Error fetching programs: {str(e)}")
        programs = []
    
    # Fetch all active accreditation types
    try:
        types = get_all_documents('accreditation_types')
        
        # Filter only active and non-archived types
        types = [
            t for t in types 
            if t.get('is_active', True) and not t.get('is_archived', False)
        ]
        
        # Sort by type name
        types.sort(key=lambda x: x.get('type', ''))
        
    except Exception as e:
        print(f"Error fetching types: {str(e)}")
        types = []
    
    # Fetch all active areas
    try:
        areas = get_all_documents('areas')
        
        # Filter only active and non-archived areas
        areas = [
            mod for mod in areas 
            if mod.get('is_active', True) and not mod.get('is_archived', False)
        ]
        
        # Sort by name
        areas.sort(key=lambda x: x.get('name', ''))
        
    except Exception as e:
        print(f"Error fetching areas: {str(e)}")
        areas = []
    
    # Calculate progress for each department
    try:
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for dept in departments:
            dept_id = dept.get('id')
            # Get all programs for this department
            dept_programs = [p for p in programs if p.get('department_id') == dept_id]
            
            if not dept_programs:
                dept['progress'] = 0
                continue
            
            # Calculate progress for each program first
            program_progresses = []
            for prog in dept_programs:
                prog_id = prog.get('id')
                prog_types = [t for t in types if t.get('program_id') == prog_id]
                
                if not prog_types:
                    program_progresses.append(0)
                    continue
                
                # Calculate progress for each type
                type_progresses = []
                for prog_type in prog_types:
                    type_id = prog_type.get('id')
                    type_areas = [a for a in areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                    
                    if not type_areas:
                        type_progresses.append(0)
                        continue
                    
                    # Calculate progress for each area
                    area_progresses = []
                    for area in type_areas:
                        area_id = area.get('id')
                        area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
                        
                        if not area_checklists:
                            area_progresses.append(0)
                            continue
                        
                        total_checklists = len(area_checklists)
                        completed_checklists = 0
                        
                        for checklist in area_checklists:
                            checklist_id = checklist.get('id')
                            required_docs = [
                                doc for doc in all_documents 
                                if doc.get('checklist_id') == checklist_id 
                                and doc.get('is_required', False)
                                and not doc.get('is_archived', False)
                                and doc.get('status') == 'approved'
                            ]
                            if len(required_docs) > 0:
                                completed_checklists += 1
                        
                        area_progress = (completed_checklists / total_checklists) * 100 if total_checklists > 0 else 0
                        area_progresses.append(area_progress)
                    
                    # Type progress is average of its areas
                    type_progress = sum(area_progresses) / len(area_progresses) if area_progresses else 0
                    type_progresses.append(type_progress)
                
                # Program progress is average of its types
                program_progress = sum(type_progresses) / len(type_progresses) if type_progresses else 0
                program_progresses.append(program_progress)
            
            # Department progress is the average of its programs' progress
            dept['progress'] = round(sum(program_progresses) / len(program_progresses)) if program_progresses else 0
            
    except Exception as e:
        print(f"Error calculating department progress: {str(e)}")
        for dept in departments:
            dept['progress'] = 0
    
    # Get user's department object
    user_department = None
    if user_department_id:
        user_department = next((d for d in departments if d.get('id') == user_department_id), None)
    
    context = {
        'active_page': 'my_accreditation',
        'user': user,
        'departments': departments,
        'programs': programs,
        'types': types,
        'areas': areas,
        'user_department': user_department,
        'user_department_id': user_department_id,
    }
    return render(request, 'dashboard/my_accreditation.html', context)


@login_required
def my_accreditation_department_programs_view(request, dept_id):
    """My Accreditation Programs page - read-only view of programs under a department"""
    user = get_user_from_session(request)
    user_department_id = request.session.get('user', {}).get('department_id', '')
    
    # Get department info
    try:
        department = get_document('departments', dept_id)
        if not department:
            messages.error(request, 'Department not found.')
            return redirect('dashboard:my_accreditation')
    except Exception as e:
        print(f"Error fetching department: {str(e)}")
        messages.error(request, 'Error fetching department.')
        return redirect('dashboard:my_accreditation')
    
    # Get all active programs for this department
    try:
        all_programs = get_all_documents('programs')
        programs = [
            prog for prog in all_programs 
            if prog.get('department_id') == dept_id 
            and prog.get('is_active', True) 
            and not prog.get('is_archived', False)
        ]
        programs.sort(key=lambda x: x.get('code', ''))
        
        # Calculate progress for each program based on its types
        all_types = get_all_documents('accreditation_types')
        all_areas = get_all_documents('areas')
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for prog in programs:
            prog_id = prog.get('id')
            # Get all types for this program
            prog_types = [t for t in all_types if t.get('program_id') == prog_id]
            
            if not prog_types:
                prog['progress'] = 0
                continue
            
            # Calculate progress for each type first
            type_progresses = []
            for prog_type in prog_types:
                type_id = prog_type.get('id')
                type_areas = [a for a in all_areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
                
                if not type_areas:
                    type_progresses.append(0)
                    continue
                
                # Calculate progress for each area
                area_progresses = []
                for area in type_areas:
                    area_id = area.get('id')
                    area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
                    
                    if not area_checklists:
                        area_progresses.append(0)
                        continue
                    
                    total_checklists = len(area_checklists)
                    completed_checklists = 0
                    
                    for checklist in area_checklists:
                        checklist_id = checklist.get('id')
                        required_docs = [
                            doc for doc in all_documents 
                            if doc.get('checklist_id') == checklist_id 
                            and doc.get('is_required', False)
                            and not doc.get('is_archived', False)
                            and doc.get('status') == 'approved'
                        ]
                        if len(required_docs) > 0:
                            completed_checklists += 1
                    
                    area_progress = (completed_checklists / total_checklists) * 100 if total_checklists > 0 else 0
                    area_progresses.append(area_progress)
                
                # Type progress is average of its areas
                type_progress = sum(area_progresses) / len(area_progresses) if area_progresses else 0
                type_progresses.append(type_progress)
            
            # Program progress is the average of its types' progress
            prog['progress'] = round(sum(type_progresses) / len(type_progresses)) if type_progresses else 0
            
    except Exception as e:
        print(f"Error fetching programs: {str(e)}")
        programs = []
    
    # Check if this is user's assigned department
    print(f"DEBUG: dept_id = '{dept_id}', user_department_id = '{user_department_id}'")
    is_user_department = (dept_id == user_department_id)
    print(f"DEBUG: is_user_department = {is_user_department}")
    
    context = {
        'active_page': 'my_accreditation',
        'user': user,
        'department': department,
        'programs': programs,
        'dept_id': dept_id,
        'is_user_department': is_user_department,
        'user_department_id': user_department_id,
    }
    return render(request, 'dashboard/my_accreditation_programs.html', context)


@login_required
def my_accreditation_program_types_view(request, dept_id, prog_id):
    """My Accreditation Types page - read-only view of types under a program"""
    user = get_user_from_session(request)
    user_department_id = request.session.get('user', {}).get('department_id', '')
    
    # Get department and program info
    try:
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        if not department or not program:
            messages.error(request, 'Department or Program not found.')
            return redirect('dashboard:my_accreditation')
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        messages.error(request, 'Error fetching data.')
        return redirect('dashboard:my_accreditation')
    
    # Get all active types for this program
    try:
        all_types = get_all_documents('accreditation_types')
        types = [
            t for t in all_types 
            if t.get('program_id') == prog_id 
            and t.get('is_active', True) 
            and not t.get('is_archived', False)
        ]
        types.sort(key=lambda x: x.get('type', ''))
        
        # Calculate progress for each type based on its areas
        all_areas = get_all_documents('areas')
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for accred_type in types:
            type_id = accred_type.get('id')
            # Get all areas for this type (check both type_id and accreditation_type_id)
            type_areas = [a for a in all_areas if (a.get('type_id') == type_id or a.get('accreditation_type_id') == type_id)]
            
            if not type_areas:
                accred_type['progress'] = 0
                continue
            
            # Calculate progress for each area first
            area_progresses = []
            for area in type_areas:
                area_id = area.get('id')
                area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
                
                if not area_checklists:
                    area_progresses.append(0)
                    continue
                
                total_checklists = len(area_checklists)
                completed_checklists = 0
                
                for checklist in area_checklists:
                    checklist_id = checklist.get('id')
                    required_docs = [
                        doc for doc in all_documents 
                        if doc.get('checklist_id') == checklist_id 
                        and doc.get('is_required', False)
                        and not doc.get('is_archived', False)
                        and doc.get('status') == 'approved'
                    ]
                    if len(required_docs) > 0:
                        completed_checklists += 1
                
                area_progress = (completed_checklists / total_checklists) * 100 if total_checklists > 0 else 0
                area_progresses.append(area_progress)
            
            # Type progress is the average of its areas' progress
            accred_type['progress'] = round(sum(area_progresses) / len(area_progresses)) if area_progresses else 0
            
    except Exception as e:
        print(f"Error fetching types: {str(e)}")
        types = []
    
    # Check if this is user's assigned department
    is_user_department = (dept_id == user_department_id)
    
    context = {
        'active_page': 'my_accreditation',
        'user': user,
        'department': department,
        'program': program,
        'types': types,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'is_user_department': is_user_department,
        'user_department_id': user_department_id,
    }
    return render(request, 'dashboard/my_accreditation_types.html', context)


@login_required
def my_accreditation_type_areas_view(request, dept_id, prog_id, type_id):
    """My Accreditation Areas page - read-only view of areas under a type"""
    user = get_user_from_session(request)
    user_department_id = request.session.get('user', {}).get('department_id', '')
    
    # Get breadcrumb info
    try:
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        accreditation_type = get_document('accreditation_types', type_id)
        if not department or not program or not accreditation_type:
            messages.error(request, 'Data not found.')
            return redirect('dashboard:my_accreditation')
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        messages.error(request, 'Error fetching data.')
        return redirect('dashboard:my_accreditation')
    
    # Get all active areas for this type
    try:
        all_areas = get_all_documents('areas')
        areas = [
            mod for mod in all_areas 
            if (mod.get('type_id') == type_id or mod.get('accreditation_type_id') == type_id)
            and mod.get('is_active', True) 
            and not mod.get('is_archived', False)
        ]
        areas.sort(key=lambda x: x.get('name', ''))
        
        # Calculate progress for each area
        all_checklists = get_all_documents('checklists')
        all_documents = get_all_documents('documents')
        
        for area in areas:
            area_id = area.get('id')
            # Get all checklists for this area
            area_checklists = [c for c in all_checklists if c.get('area_id') == area_id]
            
            if not area_checklists:
                area['progress'] = 0
                continue
            
            # Calculate progress based on required documents
            total_checklists = len(area_checklists)
            completed_checklists = 0
            
            for checklist in area_checklists:
                checklist_id = checklist.get('id')
                required_docs = [
                    doc for doc in all_documents 
                    if doc.get('checklist_id') == checklist_id 
                    and doc.get('is_required', False)
                    and not doc.get('is_archived', False)
                    and doc.get('status') == 'approved'
                ]
                if len(required_docs) > 0:
                    completed_checklists += 1
            
            area['progress'] = round((completed_checklists / total_checklists) * 100) if total_checklists > 0 else 0
            
    except Exception as e:
        print(f"Error fetching areas: {str(e)}")
        areas = []
    
    # Check if this is user's assigned department
    is_user_department = (dept_id == user_department_id)
    
    context = {
        'active_page': 'my_accreditation',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'areas': areas,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'type_id': type_id,
        'is_user_department': is_user_department,
        'user_department_id': user_department_id,
    }
    return render(request, 'dashboard/my_accreditation_areas.html', context)


@login_required
def my_accreditation_area_checklists_view(request, dept_id, prog_id, type_id, area_id):
    """My Accreditation Checklists page - read-only view of checklists under an area"""
    user = get_user_from_session(request)
    user_department_id = request.session.get('user', {}).get('department_id', '')
    
    # Get breadcrumb info
    try:
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        accreditation_type = get_document('accreditation_types', type_id)
        area = get_document('areas', area_id)
        if not department or not program or not accreditation_type or not area:
            messages.error(request, 'Data not found.')
            return redirect('dashboard:my_accreditation')
    except Exception as e:
        print(f"Error fetching data: {str(e)}")
        messages.error(request, 'Error fetching data.')
        return redirect('dashboard:my_accreditation')
    
    # Get all active checklists for this area
    try:
        all_checklists = get_all_documents('checklists')
        checklists = [
            checklist for checklist in all_checklists 
            if checklist.get('area_id') == area_id 
            and checklist.get('is_active', True) 
            and not checklist.get('is_archived', False)
        ]
        checklists.sort(key=lambda x: x.get('name', ''))
        
        # Get all documents to calculate progress for each checklist
        all_documents = get_all_documents('documents')
        
        # Add progress percentage to each checklist
        for checklist in checklists:
            checklist_id = checklist.get('id')
            # Get required documents for this checklist
            required_docs = [
                doc for doc in all_documents 
                if doc.get('checklist_id') == checklist_id 
                and doc.get('is_required', False)
                and not doc.get('is_archived', False)
                and doc.get('status') == 'approved'
            ]
            # Progress is 100% if there's at least 1 required document, otherwise 0%
            checklist['progress'] = 100 if len(required_docs) > 0 else 0
            
    except Exception as e:
        print(f"Error fetching checklists: {str(e)}")
        checklists = []
    
    # Check if this is user's assigned department
    is_user_department = (dept_id == user_department_id)
    
    context = {
        'active_page': 'my_accreditation',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'area': area,
        'checklists': checklists,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'type_id': type_id,
        'area_id': area_id,
        'is_user_department': is_user_department,
        'user_department_id': user_department_id,
    }
    return render(request, 'dashboard/my_accreditation_checklists.html', context)


@login_required
def my_accreditation_checklist_documents_view(request, dept_id, prog_id, type_id, area_id, checklist_id):
    """View documents for a specific checklist - My Accreditation version (no approve/disapprove buttons)"""
    user = get_user_from_session(request)
    user_department_id = request.session.get('user', {}).get('department_id', '')
    
    try:
        # Get breadcrumb data
        department = get_document('departments', dept_id)
        program = get_document('programs', prog_id)
        accreditation_type = get_document('accreditation_types', type_id)
        area = get_document('areas', area_id)
        checklist = get_document('checklists', checklist_id)
        
        if not all([department, program, accreditation_type, area, checklist]):
            messages.error(request, 'Data not found.')
            return redirect('dashboard:my_accreditation')
        
        # Get documents for this checklist
        all_documents = get_all_documents('documents')
        documents = [
            doc for doc in all_documents 
            if doc.get('checklist_id') == checklist_id 
            and not doc.get('is_archived', False)
        ]
        
        # Separate required and additional documents
        required_documents = [doc for doc in documents if doc.get('is_required', False)]
        additional_documents = [doc for doc in documents if not doc.get('is_required', False)]
        
        # Sort required documents by creation date (most recent first)
        required_documents.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        # Sort additional documents by creation date
        additional_documents.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        # Get the most recent required document for backward compatibility
        required_document = required_documents[0] if required_documents else None
        
    except Exception as e:
        print(f"Error fetching documents: {str(e)}")
        messages.error(request, 'Error fetching documents.')
        return redirect('dashboard:my_accreditation')
    
    # Check if this is user's assigned department
    is_user_department = (dept_id == user_department_id)
    
    context = {
        'active_page': 'my_accreditation',
        'user': user,
        'department': department,
        'program': program,
        'accreditation_type': accreditation_type,
        'area': area,
        'checklist': checklist,
        'required_document': required_document,
        'required_documents': required_documents,
        'additional_documents': additional_documents,
        'dept_id': dept_id,
        'prog_id': prog_id,
        'type_id': type_id,
        'area_id': area_id,
        'checklist_id': checklist_id,
        'is_user_department': is_user_department,
        'user_department_id': user_department_id,
    }
    return render(request, 'dashboard/my_accreditation_checklist_documents.html', context)


@login_required
def my_accreditation_view_document(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):
    """View document details - My Accreditation version"""
    try:
        # Get document
        document = get_document('documents', document_id)
        if not document:
            return JsonResponse({'success': False, 'error': 'Document not found'})
        
        # Verify document belongs to this checklist
        if document.get('checklist_id') != checklist_id:
            return JsonResponse({'success': False, 'error': 'Invalid document'})
        
        # Return document data
        return JsonResponse({
            'success': True,
            'document': {
                'id': document.get('id'),
                'name': document.get('name'),
                'format': document.get('format'),
                'file_url': document.get('file_url'),
                'uploaded_by': document.get('uploaded_by'),
                'uploaded_at': document.get('created_at'),
                'status': document.get('status', 'submitted'),
                'is_required': document.get('is_required', False)
            }
        })
        
    except Exception as e:
        print(f"Error viewing document: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An error occurred while loading the document'})


@login_required
@require_http_methods(["GET"])
def my_accreditation_download_document(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):

    """Download document with audit logging - My Accreditation version"""
    from .audit_utils import log_audit
    
    try:
        user = get_user_from_session(request)
        
        # Get document
        document = get_document('documents', document_id)
        if not document:
            return JsonResponse({'success': False, 'error': 'Document not found'})
        
        # Verify document belongs to this checklist
        if document.get('checklist_id') != checklist_id:
            return JsonResponse({'success': False, 'error': 'Invalid document'})
        
        # Only allow download of approved documents
        if document.get('status') != 'approved':
            return JsonResponse({'success': False, 'error': 'Only approved documents can be downloaded'})
        
        # Get document URL
        download_url = document.get('file_url', '')
        if not download_url:
            return JsonResponse({'success': False, 'error': 'Document file not found'})
        
        # Log audit trail
        log_audit(
            user_id=user.get('id'),
            user_email=user.get('email'),
            action='DOWNLOAD_DOCUMENT',
            target_type='document',
            target_id=document_id,
            details={
                'document_name': document.get('name'),
                'document_format': document.get('format'),
                'department_id': dept_id,
                'program_id': prog_id,
                'type_id': type_id,
                'area_id': area_id,
                'checklist_id': checklist_id,
                'page': 'my_accreditation'
            }
        )
        
        return JsonResponse({
            'success': True,
            'download_url': download_url,
            'document_name': document.get('name')
        })
        
    except Exception as e:
        print(f"Error downloading document: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An error occurred while preparing the download'})


@require_http_methods(["GET"])
def accreditation_download_document(request, dept_id, prog_id, type_id, area_id, checklist_id, document_id):
    """Download document - Accreditation (QA Head) version"""
    from .audit_utils import log_audit
    from django.http import FileResponse, HttpResponse
    import requests
    import tempfile
    import os
    import subprocess
    import platform
    
    try:
        user = get_user_from_session(request)
        
        # Only QA Head can download from this view
        if user.get('role') != 'qa_head':
            return JsonResponse({'success': False, 'error': 'Unauthorized access'})
        
        # Get document
        document = get_document('documents', document_id)
        if not document:
            return JsonResponse({'success': False, 'error': 'Document not found'})
        
        # Verify document belongs to this checklist
        if document.get('checklist_id') != checklist_id:
            return JsonResponse({'success': False, 'error': 'Invalid document'})
        
        # Only allow download of approved documents
        if document.get('status') != 'approved':
            return JsonResponse({'success': False, 'error': 'Only approved documents can be downloaded'})
        
        # Get document info
        download_url = document.get('file_url', '')
        if not download_url:
            return JsonResponse({'success': False, 'error': 'Document file not found'})
        
        document_name = document.get('name', 'document')
        document_format = document.get('format', 'docx').lower()
        is_required = document.get('is_required', False)
        
        print(f"Download request - Name: {document_name}, Format: {document_format}, Required: {is_required}")
        
        # Convert to PDF for ALL required documents with DOCX/DOC format
        should_convert = (is_required or is_required is None) and document_format in ['doc', 'docx']
        
        if should_convert:
            print(f"Converting {document_name}.{document_format} to PDF using LibreOffice...")
            try:
                # Download the DOCX file from Cloudinary
                response = requests.get(download_url, timeout=30)
                response.raise_for_status()
                
                # Create temporary files
                temp_docx = tempfile.NamedTemporaryFile(delete=False, suffix='.docx')
                temp_dir = tempfile.mkdtemp()
                
                # Save DOCX content
                temp_docx.write(response.content)
                temp_docx.close()
                
                # Detect LibreOffice path based on OS
                system = platform.system()
                if system == 'Windows':
                    # Common Windows paths
                    libreoffice_paths = [
                        r'C:\Program Files\LibreOffice\program\soffice.exe',
                        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
                    ]
                else:
                    # Linux/Mac paths
                    libreoffice_paths = [
                        '/usr/bin/libreoffice',
                        '/usr/bin/soffice',
                        '/snap/bin/libreoffice',
                    ]
                
                # Find available LibreOffice
                libreoffice_cmd = None
                for path in libreoffice_paths:
                    if os.path.exists(path):
                        libreoffice_cmd = path
                        break
                
                if not libreoffice_cmd:
                    # Try to find in PATH
                    try:
                        result = subprocess.run(['which', 'libreoffice'], capture_output=True, text=True)
                        if result.returncode == 0:
                            libreoffice_cmd = result.stdout.strip()
                    except:
                        pass
                
                if not libreoffice_cmd:
                    print("LibreOffice not found, falling back to original file")
                    raise Exception("LibreOffice not installed")
                
                # Convert DOCX to PDF using LibreOffice headless mode
                # This preserves ALL formatting, images, headers, footers, etc.
                cmd = [
                    libreoffice_cmd,
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', temp_dir,
                    temp_docx.name
                ]
                
                print(f"Running LibreOffice conversion: {' '.join(cmd)}")
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                
                if result.returncode != 0:
                    print(f"LibreOffice conversion failed: {result.stderr}")
                    raise Exception(f"Conversion failed: {result.stderr}")
                
                # Find the generated PDF file
                pdf_filename = os.path.splitext(os.path.basename(temp_docx.name))[0] + '.pdf'
                pdf_path = os.path.join(temp_dir, pdf_filename)
                
                if not os.path.exists(pdf_path):
                    print(f"PDF file not found at {pdf_path}")
                    raise Exception("PDF file not generated")
                
                # Log audit trail
                try:
                    log_audit(
                        user, 
                        action_type='download', 
                        resource_type='document', 
                        resource_id=document_id, 
                        details=f"Downloaded document as PDF: {document_name} from {dept_id}/{prog_id}", 
                        status='success'
                    )
                except Exception as audit_error:
                    print(f"Audit logging failed (non-critical): {audit_error}")
                
                # Read PDF file
                with open(pdf_path, 'rb') as pdf_file:
                    pdf_content = pdf_file.read()
                
                # Clean up temporary files
                os.unlink(temp_docx.name)
                os.unlink(pdf_path)
                os.rmdir(temp_dir)
                
                # Return PDF file
                response = HttpResponse(pdf_content, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{document_name}.pdf"'
                print(f"Successfully converted {document_name} to PDF using LibreOffice, returning file")
                return response
                
            except Exception as convert_error:
                import traceback
                print(f"PDF conversion error: {convert_error}")
                print(f"Traceback: {traceback.format_exc()}")
                # Fall back to original file if conversion fails
                print("PDF conversion failed, downloading original file")
        
        # For non-required or non-DOCX files, return original URL
        # Modify Cloudinary URL to force download
        if 'cloudinary.com' in download_url:
            if '/upload/' in download_url:
                download_url = download_url.replace('/upload/', '/upload/fl_attachment/')
                
                if '?' in download_url:
                    download_url += f'&filename={document_name}.{document_format}'
                else:
                    download_url += f'?filename={document_name}.{document_format}'
        
        # Log audit trail
        try:
            log_audit(
                user, 
                action_type='download', 
                resource_type='document', 
                resource_id=document_id, 
                details=f"Downloaded document: {document_name} from {dept_id}/{prog_id}", 
                status='success'
            )
        except Exception as audit_error:
            print(f"Audit logging failed (non-critical): {audit_error}")
        
        # Return success with download URL
        return JsonResponse({
            'success': True,
            'download_url': download_url,
            'document_name': f"{document_name}.{document_format}",
            'direct_download': True
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"Error downloading document (accreditation_download_document): {error_msg}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Error: {error_msg}'})


@login_required
def contact_us_view(request):
    """Contact Us page (Department Users only)"""
    user = get_user_from_session(request)
    
    # Check if user is Department User
    if user.get('role') != 'department_user':
        messages.error(request, 'Access denied. Contact Us page is only available for Department Users.')
        return redirect('dashboard:home')
    
    context = {
        'active_page': 'contact_us',
        'user': user,
    }
    return render(request, 'dashboard/contact_us.html', context)


@login_required
@require_http_methods(["POST"])
def contact_us_submit(request):
    """Handle contact form submission"""
    from django.core.mail import send_mail
    from django.conf import settings
    
    user = get_user_from_session(request)
    
    # Check if user is Department User
    if user.get('role') != 'department_user':
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        
        # Validation
        if not all([name, email, subject, message]):
            return JsonResponse({
                'success': False,
                'message': 'All fields are required.'
            })
        
        # Create contact message document
        contact_data = {
            'name': name,
            'email': email,
            'subject': subject,
            'message': message,
            'user_id': user.get('id'),
            'user_role': user.get('role'),
            'status': 'unread',
            'created_at': datetime.now().isoformat(),
        }
        
        # Save to Firestore
        doc_id = create_document('contact_messages', contact_data)
        
        # Send email notification to QA Office
        try:
            email_subject = f"[Contact Form] {subject}"
            email_body = f"""
New Contact Form Submission

From: {name}
Email: {email}
Subject: {subject}

Message:
{message}

---
Sent from PLP Accreditation System
Date: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
User Role: {user.get('role', 'N/A')}
            """
            
            # Send email to QA Office
            send_mail(
                subject=email_subject,
                message=email_body,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[settings.EMAIL_HOST_USER],  # Send to QA Office email
                fail_silently=False,
            )
            
            # Send confirmation email to user
            confirmation_subject = f"Your message has been received: {subject}"
            confirmation_body = f"""
Dear {name},

Thank you for contacting the Quality Assurance Office at Pamantasan ng Lungsod ng Pasig.

We have received your message regarding: {subject}

Our team will review your inquiry and get back to you as soon as possible.

Your Message:
{message}

---
This is an automated confirmation email from the PLP Accreditation System.
Please do not reply to this email.

For urgent matters, you may contact us directly at:
Email: qa@plpasig.edu.ph
Phone: (02) 8643-7000

Best regards,
Quality Assurance Office
Pamantasan ng Lungsod ng Pasig
            """
            
            send_mail(
                subject=confirmation_subject,
                message=confirmation_body,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[email],
                fail_silently=True,  # Don't fail if user email doesn't work
            )
            
        except Exception as email_error:
            print(f"Error sending email: {str(email_error)}")
            # Continue even if email fails - message is still saved in DB
        
        # Log audit trail
        log_audit(
            user_id=user.get('id'),
            user_email=user.get('email'),
            action='CONTACT_SUBMIT',
            target_type='contact_message',
            target_id=doc_id,
            details={
                'subject': subject,
                'page': 'contact_us'
            },
            ip_address=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Your message has been sent successfully! We will get back to you soon.'
        })
        
    except Exception as e:
        print(f"Error submitting contact form: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        }, status=500)


# ============================================
# Simple Archive Endpoints for Accreditation Navigation
# ============================================

@require_http_methods(["POST"])
def archive_department_simple(request, dept_code):
    """Simple archive endpoint for departments (used in accreditation navigation)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Find department by code
        depts = get_all_documents('departments')
        dept = next((d for d in depts if d.get('code') == dept_code), None)
        
        if not dept:
            return JsonResponse({
                'success': False,
                'message': 'Department not found'
            }, status=404)
        
        dept_id = dept.get('id')
        dept_name = dept.get('name')
        
        # Archive the department
        update_document('departments', dept_id, {'is_archived': True})
        
        # Log audit trail
        log_audit(
            user,
            action_type='ARCHIVE',
            resource_type='department',
            resource_id=dept_id,
            details=f"Archived department: {dept_name} (Code: {dept_code})",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Department "{dept_name}" archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving department: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error archiving department: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def archive_program_simple(request, prog_code):
    """Simple archive endpoint for programs (used in accreditation navigation)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Find program by code
        progs = get_all_documents('programs')
        prog = next((p for p in progs if p.get('code') == prog_code), None)
        
        if not prog:
            return JsonResponse({
                'success': False,
                'message': 'Program not found'
            }, status=404)
        
        prog_id = prog.get('id')
        prog_name = prog.get('name')
        
        # Archive the program
        update_document('programs', prog_id, {'is_archived': True})
        
        # Log audit trail
        log_audit(
            user,
            action_type='ARCHIVE',
            resource_type='program',
            resource_id=prog_id,
            details=f"Archived program: {prog_name} (Code: {prog_code})",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Program "{prog_name}" archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving program: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error archiving program: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def archive_type_simple(request, type_id):
    """Simple archive endpoint for accreditation types (used in accreditation navigation)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Get accreditation type
        acc_type = get_document('accreditation_types', type_id)
        
        if not acc_type:
            return JsonResponse({
                'success': False,
                'message': 'Accreditation type not found'
            }, status=404)
        
        type_name = acc_type.get('name')
        
        # Archive the type
        update_document('accreditation_types', type_id, {'is_archived': True})
        
        # Log audit trail
        log_audit(
            user,
            action_type='ARCHIVE',
            resource_type='accreditation_type',
            resource_id=type_id,
            details=f"Archived accreditation type: {type_name}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Type "{type_name}" archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving type: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error archiving type: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def archive_area_simple(request, area_id):
    """Simple archive endpoint for areas (used in accreditation navigation)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Get area
        area = get_document('areas', area_id)
        
        if not area:
            return JsonResponse({
                'success': False,
                'message': 'Area not found'
            }, status=404)
        
        area_name = area.get('name')
        
        # Archive the area
        update_document('areas', area_id, {'is_archived': True})
        
        # Log audit trail
        log_audit(
            user,
            action_type='ARCHIVE',
            resource_type='area',
            resource_id=area_id,
            details=f"Archived area: {area_name}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Area "{area_name}" archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving area: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error archiving area: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def archive_checklist_simple(request, checklist_id):
    """Simple archive endpoint for checklists (used in accreditation navigation)"""
    user = get_user_from_session(request)
    
    # Check if user is QA Head or QA Admin
    if user.get('role') not in ['qa_head', 'qa_admin']:
        return JsonResponse({
            'success': False,
            'message': 'Access denied.'
        }, status=403)
    
    try:
        # Get checklist
        checklist = get_document('checklists', checklist_id)
        
        if not checklist:
            return JsonResponse({
                'success': False,
                'message': 'Checklist not found'
            }, status=404)
        
        checklist_name = checklist.get('name')
        
        # Archive the checklist
        update_document('checklists', checklist_id, {'is_archived': True})
        
        # Log audit trail
        log_audit(
            user,
            action_type='ARCHIVE',
            resource_type='checklist',
            resource_id=checklist_id,
            details=f"Archived checklist: {checklist_name}",
            status='success',
            ip=get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Checklist "{checklist_name}" archived successfully'
        })
        
    except Exception as e:
        print(f"Error archiving checklist: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error archiving checklist: {str(e)}'
        }, status=500)


# ============================================
# Department User Pages
# ============================================

@login_required
def dept_home(request):
    """Department User Home Page with welcome message and overlay background"""
    user = get_user_from_session(request)
    
    # Get appearance settings
    appearance_settings = get_document('settings', 'appearance')
    theme_color = appearance_settings.get('theme_color', '#4a9d4f') if appearance_settings else '#4a9d4f'
    logo_url = appearance_settings.get('logo_url', '') if appearance_settings else ''
    system_title = appearance_settings.get('system_title', 'PLP Accreditation System') if appearance_settings else 'PLP Accreditation System'
    
    context = {
        'active_page': 'dept_home',
        'user': user,
        'theme_color': theme_color,
        'logo_url': logo_url,
        'system_title': system_title,
    }
    
    return render(request, 'dashboard/dept_home.html', context)


@login_required
def about(request):
    """About PLP Page"""
    user = get_user_from_session(request)
    
    # Get appearance settings
    appearance_settings = get_document('settings', 'appearance')
    theme_color = appearance_settings.get('theme_color', '#4a9d4f') if appearance_settings else '#4a9d4f'
    logo_url = appearance_settings.get('logo_url', '') if appearance_settings else ''
    system_title = appearance_settings.get('system_title', 'PLP Accreditation System') if appearance_settings else 'PLP Accreditation System'
    
    context = {
        'active_page': 'about',
        'user': user,
        'theme_color': theme_color,
        'logo_url': logo_url,
        'system_title': system_title,
    }
    
    return render(request, 'dashboard/about.html', context)


@login_required
def location(request):
    """Location/Campus Map Page"""
    user = get_user_from_session(request)
    
    # Get appearance settings
    appearance_settings = get_document('settings', 'appearance')
    theme_color = appearance_settings.get('theme_color', '#4a9d4f') if appearance_settings else '#4a9d4f'
    logo_url = appearance_settings.get('logo_url', '') if appearance_settings else ''
    system_title = appearance_settings.get('system_title', 'PLP Accreditation System') if appearance_settings else 'PLP Accreditation System'
    
    context = {
        'active_page': 'location',
        'user': user,
        'theme_color': theme_color,
        'logo_url': logo_url,
        'system_title': system_title,
    }
    
    return render(request, 'dashboard/location.html', context)


@login_required
def mission_vision(request):
    """Mission & Vision Page"""
    user = get_user_from_session(request)
    
    # Get appearance settings
    appearance_settings = get_document('settings', 'appearance')
    theme_color = appearance_settings.get('theme_color', '#4a9d4f') if appearance_settings else '#4a9d4f'
    logo_url = appearance_settings.get('logo_url', '') if appearance_settings else ''
    system_title = appearance_settings.get('system_title', 'PLP Accreditation System') if appearance_settings else 'PLP Accreditation System'
    
    context = {
        'active_page': 'mission_vision',
        'user': user,
        'theme_color': theme_color,
        'logo_url': logo_url,
        'system_title': system_title,
    }
    
    return render(request, 'dashboard/mission_vision.html', context)


# ============================================
# Notification Endpoints
# ============================================

@login_required
def notifications_list_view(request):
    """Get user's notifications"""
    from accreditation.notification_utils import get_user_notifications, get_unread_count
    
    user = get_user_from_session(request)
    user_id = user.get('id')
    
    try:
        notifications = get_user_notifications(user_id)
        unread_count = get_unread_count(user_id)
        
        return JsonResponse({
            'success': True,
            'notifications': notifications,
            'unread_count': unread_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading notifications: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def notification_mark_read_view(request, notification_id):
    """Mark a notification as read"""
    from accreditation.notification_utils import mark_notification_as_read
    
    user = get_user_from_session(request)
    
    try:
        # Verify notification belongs to user
        notification = get_document('notifications', notification_id)
        
        if not notification:
            return JsonResponse({
                'success': False,
                'message': 'Notification not found'
            }, status=404)
        
        if notification.get('user_id') != user.get('id'):
            return JsonResponse({
                'success': False,
                'message': 'Unauthorized'
            }, status=403)
        
        success = mark_notification_as_read(notification_id)
        
        if success:
            return JsonResponse({
                'success': True,
                'message': 'Notification marked as read'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Failed to mark as read'
            }, status=500)
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def notifications_mark_all_read_view(request):
    """Mark all notifications as read for the user"""
    from accreditation.notification_utils import mark_all_as_read
    
    user = get_user_from_session(request)
    user_id = user.get('id')
    
    try:
        count = mark_all_as_read(user_id)
        
        return JsonResponse({
            'success': True,
            'message': f'{count} notifications marked as read',
            'count': count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

